#!/usr/bin/env python3
# =============================================================================
# DISCLAIMER: This code is provided on a best-effort basis and is not in any
# way officially supported or sanctioned by Cohesity. The code is intentionally
# kept simple to retain value as example code. The code in this repository is
# provided as-is and the author accepts no liability for damages resulting
# from its use.
# =============================================================================
"""
alert_summary_report.py
-----------------------
Reports active Cohesity alerts grouped by severity and category across all
Helios-connected clusters (or a single named cluster).

Outputs an Excel workbook with two sheets:
  - Summary  — alert count by cluster × severity × category
  - Alerts   — one row per alert with full detail

API endpoints used:
  Cluster list:  GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  Alerts:        GET  https://helios.cohesity.com/v2/alerts
                      ?alertStateList=kOpen&startDateUsecs=<us>&endDateUsecs=<us>
                      &maxAlerts=1000

Supports Helios API key auth only (multi-cluster feature).

Version history:
  1.0 (2026-04-06) — Initial release. Active alerts by severity and category,
                     dual-sheet Excel output (Summary + Alerts detail),
                     Cohesity green header styling, auto-fit columns.

Usage:
  python3 alert_summary_report.py                        # prompts for key on first run
  python3 alert_summary_report.py --days 7               # last 7 days (default 1)
  python3 alert_summary_report.py --cluster <name>       # one cluster via Helios
  python3 alert_summary_report.py --clear-credentials    # remove stored key
"""

__version__ = "1.0"

import argparse
import getpass
import os
import sys
import urllib3
from collections import defaultdict
from datetime import datetime, timedelta, timezone

import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HELIOS_HOST     = "helios.cohesity.com"
_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
COHESITY_GREEN  = "00B388"


# ---------------------------------------------------------------------------
# Credential helpers
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring  # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def clear_stored_credentials():
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
        keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        print("[*] Stored Helios API key removed from system keychain.")
    else:
        print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Helios helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)
    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])
    result = [{"name": c.get("name", ""), "clusterId": c.get("clusterId")}
              for c in clusters]
    print(f"[+] Connected to Helios — {len(result)} cluster(s)")
    return result


# ---------------------------------------------------------------------------
# Alert fetch
# ---------------------------------------------------------------------------

def usecs_to_str(usecs: int) -> str:
    if not usecs:
        return ""
    return datetime.fromtimestamp(usecs / 1_000_000, tz=timezone.utc).strftime(
        "%Y-%m-%d %H:%M:%S UTC")


def get_alerts(api_key: str, cluster_id: int, start_usecs: int,
               end_usecs: int, debug: bool = False) -> list:
    """Fetch open alerts for a single cluster via Helios routing."""
    url = f"https://{HELIOS_HOST}/irisservices/api/v1/public/alerts"
    params = {
        "alertStateList": "kOpen",
        "startDateUsecs": start_usecs,
        "endDateUsecs":   end_usecs,
        "maxAlerts":      1000,
    }
    if debug:
        print(f"    GET {url}  params={params}")
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         params=params, verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.HTTPError:
        print(f"    WARN: alerts fetch failed ({r.status_code}) — skipping cluster")
        return []
    data = r.json()
    # v1 endpoint returns a direct array
    return data if isinstance(data, list) else data.get("alerts", [])


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def style_ws(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=COHESITY_GREEN)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def auto_fit(ws, min_w=10, max_w=60):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value is not None:
                best = max(best, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(best, max_w)


def write_excel(all_alerts: list, output_path: str):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("ERROR: openpyxl not installed.  pip install openpyxl")
        sys.exit(1)

    wb = Workbook()

    # ---- Summary sheet ----
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = ["Cluster", "Severity", "Category", "Alert Count"]
    ws_sum.append(sum_headers)

    summary: dict = defaultdict(int)
    for a in all_alerts:
        key = (
            a.get("_cluster_name", ""),
            a.get("severity", ""),
            a.get("alertCategory", ""),
        )
        summary[key] += 1

    for (cluster, sev, cat), count in sorted(summary.items()):
        ws_sum.append([cluster, sev, cat, count])

    style_ws(ws_sum)
    auto_fit(ws_sum)

    # ---- Alerts detail sheet ----
    ws_det = wb.create_sheet("Alerts")
    det_headers = [
        "Cluster", "Alert ID", "Alert Code", "Severity", "Category",
        "State", "Alert Name", "Description", "First Occurred", "Last Updated",
        "Node ID", "Acknowledged",
    ]
    ws_det.append(det_headers)

    for a in all_alerts:
        props  = {p.get("key"): p.get("value")
                  for p in a.get("propertyList", [])}
        ws_det.append([
            a.get("_cluster_name", ""),
            a.get("id", ""),
            a.get("alertCode", ""),
            a.get("severity", ""),
            a.get("alertCategory", ""),
            a.get("alertState", ""),
            a.get("alertDocument", {}).get("alertName", ""),
            a.get("alertDocument", {}).get("alertDescription", ""),
            usecs_to_str(a.get("firstTimestampUsecs")),
            usecs_to_str(a.get("latestTimestampUsecs")),
            props.get("nodeId", ""),
            "Yes" if a.get("acknowledged") else "No",
        ])

    style_ws(ws_det)
    auto_fit(ws_det)

    wb.save(output_path)
    print(f"\n[+] Report saved: {output_path}")
    print(f"    Clusters: {len(set(a.get('_cluster_name') for a in all_alerts))}")
    print(f"    Total open alerts: {len(all_alerts)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=f"Cohesity alert summary report v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--apikey",            help="Helios API key (stored in keychain after first use)")
    parser.add_argument("--cluster",           help="Filter to one Helios-connected cluster by name")
    parser.add_argument("--days",   type=int,  default=1,
                        help="Look back N days for open alerts (default: 1)")
    parser.add_argument("--output",            help="Output .xlsx path (auto-generated if omitted)")
    parser.add_argument("--debug",  action="store_true", help="Print API request details")
    parser.add_argument("--clear-credentials", action="store_true",
                        help="Remove stored Helios API key and exit")
    args = parser.parse_args()

    if args.clear_credentials:
        clear_stored_credentials()
        sys.exit(0)

    api_key = get_api_key(args.apikey)
    output  = args.output or default_output_path()

    now_usecs   = int(datetime.now(tz=timezone.utc).timestamp() * 1_000_000)
    start_usecs = int((datetime.now(tz=timezone.utc) - timedelta(days=args.days)).timestamp()
                      * 1_000_000)

    print(f"\n[*] Alert window: last {args.days} day(s)")
    print(f"    From: {usecs_to_str(start_usecs)}")
    print(f"    To:   {usecs_to_str(now_usecs)}")

    clusters = get_helios_clusters(api_key)
    if args.cluster:
        clusters = [c for c in clusters
                    if c["name"].lower() == args.cluster.lower()]
        if not clusters:
            print(f"ERROR: Cluster '{args.cluster}' not found in Helios.")
            sys.exit(1)

    all_alerts = []
    for c in clusters:
        print(f"\n[*] Fetching alerts: {c['name']} (id={c['clusterId']})")
        alerts = get_alerts(api_key, c["clusterId"], start_usecs,
                            now_usecs, debug=args.debug)
        for a in alerts:
            a["_cluster_name"] = c["name"]
        all_alerts.extend(alerts)
        print(f"    {len(alerts)} open alert(s)")

    write_excel(all_alerts, output)


if __name__ == "__main__":
    main()
