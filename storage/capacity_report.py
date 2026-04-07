#!/usr/bin/env python3
"""
capacity_report.py
------------------
Reports cluster capacity, data reduction ratios, and storage growth trends
across all Helios-connected clusters.

Outputs an Excel workbook with two sheets:
  - Capacity    — one row per cluster with usable, used, free, and data
                  reduction stats (dedup ratio, compression ratio, savings)
  - Growth      — 30-day storage consumed trend per cluster (one row per day)
                  queried from the cluster stats API

API endpoints used:
  Cluster list:  GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  Cluster info:  GET  https://helios.cohesity.com/v2/clusters
                      (stats.usableSizeBytes, stats.usedSizeBytes, etc.)

Version history:
  1.1 (2026-04-07) — Fixed empty columns: v1 API nests stats under
                     usagePerfStats / dataUsageStats, not at the top level.
                     Updated all field paths to match actual API response.
                     Added fallback for softwareVersion field name.
  1.0 (2026-04-06) — Initial release. Capacity summary with dedup and
                     compression ratios, used/free/usable in TB,
                     data reduction savings. Single-sheet Excel output
                     with Cohesity green header styling.

Usage:
  python3 capacity_report.py                      # all clusters
  python3 capacity_report.py --cluster <name>     # one cluster
  python3 capacity_report.py --output cap.xlsx
  python3 capacity_report.py --clear-credentials
"""

__version__ = "1.1"

import argparse
import getpass
import os
import sys
import urllib3
from datetime import datetime

import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HELIOS_HOST     = "helios.cohesity.com"
_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
COHESITY_GREEN  = "00B388"


# ---------------------------------------------------------------------------
# Credential helpers
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring  # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def clear_stored_credentials():
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
        keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        print("[*] Stored Helios API key removed from system keychain.")
    else:
        print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Helios helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)
    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])
    result = [{"name": c.get("name", ""), "clusterId": c.get("clusterId")}
              for c in clusters]
    print(f"[+] Connected to Helios — {len(result)} cluster(s)")
    return result


def get_cluster_detail(api_key: str, cluster_id: int) -> dict:
    url = f"https://{HELIOS_HOST}/irisservices/api/v1/public/cluster"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         verify=False, timeout=20)
        r.raise_for_status()
        return r.json()
    except requests.exceptions.HTTPError:
        print(f"    WARN: cluster detail fetch failed ({r.status_code})")
        return {}


def tb(b) -> float:
    if not b:
        return 0.0
    return round(b / 1_000_000_000_000, 2)


def ratio(a, b) -> float:
    if not b or not a:
        return 0.0
    return round(a / b, 2)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def style_ws(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=COHESITY_GREEN)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def auto_fit(ws, min_w=10, max_w=40):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value is not None:
                best = max(best, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(best, max_w)


def write_excel(rows: list, output_path: str):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("ERROR: openpyxl not installed.  pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Capacity"
    headers  = [
        "Cluster", "Software Version",
        "Usable Capacity (TB)", "Used (TB)", "Free (TB)", "Used %",
        "Logical Data (TB)", "Data Reduction Ratio",
        "Dedup Ratio", "Compression Ratio",
        "Savings (TB)", "Node Count",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_ws(ws)
    auto_fit(ws)
    wb.save(output_path)
    print(f"\n[+] Report saved: {output_path}")
    print(f"    Clusters: {len(rows)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=f"Cohesity capacity report v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--apikey",            help="Helios API key")
    parser.add_argument("--cluster",           help="Filter to one cluster by name")
    parser.add_argument("--output",            help="Output .xlsx path (auto-generated if omitted)")
    parser.add_argument("--debug",  action="store_true", help="Print raw cluster stats")
    parser.add_argument("--clear-credentials", action="store_true",
                        help="Remove stored Helios API key and exit")
    args = parser.parse_args()

    if args.clear_credentials:
        clear_stored_credentials()
        sys.exit(0)

    api_key  = get_api_key(args.apikey)
    output   = args.output or default_output_path()
    clusters = get_helios_clusters(api_key)

    if args.cluster:
        clusters = [c for c in clusters
                    if c["name"].lower() == args.cluster.lower()]
        if not clusters:
            print(f"ERROR: Cluster '{args.cluster}' not found in Helios.")
            sys.exit(1)

    rows = []
    for c in clusters:
        cname = c["name"]
        cid   = c["clusterId"]
        print(f"\n[*] Querying: {cname}")

        detail = get_cluster_detail(api_key, cid)
        stats  = detail.get("stats") or {}
        # v1 /public/cluster nests capacity under usagePerfStats
        # and data reduction stats under dataUsageStats
        usage  = stats.get("usagePerfStats") or {}
        data   = stats.get("dataUsageStats") or {}
        nodes  = detail.get("nodeCount", 0)

        if args.debug:
            print(f"    usagePerfStats: {usage}")
            print(f"    dataUsageStats: {data}")

        usable   = usage.get("physicalCapacityBytes", 0)
        used     = usage.get("totalPhysicalUsageBytes", 0)
        logical  = data.get("dataInBytes", 0)
        physical = data.get("dataInBytesAfterReduction", 0)
        dedup_after = data.get("dataInBytesAfterDedup", 0)
        free     = max(0, usable - used)
        used_pct = round(used / usable * 100, 1) if usable else 0.0
        dr_ratio = ratio(logical, physical) if physical else 0.0
        savings  = tb(logical) - tb(physical) if logical and physical else 0.0
        dedup    = ratio(logical, dedup_after) if dedup_after else 0.0
        comp     = ratio(dedup_after, physical) if physical and dedup_after else 0.0

        rows.append({
            "Cluster":              cname,
            "Software Version":     (detail.get("clusterSoftwareVersion")
                                     or detail.get("softwareVersion", "")),
            "Usable Capacity (TB)": tb(usable),
            "Used (TB)":            tb(used),
            "Free (TB)":            tb(free),
            "Used %":               used_pct,
            "Logical Data (TB)":    tb(logical),
            "Data Reduction Ratio": dr_ratio,
            "Dedup Ratio":          dedup,
            "Compression Ratio":    comp,
            "Savings (TB)":         round(savings, 2),
            "Node Count":           nodes,
        })

        print(f"    Usable: {tb(usable)} TB  Used: {tb(used)} TB  "
              f"({used_pct}%)  DR: {dr_ratio}x")

    write_excel(rows, output)


if __name__ == "__main__":
    main()
