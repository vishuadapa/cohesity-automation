#!/usr/bin/env python3
# =============================================================================
# DISCLAIMER: This code is provided on a best-effort basis and is not in any
# way officially supported or sanctioned by Cohesity. The code is intentionally
# kept simple to retain value as example code. The code in this repository is
# provided as-is and the author accepts no liability for damages resulting
# from its use.
# =============================================================================
"""
list_views.py
-------------
Enumerate all Cohesity NAS views (SMB shares / NFS exports) across
Helios-connected clusters with quota and usage information.

Outputs an Excel workbook — one row per view per cluster. Useful for
NAS capacity planning, quota audits, and identifying unconfigured quotas.

API endpoints used:
  Cluster list:  GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  Views:         GET  https://helios.cohesity.com/v2/file-services/views
                      ?maxCount=2000&includeStats=true

Version history:
  1.2 (2026-04-17) — Security: TLS verification enabled by default; added
                     --ca-bundle and --insecure CLI flags. Excel formula
                     injection hardening via _safe_cell().
  1.1 (2026-04-07) — Fixed AttributeError crashes: API returns explicit null
                     for logicalQuota, storagePolicy, stats, smbMountPaths,
                     nfsMountPaths — replaced .get(k, {}) with (... or {})
                     throughout so null values are treated as absent.
  1.0 (2026-04-06) — Initial release. View name, storage domain, protocol
                     access (SMB/NFS/S3), quota limit, logical usage,
                     creation time, and share path.

Usage:
  python3 list_views.py                        # all clusters
  python3 list_views.py --cluster <name>       # one cluster
  python3 list_views.py --protocol nfs         # filter by protocol
  python3 list_views.py --output views.xlsx
  python3 list_views.py --insecure             # skip TLS verification
  python3 list_views.py --ca-bundle /path/to/ca.crt
  python3 list_views.py --clear-credentials
"""

__version__ = "1.2"

import argparse
import getpass
import os
import sys
from datetime import datetime, timezone

import requests

HELIOS_HOST     = "helios.cohesity.com"
_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
COHESITY_GREEN  = "70AD47"

_verify = True   # overridden in main() via --insecure / --ca-bundle

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(val):
    if isinstance(val, str) and val and val[0] in _FORMULA_PREFIXES:
        return "'" + val
    return val


# ---------------------------------------------------------------------------
# Credential helpers
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring  # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def clear_stored_credentials():
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
        keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        print("[*] Stored Helios API key removed from system keychain.")
    else:
        print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_v{__version__}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Helios helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=_verify, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)
    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])
    result = [{"name": c.get("name", ""), "clusterId": c.get("clusterId")}
              for c in clusters]
    print(f"[+] Connected to Helios — {len(result)} cluster(s)")
    return result


def get_views(api_key: str, cluster_id: int) -> list:
    url    = f"https://{HELIOS_HOST}/v2/file-services/views"
    params = {"maxCount": 2000, "includeStats": "true"}
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         params=params, verify=_verify, timeout=20)
        r.raise_for_status()
        return r.json().get("views") or []
    except requests.exceptions.HTTPError:
        print(f"    WARN: views fetch failed ({r.status_code}): {r.text[:200]}")
        return []
    except Exception as e:
        print(f"    WARN: views fetch error: {e}")
        return []


def gb(b) -> float:
    if not b:
        return 0.0
    return round(b / 1_000_000_000, 2)


def usecs_to_date(usecs: int) -> str:
    if not usecs:
        return ""
    return datetime.fromtimestamp(usecs / 1_000_000,
                                  tz=timezone.utc).strftime("%Y-%m-%d")


def protocols(view: dict) -> str:
    parts = []
    if view.get("enableSmbViewDiscovery") or view.get("smbMountPaths"):
        parts.append("SMB")
    nfs = view.get("nfsMountPaths") or view.get("enableNfsKerberos")
    if nfs:
        parts.append("NFS")
    if view.get("s3AccessEnabled"):
        parts.append("S3")
    return ", ".join(parts) if parts else "Unknown"


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def style_ws(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=COHESITY_GREEN)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def auto_fit(ws, min_w=10, max_w=55):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value is not None:
                best = max(best, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(best, max_w)


def write_excel(rows: list, output_path: str):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("ERROR: openpyxl not installed.  pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Views"
    headers  = [
        "Cluster", "View Name", "Storage Domain", "Protocols",
        "Quota Limit (GB)", "Logical Usage (GB)", "Usage %",
        "Case Insensitive", "Created",
        "SMB Paths", "NFS Paths",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([_safe_cell(row.get(h, "")) if isinstance(row.get(h, ""), str) else row.get(h, "") for h in headers])
    style_ws(ws)
    auto_fit(ws)
    wb.save(output_path)
    print(f"\n[+] Report saved: {output_path}")
    print(f"    Views: {len(rows)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=f"Cohesity list views v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--apikey",            help="Helios API key")
    parser.add_argument("--cluster",           help="Filter to one cluster by name")
    parser.add_argument("--protocol",          choices=["smb", "nfs", "s3"],
                        help="Filter views by protocol")
    parser.add_argument("--output",            help="Output .xlsx path (auto-generated if omitted)")
    parser.add_argument("--debug",  action="store_true", help="Print sample raw view object")
    parser.add_argument("--clear-credentials", action="store_true",
                        help="Remove stored Helios API key and exit")
    parser.add_argument("--ca-bundle", dest="ca_bundle", default=None, metavar="PATH",
                        help="Path to CA bundle for TLS verification (e.g. corporate proxy cert)")
    parser.add_argument("--insecure", action="store_true",
                        help="Disable TLS certificate verification (NOT recommended)")
    args = parser.parse_args()

    global _verify
    if args.insecure:
        _verify = False
        print("=" * 65)
        print("  WARN: --insecure — TLS certificate validation DISABLED.")
        print("        Credentials and tokens may be intercepted (MITM).")
        print("=" * 65)
        try:
            import urllib3 as _u3
            _u3.disable_warnings(_u3.exceptions.InsecureRequestWarning)
        except ImportError:
            requests.packages.urllib3.disable_warnings()
    elif args.ca_bundle:
        _verify = args.ca_bundle

    if args.clear_credentials:
        clear_stored_credentials()
        sys.exit(0)

    api_key  = get_api_key(args.apikey)
    output   = args.output or default_output_path()
    clusters = get_helios_clusters(api_key)

    if args.cluster:
        clusters = [c for c in clusters
                    if c["name"].lower() == args.cluster.lower()]
        if not clusters:
            print(f"ERROR: Cluster '{args.cluster}' not found in Helios.")
            sys.exit(1)

    rows = []
    for c in clusters:
        cname = c["name"]
        cid   = c["clusterId"]
        print(f"\n[*] Fetching views: {cname}")
        views = get_views(api_key, cid)

        if args.debug and views:
            print(f"    sample view: {views[0]}")

        for v in views:
            proto = protocols(v)
            if args.protocol and args.protocol.upper() not in proto:
                continue

            quota_bytes = (
                (v.get("logicalQuota") or {}).get("hardLimitBytes")
                or ((v.get("storagePolicy") or {})
                    .get("storageQuotaPolicy") or {}).get("hardLimitBytes")
            )
            usage_bytes = (
                ((v.get("stats") or {}).get("dataUsageStats") or {})
                .get("totalLogicalUsageBytes")
            )
            quota_gb    = gb(quota_bytes)
            usage_gb    = gb(usage_bytes)
            usage_pct   = (round(usage_gb / quota_gb * 100, 1)
                           if quota_gb else "N/A")

            smb_paths = ", ".join(v.get("smbMountPaths") or [])
            nfs_paths = ", ".join(v.get("nfsMountPaths") or [])

            rows.append({
                "Cluster":          cname,
                "View Name":        v.get("name", ""),
                "Storage Domain":   v.get("storageDomainName", ""),
                "Protocols":        proto,
                "Quota Limit (GB)": quota_gb if quota_gb else "",
                "Logical Usage (GB)": usage_gb,
                "Usage %":          usage_pct,
                "Case Insensitive": "Yes" if v.get("caseInsensitiveNamesEnabled") else "No",
                "Created":          usecs_to_date(v.get("createTimeMsecs", 0) * 1000
                                                  if v.get("createTimeMsecs")
                                                  else v.get("createTimeUsecs")),
                "SMB Paths":        smb_paths,
                "NFS Paths":        nfs_paths,
            })

        print(f"    {len(views)} view(s) found")

    write_excel(rows, output)


if __name__ == "__main__":
    main()
