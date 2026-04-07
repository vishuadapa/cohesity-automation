#!/usr/bin/env python3
"""
list_snapshots.py
-----------------
List available recovery points (snapshots) for a protected object.

Given an object name (VM, host, NAS share, etc.), finds its protection runs
and lists all available snapshots with their timestamps, expiry, run type,
status, and storage location (local / replicated / archived).

API endpoints used:
  Cluster list:   GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  Object search:  GET  https://helios.cohesity.com/v2/data-protect/search/protected-objects
                       ?searchString=<name>
  Snapshots:      GET  https://helios.cohesity.com/v2/data-protect/objects/{id}/snapshots

Version history:
  1.1 (2026-04-07) — Removed unused `targets` variable (dead code from
                     indexingStatus field that was never referenced).
  1.0 (2026-04-06) — Initial release. Search by object name, list all
                     snapshots with timestamp, expiry, run type, status,
                     and target locations. Console table + optional Excel.

Usage:
  python3 list_snapshots.py --object "vm-prod-01" --cluster <name>
  python3 list_snapshots.py --object "NAS-Share" --cluster <name> --output snaps.xlsx
  python3 list_snapshots.py --clear-credentials
"""

__version__ = "1.1"

import argparse
import getpass
import os
import sys
import urllib3
from datetime import datetime, timezone

import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HELIOS_HOST     = "helios.cohesity.com"
_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
COHESITY_GREEN  = "00B388"


# ---------------------------------------------------------------------------
# Credential helpers
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring  # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def clear_stored_credentials():
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
        keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        print("[*] Stored Helios API key removed from system keychain.")
    else:
        print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Helios helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)
    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])
    return [{"name": c.get("name", ""), "clusterId": c.get("clusterId")}
            for c in clusters]


def usecs_to_str(usecs: int) -> str:
    if not usecs:
        return ""
    return datetime.fromtimestamp(usecs / 1_000_000, tz=timezone.utc).strftime(
        "%Y-%m-%d %H:%M:%S UTC")


def search_objects(api_key: str, cluster_id: int, name: str) -> list:
    url    = f"https://{HELIOS_HOST}/v2/data-protect/search/protected-objects"
    params = {"searchString": name, "count": 50}
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         params=params, verify=False, timeout=20)
        r.raise_for_status()
        return r.json().get("objects", [])
    except Exception:
        return []


def get_snapshots(api_key: str, cluster_id: int, object_id: int) -> list:
    url = f"https://{HELIOS_HOST}/v2/data-protect/objects/{object_id}/snapshots"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         verify=False, timeout=20)
        r.raise_for_status()
        return r.json().get("snapshots", [])
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(rows: list, output_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("NOTE: openpyxl not installed — skipping Excel output.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshots"
    headers  = [
        "Object Name", "Object Type", "Protection Group", "Cluster",
        "Snapshot Time (UTC)", "Run Type", "Status", "Expiry (UTC)",
        "Has Local", "Has Replication", "Has Archival", "Snapshot ID",
    ]
    ws.append(headers)

    fill = PatternFill("solid", fgColor=COHESITY_GREEN)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for col in ws.columns:
        best = 10
        for cell in col:
            if cell.value:
                best = max(best, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(best, 50)

    wb.save(output_path)
    print(f"\n[+] Report saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=f"Cohesity list snapshots v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--apikey",            help="Helios API key")
    parser.add_argument("--cluster", required=True,
                        help="Cluster name (as shown in Helios)")
    parser.add_argument("--object",  required=True,
                        help="Object name to search for (partial match)")
    parser.add_argument("--output",            help="Save results to .xlsx (optional)")
    parser.add_argument("--debug",  action="store_true", help="Print raw API responses")
    parser.add_argument("--clear-credentials", action="store_true",
                        help="Remove stored Helios API key and exit")
    args = parser.parse_args()

    if args.clear_credentials:
        clear_stored_credentials()
        sys.exit(0)

    api_key  = get_api_key(args.apikey)
    clusters = get_helios_clusters(api_key)
    match_c  = [c for c in clusters
                if c["name"].lower() == args.cluster.lower()]
    if not match_c:
        print(f"ERROR: Cluster '{args.cluster}' not found in Helios.")
        sys.exit(1)
    cluster_id = match_c[0]["clusterId"]

    print(f"\n[*] Searching for '{args.object}' on {args.cluster}…")
    objects = search_objects(api_key, cluster_id, args.object)

    if not objects:
        print(f"[*] No protected objects found matching '{args.object}'.")
        sys.exit(0)

    print(f"[+] Found {len(objects)} object(s):\n")
    rows = []

    for obj in objects:
        oname = obj.get("name", "")
        oid   = obj.get("id")
        otype = obj.get("objectType", "")
        pg    = obj.get("protectionGroupName", "")

        print(f"  Object: {oname}  [{otype}]  Group: {pg}")

        snaps = get_snapshots(api_key, cluster_id, oid)
        if args.debug:
            print(f"    raw snapshots[0]: {snaps[:1]}")

        if not snaps:
            print(f"    No snapshots found.\n")
            continue

        print(f"  {'Snapshot Time (UTC)':<25} {'Run Type':<12} {'Status':<15} "
              f"{'Expiry':<25} {'Local':<7} {'Repl':<7} {'Archive':<8}")
        print(f"  {'─'*100}")

        for s in snaps:
            snap_time = usecs_to_str(s.get("snapshotTimestampUsecs"))
            expiry    = usecs_to_str(s.get("expiryTimeUsecs"))
            run_type  = s.get("runType", "")
            status    = s.get("runStatus", "")

            has_local = "Yes" if s.get("hasLocalSnapshot") else "No"
            has_repl  = "Yes" if s.get("hasReplicationSnapshot") else "No"
            has_arch  = "Yes" if s.get("hasArchivalSnapshot") else "No"

            print(f"  {snap_time:<25} {run_type:<12} {status:<15} "
                  f"{expiry:<25} {has_local:<7} {has_repl:<7} {has_arch:<8}")

            rows.append({
                "Object Name":        oname,
                "Object Type":        otype,
                "Protection Group":   pg,
                "Cluster":            args.cluster,
                "Snapshot Time (UTC)": snap_time,
                "Run Type":           run_type,
                "Status":             status,
                "Expiry (UTC)":       expiry,
                "Has Local":          has_local,
                "Has Replication":    has_repl,
                "Has Archival":       has_arch,
                "Snapshot ID":        s.get("id", ""),
            })

        print()

    if args.output and rows:
        write_excel(rows, args.output)
    elif rows and len(rows) > 20:
        out = default_output_path()
        write_excel(rows, out)


if __name__ == "__main__":
    main()
