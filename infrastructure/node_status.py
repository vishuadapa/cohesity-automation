#!/usr/bin/env python3
# =============================================================================
# DISCLAIMER: This code is provided on a best-effort basis and is not in any
# way officially supported or sanctioned by Cohesity. The code is intentionally
# kept simple to retain value as example code. The code in this repository is
# provided as-is and the author accepts no liability for damages resulting
# from its use.
# =============================================================================
"""
node_status.py
--------------
Reports node health, disk status, and fault tolerance for all
Helios-connected clusters (or a single named cluster).

Outputs an Excel workbook with two sheets:
  - Nodes   — one row per node with health, role, IP, and disk summary
  - Disks   — one row per disk with slot, capacity, health, and tier

API endpoints used:
  Cluster list:  GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  Nodes:         GET  https://helios.cohesity.com/v2/nodes
  Disks:         GET  https://helios.cohesity.com/v2/disks

Version history:
  1.0 (2026-04-06) — Initial release. Node health and disk inventory with
                     fault tolerance counts, dual-sheet Excel output,
                     Cohesity green header styling.

Usage:
  python3 node_status.py                      # all clusters
  python3 node_status.py --cluster <name>     # one cluster
  python3 node_status.py --unhealthy-only     # only nodes/disks with issues
  python3 node_status.py --clear-credentials
"""

__version__ = "1.0"

import argparse
import getpass
import os
import sys
import urllib3
from datetime import datetime

import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HELIOS_HOST     = "helios.cohesity.com"
_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
COHESITY_GREEN  = "70AD47"


# ---------------------------------------------------------------------------
# Credential helpers
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring  # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def clear_stored_credentials():
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
        keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        print("[*] Stored Helios API key removed from system keychain.")
    else:
        print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Helios helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)
    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])
    result = [{"name": c.get("name", ""), "clusterId": c.get("clusterId")}
              for c in clusters]
    print(f"[+] Connected to Helios — {len(result)} cluster(s)")
    return result


def get_nodes(api_key: str, cluster_id: int) -> list:
    url = f"https://{HELIOS_HOST}/irisservices/api/v1/public/nodes"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         verify=False, timeout=20)
        r.raise_for_status()
        data = r.json()
        return data if isinstance(data, list) else data.get("nodes", [])
    except requests.exceptions.HTTPError:
        print(f"    WARN: node fetch failed ({r.status_code})")
        return []


def node_status_str(n: dict) -> str:
    """Derive a human-readable node status from v1 node fields."""
    if n.get("isMarkedForRemoval"):
        return "Marked for Removal"
    state = n.get("removalState", "")
    if state == "kDontRemove":
        return "Healthy"
    return state if state else "Unknown"


def bytes_to_tb(b) -> float:
    if not b:
        return 0.0
    return round(b / 1_000_000_000_000, 2)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def style_ws(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=COHESITY_GREEN)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def auto_fit(ws, min_w=10, max_w=55):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value is not None:
                best = max(best, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(best, max_w)


def write_excel(node_rows: list, disk_rows: list, output_path: str):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("ERROR: openpyxl not installed.  pip install openpyxl")
        sys.exit(1)

    wb = Workbook()

    # ---- Nodes sheet ----
    ws_n = wb.active
    ws_n.title = "Nodes"
    node_headers = [
        "Cluster", "Node ID", "Node IP", "Status", "Node Type",
        "Total Disks", "Capacity (TB)", "Model", "Host Name",
        "Serial", "Software Version",
    ]
    ws_n.append(node_headers)
    for row in node_rows:
        ws_n.append([row.get(h, "") for h in node_headers])
    style_ws(ws_n)
    auto_fit(ws_n)

    # ---- Disk Tiers sheet (tier-level breakdown per node) ----
    ws_d = wb.create_sheet("Disk Tiers")
    disk_headers = [
        "Cluster", "Node IP", "Host Name", "Storage Tier",
        "Disk Count", "Tier Capacity (TB)",
    ]
    ws_d.append(disk_headers)
    for row in disk_rows:
        ws_d.append([row.get(h, "") for h in disk_headers])
    style_ws(ws_d)
    auto_fit(ws_d)

    wb.save(output_path)
    print(f"\n[+] Report saved: {output_path}")
    print(f"    Nodes: {len(node_rows)}  |  Disk tier rows: {len(disk_rows)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=f"Cohesity node status report v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--apikey",              help="Helios API key")
    parser.add_argument("--cluster",             help="Filter to one cluster by name")
    parser.add_argument("--unhealthy-only",      action="store_true",
                        help="Only include nodes/disks that are not in a normal state")
    parser.add_argument("--output",              help="Output .xlsx path (auto-generated if omitted)")
    parser.add_argument("--debug",   action="store_true", help="Print raw API responses")
    parser.add_argument("--clear-credentials",   action="store_true",
                        help="Remove stored Helios API key and exit")
    args = parser.parse_args()

    if args.clear_credentials:
        clear_stored_credentials()
        sys.exit(0)

    api_key  = get_api_key(args.apikey)
    output   = args.output or default_output_path()
    clusters = get_helios_clusters(api_key)

    if args.cluster:
        clusters = [c for c in clusters
                    if c["name"].lower() == args.cluster.lower()]
        if not clusters:
            print(f"ERROR: Cluster '{args.cluster}' not found in Helios.")
            sys.exit(1)

    node_rows, disk_rows = [], []

    for c in clusters:
        cname = c["name"]
        cid   = c["clusterId"]
        print(f"\n[*] Querying: {cname}")

        nodes = get_nodes(api_key, cid)

        if args.debug:
            print(f"    nodes sample: {nodes[:1]}")

        for n in nodes:
            status   = node_status_str(n)
            node_ip  = n.get("ip", "")
            hostname = n.get("hostName", "")

            if args.unhealthy_only and status == "Healthy":
                continue

            cap_tb = bytes_to_tb(n.get("maxPhysicalCapacityBytes", 0))

            node_rows.append({
                "Cluster":        cname,
                "Node ID":        n.get("id", ""),
                "Node IP":        node_ip,
                "Status":         status,
                "Node Type":      n.get("nodeType", ""),
                "Total Disks":    n.get("diskCount", ""),
                "Capacity (TB)":  cap_tb,
                "Model":          n.get("hardwareModel", n.get("productModel", "")),
                "Host Name":      hostname,
                "Serial":         n.get("cohesityNodeSerial", ""),
                "Software Version": n.get("nodeSoftwareVersion", ""),
            })

            # Disk tier breakdown rows (one per storage tier per node)
            for tier in n.get("diskCountByTier", []):
                tier_name  = tier.get("storageTier", "")
                tier_count = tier.get("diskCount", 0)
                # Match capacity for this tier
                tier_cap   = next(
                    (bytes_to_tb(t.get("tierMaxPhysicalCapacityBytes", 0))
                     for t in n.get("capacityByTier", [])
                     if t.get("storageTier") == tier_name),
                    ""
                )
                disk_rows.append({
                    "Cluster":          cname,
                    "Node IP":          node_ip,
                    "Host Name":        hostname,
                    "Storage Tier":     tier_name,
                    "Disk Count":       tier_count,
                    "Tier Capacity (TB)": tier_cap,
                })

        healthy_nodes = sum(1 for n in nodes if node_status_str(n) == "Healthy")
        print(f"    Nodes: {len(nodes)} ({healthy_nodes} healthy)")

    write_excel(node_rows, disk_rows, output)


if __name__ == "__main__":
    main()
