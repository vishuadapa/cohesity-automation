#!/usr/bin/env python3
"""
cluster_info_report.py
----------------------
Reports software version, node count, cluster ID, and key configuration
across all Helios-connected clusters (or a single named cluster).

Outputs an Excel workbook with one row per cluster. Useful as a quick
inventory snapshot before upgrades, for presales assessments, or for
multi-cluster environment audits.

API endpoints used:
  Cluster list:  GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  Cluster info:  GET  https://helios.cohesity.com/v2/clusters
  Node list:     GET  https://helios.cohesity.com/v2/nodes

Version history:
  1.0 (2026-04-06) — Initial release. Software version, node count, cluster
                     ID, incarnation ID, cluster type, domain name, DNS
                     servers, NTP servers. Excel output with Cohesity green
                     header styling and auto-fit columns.

Usage:
  python3 cluster_info_report.py                     # all clusters via Helios
  python3 cluster_info_report.py --cluster <name>    # one cluster
  python3 cluster_info_report.py --output report.xlsx
  python3 cluster_info_report.py --clear-credentials
"""

__version__ = "1.0"

import argparse
import getpass
import os
import sys
import urllib3
from datetime import datetime, timezone

import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HELIOS_HOST     = "helios.cohesity.com"
_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
COHESITY_GREEN  = "00B388"


# ---------------------------------------------------------------------------
# Credential helpers
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring  # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def clear_stored_credentials():
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
        keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        print("[*] Stored Helios API key removed from system keychain.")
    else:
        print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Helios helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)
    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])
    result = [{"name": c.get("name", ""), "clusterId": c.get("clusterId"),
               "heliosConnectedAt": c.get("connectedToHelios")}
              for c in clusters]
    print(f"[+] Connected to Helios — {len(result)} cluster(s)")
    return result


def get_cluster_detail(api_key: str, cluster_id: int) -> dict:
    """GET /irisservices/api/v1/public/cluster — returns cluster config including version."""
    url = f"https://{HELIOS_HOST}/irisservices/api/v1/public/cluster"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         verify=False, timeout=20)
        r.raise_for_status()
        return r.json()
    except requests.exceptions.HTTPError:
        print(f"    WARN: cluster detail fetch failed ({r.status_code})")
        return {}


def get_node_list(api_key: str, cluster_id: int) -> list:
    """GET /irisservices/api/v1/public/nodes — returns list of nodes."""
    url = f"https://{HELIOS_HOST}/irisservices/api/v1/public/nodes"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         verify=False, timeout=20)
        r.raise_for_status()
        data = r.json()
        return data if isinstance(data, list) else data.get("nodes", [])
    except requests.exceptions.HTTPError:
        print(f"    WARN: node list fetch failed ({r.status_code})")
        return []


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def style_ws(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=COHESITY_GREEN)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def auto_fit(ws, min_w=10, max_w=60):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value is not None:
                best = max(best, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(best, max_w)


def write_excel(rows: list, output_path: str):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("ERROR: openpyxl not installed.  pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cluster Info"

    headers = [
        "Cluster Name", "Cluster ID", "Incarnation ID", "Software Version",
        "Cluster Type", "Node Count", "Healthy Nodes", "Domain Name",
        "DNS Servers", "NTP Servers", "Timezone", "Encryption Enabled",
        "FIPS Enabled", "Helios Connected",
    ]
    ws.append(headers)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    style_ws(ws)
    auto_fit(ws)
    wb.save(output_path)
    print(f"\n[+] Report saved: {output_path}")
    print(f"    Clusters: {len(rows)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=f"Cohesity cluster info report v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--apikey",            help="Helios API key (stored in keychain after first use)")
    parser.add_argument("--cluster",           help="Filter to one Helios-connected cluster by name")
    parser.add_argument("--output",            help="Output .xlsx path (auto-generated if omitted)")
    parser.add_argument("--debug",  action="store_true", help="Print raw API responses")
    parser.add_argument("--clear-credentials", action="store_true",
                        help="Remove stored Helios API key and exit")
    args = parser.parse_args()

    if args.clear_credentials:
        clear_stored_credentials()
        sys.exit(0)

    api_key  = get_api_key(args.apikey)
    output   = args.output or default_output_path()
    clusters = get_helios_clusters(api_key)

    if args.cluster:
        clusters = [c for c in clusters
                    if c["name"].lower() == args.cluster.lower()]
        if not clusters:
            print(f"ERROR: Cluster '{args.cluster}' not found in Helios.")
            sys.exit(1)

    rows = []
    for c in clusters:
        print(f"\n[*] Querying: {c['name']} (id={c['clusterId']})")
        detail = get_cluster_detail(api_key, c["clusterId"])
        nodes  = get_node_list(api_key, c["clusterId"])

        if args.debug:
            print(f"    cluster detail: {detail}")

        healthy = sum(1 for n in nodes
                      if n.get("nodeStatus", {}).get("overallStatus") == "kNormal")

        dns = ", ".join(detail.get("dnsServerIps", []))
        ntp = ", ".join(detail.get("ntpSettings", {}).get("ntpServers", []))

        rows.append({
            "Cluster Name":       c["name"],
            "Cluster ID":         detail.get("id", c["clusterId"]),
            "Incarnation ID":     detail.get("incarnationId", ""),
            "Software Version":   detail.get("clusterSoftwareVersion", ""),
            "Cluster Type":       detail.get("clusterType", ""),
            "Node Count":         len(nodes),
            "Healthy Nodes":      healthy,
            "Domain Name":        detail.get("domainNames", [""])[0]
                                  if detail.get("domainNames") else "",
            "DNS Servers":        dns,
            "NTP Servers":        ntp,
            "Timezone":           detail.get("timezone", ""),
            "Encryption Enabled": "Yes" if detail.get("encryptionEnabled") else "No",
            "FIPS Enabled":       "Yes" if detail.get("fipsEnabled") else "No",
            "Helios Connected":   "Yes" if c.get("heliosConnectedAt") else "No",
        })

        print(f"    Version: {detail.get('clusterSoftwareVersion', 'N/A')}  "
              f"Nodes: {len(nodes)} ({healthy} healthy)")

    write_excel(rows, output)


if __name__ == "__main__":
    main()
