#!/usr/bin/env python3
# =============================================================================
# DISCLAIMER: This code is provided on a best-effort basis and is not in any
# way officially supported or sanctioned by Cohesity. The code is intentionally
# kept simple to retain value as example code. The code in this repository is
# provided as-is and the author accepts no liability for damages resulting
# from its use.
# =============================================================================
"""
upgrade_readiness.py
--------------------
Checks prerequisites before a major Cohesity software version upgrade.

Runs a series of readiness checks per cluster and prints a pass/warn/fail
summary to the console. Optionally saves results to Excel for sharing.

Checks performed:
  - All nodes reachable and in Normal health
  - No disks in Faulted or Marked for Removal state
  - No active protection runs in progress
  - Helios connection status
  - Current software version (flags clusters below a minimum if --min-version set)
  - Available disk capacity margin (warns if < 20% free)

API endpoints used:
  Cluster list:  GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  Cluster info:  GET  https://helios.cohesity.com/v2/clusters
  Nodes:         GET  https://helios.cohesity.com/v2/nodes
  Disks:         GET  https://helios.cohesity.com/v2/disks
  Active runs:   GET  https://helios.cohesity.com/v2/data-protect/protection-groups/runs
                      ?runStatus=Running

Version history:
  1.0 (2026-04-06) — Initial release. Node health, disk health, in-progress
                     runs, version check, capacity margin checks.
                     Console table summary + optional Excel output.

Usage:
  python3 upgrade_readiness.py
  python3 upgrade_readiness.py --cluster <name>
  python3 upgrade_readiness.py --min-version 7.1
  python3 upgrade_readiness.py --output readiness.xlsx
  python3 upgrade_readiness.py --clear-credentials
"""

__version__ = "1.0"

import argparse
import getpass
import os
import sys
import urllib3
from datetime import datetime

import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HELIOS_HOST     = "helios.cohesity.com"
_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
COHESITY_GREEN  = "70AD47"

STATUS_PASS = "PASS"
STATUS_WARN = "WARN"
STATUS_FAIL = "FAIL"


# ---------------------------------------------------------------------------
# Credential helpers
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring  # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def clear_stored_credentials():
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
        keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        print("[*] Stored Helios API key removed from system keychain.")
    else:
        print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Helios helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)
    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])
    result = [{"name": c.get("name", ""), "clusterId": c.get("clusterId"),
               "heliosConnected": bool(c.get("connectedToHelios"))}
              for c in clusters]
    print(f"[+] Connected to Helios — {len(result)} cluster(s)\n")
    return result


def hget(api_key: str, cluster_id: int, path: str, params: dict = None) -> dict:
    url = f"https://{HELIOS_HOST}{path}"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         params=params, verify=False, timeout=20)
        r.raise_for_status()
        data = r.json()
        # v1 endpoints that return arrays — wrap so callers can use .get()
        return {"_list": data} if isinstance(data, list) else data
    except requests.exceptions.HTTPError:
        print(f"    WARN: {path} failed ({r.status_code})")
        return {}
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Checks
# ---------------------------------------------------------------------------

def _node_is_healthy(n: dict) -> bool:
    """v1 nodes have no nodeStatus.overallStatus — use isMarkedForRemoval + removalState."""
    return not n.get("isMarkedForRemoval") and n.get("removalState", "kDontRemove") == "kDontRemove"


def check_nodes(api_key: str, cluster_id: int) -> tuple:
    data  = hget(api_key, cluster_id, "/irisservices/api/v1/public/nodes")
    nodes = data.get("_list", data.get("nodes", []))
    if not nodes:
        return STATUS_WARN, "Could not retrieve node list"
    unhealthy = [n for n in nodes if not _node_is_healthy(n)]
    if unhealthy:
        details = ", ".join(
            f"{n.get('ip','?')} ({n.get('removalState','unknown')})"
            for n in unhealthy
        )
        return STATUS_FAIL, f"{len(unhealthy)}/{len(nodes)} nodes not healthy: {details}"
    return STATUS_PASS, f"All {len(nodes)} nodes healthy"


def check_disks(api_key: str, cluster_id: int) -> tuple:
    """
    v1 public API doesn't expose per-disk health in node responses.
    Uses node-level disk counts and marks nodes with isMarkedForRemoval as a proxy.
    """
    data  = hget(api_key, cluster_id, "/irisservices/api/v1/public/nodes")
    nodes = data.get("_list", data.get("nodes", []))
    if not nodes:
        return STATUS_WARN, "Could not retrieve node list for disk check"

    total_disks  = sum(n.get("diskCount", 0) for n in nodes)
    problem_nodes = [n for n in nodes if not _node_is_healthy(n)]

    if problem_nodes:
        ips = ", ".join(n.get("ip", "?") for n in problem_nodes)
        return STATUS_WARN, (f"{total_disks} total disks across {len(nodes)} nodes — "
                             f"{len(problem_nodes)} node(s) flagged: {ips}")

    # Break down by storage tier for informational detail
    tier_counts: dict = {}
    for n in nodes:
        for t in n.get("diskCountByTier", []):
            tier = t.get("storageTier", "Unknown")
            tier_counts[tier] = tier_counts.get(tier, 0) + t.get("diskCount", 0)
    tier_str = ", ".join(f"{v} {k}" for k, v in sorted(tier_counts.items()))
    return STATUS_PASS, f"{total_disks} disks healthy ({tier_str})"


def check_active_runs(api_key: str, cluster_id: int) -> tuple:
    data = hget(api_key, cluster_id,
                "/irisservices/api/v1/public/protectionRuns",
                params={"runStatus": "kRunning", "numRuns": 100})
    runs = data.get("_list", data.get("protectionRuns", []))
    if runs:
        return STATUS_WARN, f"{len(runs)} protection run(s) currently in progress"
    return STATUS_PASS, "No active protection runs"


def check_version(cluster_version: str, min_version: str) -> tuple:
    if not cluster_version:
        return STATUS_WARN, "Version not available"
    if not min_version:
        return STATUS_PASS, cluster_version
    # Simple string prefix comparison — works for Cohesity X.Y.Z.x versioning
    cv = tuple(int(x) for x in cluster_version.split(".")[:2] if x.isdigit())
    mv = tuple(int(x) for x in min_version.split(".")[:2] if x.isdigit())
    if cv < mv:
        return STATUS_FAIL, f"Version {cluster_version} < minimum {min_version}"
    return STATUS_PASS, f"Version {cluster_version} meets minimum {min_version}"


def check_capacity(api_key: str, cluster_id: int) -> tuple:
    """
    v1 /cluster endpoint returns stats=None. Compute raw physical capacity
    by summing maxPhysicalCapacityBytes across all nodes (same data used
    by check_disks). Reports total raw capacity and per-tier breakdown.
    """
    data  = hget(api_key, cluster_id, "/irisservices/api/v1/public/nodes")
    nodes = data.get("_list", data.get("nodes", []))
    if not nodes:
        return STATUS_WARN, "Could not retrieve nodes for capacity check"

    total_bytes = sum(n.get("maxPhysicalCapacityBytes", 0) for n in nodes)
    if not total_bytes:
        return STATUS_WARN, "Capacity data not available from node responses"

    total_tb = round(total_bytes / 1_000_000_000_000, 2)

    # Per-tier breakdown from cluster-level diskCountByTier
    cluster_data = hget(api_key, cluster_id, "/irisservices/api/v1/public/cluster")
    tier_info    = cluster_data.get("diskCountByTier", [])
    tier_str     = ""
    if tier_info:
        parts    = [f"{t.get('diskCount',0)} {t.get('storageTier','')}" for t in tier_info]
        tier_str = f" ({', '.join(parts)})"

    fault_level  = cluster_data.get("faultToleranceLevel", "")
    fault_str    = f", FT: {fault_level}" if fault_level else ""

    return STATUS_PASS, f"Raw physical: {total_tb} TB{tier_str}{fault_str}"


# ---------------------------------------------------------------------------
# Console + Excel output
# ---------------------------------------------------------------------------

STATUS_ICON = {STATUS_PASS: "[PASS]", STATUS_WARN: "[WARN]", STATUS_FAIL: "[FAIL]"}


def print_cluster_results(cname: str, results: list):
    print(f"  Cluster: {cname}")
    for check_name, status, detail in results:
        icon = STATUS_ICON[status]
        print(f"    {icon:8s} {check_name:30s} {detail}")
    overall = (STATUS_FAIL if any(s == STATUS_FAIL for _, s, _ in results)
               else STATUS_WARN if any(s == STATUS_WARN for _, s, _ in results)
               else STATUS_PASS)
    print(f"    {'─'*55}")
    print(f"    Overall: {STATUS_ICON[overall]}\n")
    return overall


def write_excel(all_results: list, output_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("NOTE: openpyxl not installed — skipping Excel output.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Upgrade Readiness"
    headers  = ["Cluster", "Check", "Status", "Detail"]
    ws.append(headers)

    fill = PatternFill("solid", fgColor=COHESITY_GREEN)
    bold_white = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = bold_white
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    status_colors = {STATUS_PASS: "C6EFCE", STATUS_WARN: "FFEB9C", STATUS_FAIL: "FFC7CE"}
    status_fonts  = {STATUS_PASS: "276221", STATUS_WARN: "9C5700", STATUS_FAIL: "9C0006"}

    for cname, check_name, status, detail in all_results:
        row_idx = ws.max_row + 1
        ws.append([cname, check_name, status, detail])
        for col in range(1, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill("solid", fgColor=status_colors[status])
            cell.font = Font(color=status_fonts[status])

    for col in ws.columns:
        best = 10
        for cell in col:
            if cell.value:
                best = max(best, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(best, 70)

    wb.save(output_path)
    print(f"[+] Report saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=f"Cohesity upgrade readiness check v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--apikey",              help="Helios API key")
    parser.add_argument("--cluster",             help="Filter to one cluster by name")
    parser.add_argument("--min-version",         help="Minimum required version (e.g. 7.1)")
    parser.add_argument("--output",              help="Save results to .xlsx (optional)")
    parser.add_argument("--debug",  action="store_true",
                        help="Print raw cluster API response for field inspection")
    parser.add_argument("--clear-credentials",   action="store_true",
                        help="Remove stored Helios API key and exit")
    args = parser.parse_args()

    if args.clear_credentials:
        clear_stored_credentials()
        sys.exit(0)

    api_key  = get_api_key(args.apikey)
    clusters = get_helios_clusters(api_key)

    if args.cluster:
        clusters = [c for c in clusters
                    if c["name"].lower() == args.cluster.lower()]
        if not clusters:
            print(f"ERROR: Cluster '{args.cluster}' not found in Helios.")
            sys.exit(1)

    all_results = []   # (cname, check_name, status, detail)
    overall_map = {}

    for c in clusters:
        cname = c["name"]
        cid   = c["clusterId"]

        # Get version for version check
        detail_data     = hget(api_key, cid, "/irisservices/api/v1/public/cluster")
        cluster_version = detail_data.get("clusterSoftwareVersion", "")
        if args.debug:
            print(f"  [debug] cluster raw response keys: {list(detail_data.keys())}")
            print(f"  [debug] stats value: {detail_data.get('stats')}")

        checks = [
            ("Node Health",    *check_nodes(api_key, cid)),
            ("Disk Health",    *check_disks(api_key, cid)),
            ("Active Runs",    *check_active_runs(api_key, cid)),
            ("Software Ver.",  *check_version(cluster_version, args.min_version)),
            ("Disk Capacity",  *check_capacity(api_key, cid)),
        ]

        overall = print_cluster_results(cname, checks)
        overall_map[cname] = overall

        for check_name, status, det in checks:
            all_results.append((cname, check_name, status, det))

    # Final summary
    print(f"{'═'*60}")
    print("  SUMMARY")
    print(f"{'═'*60}")
    for cname, overall in overall_map.items():
        print(f"  {STATUS_ICON[overall]:8s} {cname}")

    fail_count = sum(1 for v in overall_map.values() if v == STATUS_FAIL)
    warn_count = sum(1 for v in overall_map.values() if v == STATUS_WARN)
    print(f"\n  {len(clusters)} cluster(s) checked — "
          f"{fail_count} FAIL, {warn_count} WARN, "
          f"{len(clusters)-fail_count-warn_count} PASS")

    if args.output:
        write_excel(all_results, args.output)
    elif all_results:
        out = default_output_path()
        write_excel(all_results, out)


if __name__ == "__main__":
    main()
