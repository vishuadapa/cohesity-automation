#!/usr/bin/env python3
# =============================================================================
# DISCLAIMER: This code is provided on a best-effort basis and is not in any
# way officially supported or sanctioned by Cohesity. The code is intentionally
# kept simple to retain value as example code. The code in this repository is
# provided as-is and the author accepts no liability for damages resulting
# from its use.
# =============================================================================
"""
formatters.py
-------------
Shared formatting helpers for Cohesity scripts.

Provides:
  - usecs_to_datetime(usecs, tz)  — microsecond epoch → local datetime string
  - bytes_to_gb(b)                — bytes → GB (4 decimal places)
  - bytes_to_tb(b)                — bytes → TB (4 decimal places)
  - format_duration(start, end)   — usec timestamps → "Xh Ym Zs"
  - style_worksheet(ws)           — apply Cohesity green header style to openpyxl sheet
  - auto_fit_columns(ws)          — auto-fit column widths based on content

Typical import pattern:
  from utils.formatters import usecs_to_datetime, bytes_to_gb, style_worksheet

Version history:
  1.0 (2026-04-06) — Initial module. Extracted repeated formatting helpers
                     from reporting scripts into a shared module.
"""

__version__ = "1.0"

from datetime import datetime, timezone

# Cohesity brand green — used for header rows in all reports
COHESITY_GREEN = "00B388"


def usecs_to_datetime(usecs: int, tz=None) -> str:
    """
    Convert a microsecond epoch timestamp to a human-readable local datetime.

    Args:
        usecs: Microsecond epoch timestamp (int). Returns "" if falsy.
        tz:    ZoneInfo timezone object. If None, uses UTC.

    Returns:
        "YYYY-MM-DD HH:MM:SS" in the given timezone, or "" if usecs is 0/None.
    """
    if not usecs:
        return ""
    dt = datetime.fromtimestamp(usecs / 1_000_000, tz=timezone.utc)
    if tz is not None:
        dt = dt.astimezone(tz)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def bytes_to_gb(b) -> float:
    """Convert bytes to decimal GB (÷ 1,000,000,000). Returns 0.0 for None/0."""
    if not b:
        return 0.0
    return round(b / 1_000_000_000, 4)


def bytes_to_tb(b) -> float:
    """Convert bytes to decimal TB (÷ 1,000,000,000,000). Returns 0.0 for None/0."""
    if not b:
        return 0.0
    return round(b / 1_000_000_000_000, 4)


def format_duration(start_usecs: int, end_usecs: int) -> str:
    """
    Format the elapsed time between two microsecond timestamps as "Xh Ym Zs".
    Returns "" if either value is missing.
    """
    if not start_usecs or not end_usecs:
        return ""
    secs = max(0, (end_usecs - start_usecs) // 1_000_000)
    h, rem = divmod(secs, 3600)
    m, s   = divmod(rem, 60)
    if h:
        return f"{h}h {m}m {s}s"
    if m:
        return f"{m}m {s}s"
    return f"{s}s"


def style_worksheet(ws):
    """
    Apply Cohesity green (#00B388) header styling to an openpyxl worksheet:
      - Bold white text on Cohesity green background for row 1
      - Freeze pane below header row
    Requires openpyxl to be installed.
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return  # openpyxl not installed — skip styling silently

    header_fill = PatternFill("solid", fgColor=COHESITY_GREEN)
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"


def auto_fit_columns(ws, min_width: int = 6, max_width: int = 60):
    """
    Auto-fit column widths to content, set zoom to 130%, and left-align data cells.

    Column width = max content length + 2 padding chars, clamped to [min_width, max_width].
    Handles merged cells gracefully and multiline values (measures longest line).
    """
    try:
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils import get_column_letter as _gcl
        from openpyxl.styles import Alignment
    except ImportError:
        return

    # 130% zoom for comfortable viewing without manual adjustment
    try:
        ws.sheet_view.zoomScale = 130
    except Exception:
        pass

    # Autofit column widths based on content
    for col in ws.columns:
        if not col:
            continue
        first = col[0]
        col_letter = (first.column_letter
                      if hasattr(first, "column_letter")
                      else _gcl(first.column))
        best = min_width
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                lines    = str(cell.value).split("\n")
                cell_len = max(len(line) for line in lines) + 2
                best     = max(best, cell_len)
        ws.column_dimensions[col_letter].width = min(best, max_width)

    # Left-align data cells — only cells without explicit horizontal alignment
    _al = Alignment(horizontal="left", vertical="center")
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            if not isinstance(cell, MergedCell) and cell.alignment.horizontal is None:
                cell.alignment = _al
