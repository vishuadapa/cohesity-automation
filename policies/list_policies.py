#!/usr/bin/env python3
"""
list_policies.py
----------------
Lists all Cohesity protection policies across Helios-connected clusters
with retention, replication, and archival settings.

Outputs an Excel workbook — one row per policy per cluster. Useful for
auditing policy configurations across a multi-cluster environment before
policy consolidation or SLA reviews.

API endpoints used:
  Cluster list:  GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  Policies:      GET  https://helios.cohesity.com/v2/data-protect/policies

Version history:
  1.0 (2026-04-06) — Initial release. Retention, full/incremental schedules,
                     replication targets and retention, archival targets and
                     retention, policy type, and associated group count.
                     Excel output with Cohesity green header styling.

Usage:
  python3 list_policies.py                       # all clusters
  python3 list_policies.py --cluster <name>      # one cluster
  python3 list_policies.py --output policies.xlsx
  python3 list_policies.py --clear-credentials
"""

__version__ = "1.0"

import argparse
import getpass
import os
import sys
import urllib3
from datetime import datetime

import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HELIOS_HOST     = "helios.cohesity.com"
_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
COHESITY_GREEN  = "00B388"


# ---------------------------------------------------------------------------
# Credential helpers
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring  # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def clear_stored_credentials():
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
        keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        print("[*] Stored Helios API key removed from system keychain.")
    else:
        print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Helios helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)
    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])
    result = [{"name": c.get("name", ""), "clusterId": c.get("clusterId")}
              for c in clusters]
    print(f"[+] Connected to Helios — {len(result)} cluster(s)")
    return result


def get_policies(api_key: str, cluster_id: int) -> list:
    url = f"https://{HELIOS_HOST}/v2/data-protect/policies"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         verify=False, timeout=20)
        r.raise_for_status()
        return r.json().get("policies", [])
    except Exception:
        return []


def schedule_str(sched: dict) -> str:
    """Return a short human-readable schedule string (e.g. 'Daily', 'Weekly/Mon')."""
    if not sched:
        return ""
    unit = sched.get("unit", "")
    freq = sched.get("frequency", 1)
    if unit == "Days":
        return f"Every {freq}d" if freq > 1 else "Daily"
    if unit == "Weeks":
        days = sched.get("dayOfWeek", [])
        return f"Weekly/{','.join(days)}" if days else "Weekly"
    if unit == "Hours":
        return f"Every {freq}h"
    return unit


def retention_str(r: dict) -> str:
    if not r:
        return ""
    unit  = r.get("unit", "")
    dur   = r.get("duration", "")
    return f"{dur} {unit}" if dur and unit else ""


def replication_summary(targets: list) -> str:
    if not targets:
        return ""
    parts = []
    for t in targets:
        cname = t.get("targetName", t.get("remoteTargetConfig", {}).get("clusterName", "?"))
        ret   = retention_str(t.get("retention", {}))
        parts.append(f"{cname} ({ret})" if ret else cname)
    return "; ".join(parts)


def archival_summary(targets: list) -> str:
    if not targets:
        return ""
    parts = []
    for t in targets:
        tname = t.get("targetName", "?")
        ttype = t.get("targetType", "")
        ret   = retention_str(t.get("retention", {}))
        label = f"{tname}/{ttype} ({ret})" if ret else f"{tname}/{ttype}"
        parts.append(label)
    return "; ".join(parts)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def style_ws(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=COHESITY_GREEN)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def auto_fit(ws, min_w=10, max_w=55):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value is not None:
                best = max(best, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(best, max_w)


def write_excel(rows: list, output_path: str):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("ERROR: openpyxl not installed.  pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Policies"
    headers  = [
        "Cluster", "Policy Name", "Policy ID", "Policy Type",
        "Full Backup Schedule", "Incremental Schedule",
        "Local Retention", "Log Retention",
        "Replication Targets", "Archival Targets",
        "Is Active", "Groups Using Policy",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_ws(ws)
    auto_fit(ws)
    wb.save(output_path)
    print(f"\n[+] Report saved: {output_path}")
    print(f"    Policies: {len(rows)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=f"Cohesity list policies v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--apikey",            help="Helios API key")
    parser.add_argument("--cluster",           help="Filter to one cluster by name")
    parser.add_argument("--output",            help="Output .xlsx path (auto-generated if omitted)")
    parser.add_argument("--debug",  action="store_true", help="Print raw policy objects")
    parser.add_argument("--clear-credentials", action="store_true",
                        help="Remove stored Helios API key and exit")
    args = parser.parse_args()

    if args.clear_credentials:
        clear_stored_credentials()
        sys.exit(0)

    api_key  = get_api_key(args.apikey)
    output   = args.output or default_output_path()
    clusters = get_helios_clusters(api_key)

    if args.cluster:
        clusters = [c for c in clusters
                    if c["name"].lower() == args.cluster.lower()]
        if not clusters:
            print(f"ERROR: Cluster '{args.cluster}' not found in Helios.")
            sys.exit(1)

    rows = []
    for c in clusters:
        cname = c["name"]
        cid   = c["clusterId"]
        print(f"\n[*] Fetching policies: {cname}")
        policies = get_policies(api_key, cid)
        if args.debug:
            print(f"    sample: {policies[:1]}")

        for p in policies:
            backup = p.get("backupPolicy", {})
            reg    = backup.get("regular", {})
            full_s = schedule_str(reg.get("fullBackups", {}).get("schedule", {}))
            incr_s = schedule_str(reg.get("incremental", {}).get("schedule", {}))

            rows.append({
                "Cluster":             cname,
                "Policy Name":         p.get("name", ""),
                "Policy ID":           p.get("id", ""),
                "Policy Type":         p.get("policyCategory", ""),
                "Full Backup Schedule": full_s,
                "Incremental Schedule": incr_s,
                "Local Retention":     retention_str(
                    reg.get("retention", {})),
                "Log Retention":       retention_str(
                    backup.get("log", {}).get("retention", {})),
                "Replication Targets": replication_summary(
                    p.get("remoteTargetPolicy", {}).get("replicationTargets", [])),
                "Archival Targets":    archival_summary(
                    p.get("remoteTargetPolicy", {}).get("archivalTargets", [])),
                "Is Active":           "Yes" if not p.get("isDeleted") else "No",
                "Groups Using Policy": p.get("numProtectionGroups", ""),
            })

        print(f"    {len(policies)} policy/ies")

    write_excel(rows, output)


if __name__ == "__main__":
    main()
