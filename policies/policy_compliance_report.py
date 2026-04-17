#!/usr/bin/env python3
# =============================================================================
# DISCLAIMER: This code is provided on a best-effort basis and is not in any
# way officially supported or sanctioned by Cohesity. The code is intentionally
# kept simple to retain value as example code. The code in this repository is
# provided as-is and the author accepts no liability for damages resulting
# from its use.
# =============================================================================
"""
policy_compliance_report.py
----------------------------
Flags protection groups whose policy doesn't meet configurable SLA targets.

For each protection group across Helios-connected clusters, checks:
  - Local retention meets minimum days
  - A replication target exists (if required)
  - An archival/FortKnox target exists (if required)
  - Last run completed within the RPO window (hours since last successful run)

Groups that fail any check are marked FAIL; those with warnings are WARN;
all others are PASS. Output is an Excel workbook with color-coded rows.

API endpoints used:
  Cluster list:   GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  Policies:       GET  https://helios.cohesity.com/v2/data-protect/policies
  Prot. groups:   GET  https://helios.cohesity.com/v2/data-protect/protection-groups
                       ?isActive=true&includeTenants=true
  Last run:       GET  https://helios.cohesity.com/v2/data-protect/protection-groups/{id}/runs
                       ?numRuns=1&includeObjectDetails=false

Version history:
  1.1 (2026-04-17) — Security: TLS verification enabled by default; added
                     --ca-bundle and --insecure CLI flags. Excel formula
                     injection hardening via _safe_cell() on all API data.
  1.0 (2026-04-06) — Initial release. Retention, replication, archival, and
                     RPO checks per group. Color-coded Excel output (green/
                     yellow/red). Configurable SLA thresholds via CLI flags.

Usage:
  python3 policy_compliance_report.py
  python3 policy_compliance_report.py --min-retention 14 --require-replication
  python3 policy_compliance_report.py --require-archival --rpo-hours 24
  python3 policy_compliance_report.py --cluster <name>
  python3 policy_compliance_report.py --clear-credentials
"""

__version__ = "1.1"

import argparse
import getpass
import os
import sys
from datetime import datetime, timezone

import requests

HELIOS_HOST     = "helios.cohesity.com"
_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
COHESITY_GREEN  = "70AD47"

_verify = True   # overridden in main() via --insecure / --ca-bundle

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(val):
    if isinstance(val, str) and val and val[0] in _FORMULA_PREFIXES:
        return "'" + val
    return val

STATUS_PASS = "PASS"
STATUS_WARN = "WARN"
STATUS_FAIL = "FAIL"


# ---------------------------------------------------------------------------
# Credential helpers
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring  # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def clear_stored_credentials():
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
        keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        print("[*] Stored Helios API key removed from system keychain.")
    else:
        print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_v{__version__}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Helios helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=_verify, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)
    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])
    result = [{"name": c.get("name", ""), "clusterId": c.get("clusterId")}
              for c in clusters]
    print(f"[+] Connected to Helios — {len(result)} cluster(s)")
    return result


def hget(api_key: str, cluster_id: int, path: str, params: dict = None) -> dict:
    url = f"https://{HELIOS_HOST}{path}"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         params=params, verify=_verify, timeout=20)
        r.raise_for_status()
        return r.json()
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Policy helpers
# ---------------------------------------------------------------------------

def retention_days(ret: dict) -> int:
    """Convert a retention object to total days."""
    if not ret:
        return 0
    dur  = ret.get("duration", 0)
    unit = ret.get("unit", "Days")
    multipliers = {"Days": 1, "Weeks": 7, "Months": 30, "Years": 365}
    return dur * multipliers.get(unit, 1)


def has_replication(policy: dict) -> bool:
    targets = (policy.get("remoteTargetPolicy", {})
                     .get("replicationTargets", []))
    return len(targets) > 0


def has_archival(policy: dict) -> bool:
    targets = (policy.get("remoteTargetPolicy", {})
                     .get("archivalTargets", []))
    return len(targets) > 0


# ---------------------------------------------------------------------------
# Compliance evaluation
# ---------------------------------------------------------------------------

def evaluate_group(group: dict, policy: dict, last_run_end_usecs: int,
                   min_retention: int, require_replication: bool,
                   require_archival: bool, rpo_hours: int) -> tuple:
    """
    Returns (status, list_of_issues).
    """
    issues = []
    status = STATUS_PASS

    # Retention check
    if min_retention:
        backup   = policy.get("backupPolicy", {})
        reg      = backup.get("regular", {})
        ret_days = retention_days(reg.get("retention", {}))
        if ret_days < min_retention:
            issues.append(f"Retention {ret_days}d < {min_retention}d minimum")
            status = STATUS_FAIL

    # Replication check
    if require_replication and not has_replication(policy):
        issues.append("No replication target configured")
        status = STATUS_FAIL

    # Archival check
    if require_archival and not has_archival(policy):
        issues.append("No archival target configured")
        status = STATUS_FAIL

    # RPO check
    if rpo_hours and last_run_end_usecs:
        now_usecs = int(datetime.now(tz=timezone.utc).timestamp() * 1_000_000)
        hours_ago = (now_usecs - last_run_end_usecs) / 1_000_000 / 3600
        if hours_ago > rpo_hours:
            msg = f"Last success {hours_ago:.1f}h ago (RPO={rpo_hours}h)"
            issues.append(msg)
            if status != STATUS_FAIL:
                status = STATUS_WARN
    elif rpo_hours and not last_run_end_usecs:
        issues.append("No completed run found — cannot verify RPO")
        if status != STATUS_FAIL:
            status = STATUS_WARN

    return status, "; ".join(issues) if issues else "OK"


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

STATUS_COLORS = {STATUS_PASS: "C6EFCE", STATUS_WARN: "FFEB9C", STATUS_FAIL: "FFC7CE"}
STATUS_FONTS  = {STATUS_PASS: "276221", STATUS_WARN: "9C5700", STATUS_FAIL: "9C0006"}


def write_excel(rows: list, output_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("ERROR: openpyxl not installed.  pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance"
    headers  = [
        "Cluster", "Protection Group", "Policy Name", "Is Active", "Is Paused",
        "Local Retention (Days)", "Has Replication", "Has Archival",
        "Last Successful Run", "Status", "Issues",
    ]
    ws.append(headers)

    fill_hdr = PatternFill("solid", fgColor=COHESITY_GREEN)
    font_hdr = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = fill_hdr
        cell.font      = font_hdr
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for row in rows:
        status = row.get("Status", STATUS_PASS)
        ws.append([_safe_cell(row.get(h, "")) for h in headers])
        ri = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(ri, col).fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
            ws.cell(ri, col).font = Font(color=STATUS_FONTS[status])

    for col in ws.columns:
        best = 10
        for cell in col:
            if cell.value:
                best = max(best, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(best, 55)

    wb.save(output_path)
    fail = sum(1 for r in rows if r.get("Status") == STATUS_FAIL)
    warn = sum(1 for r in rows if r.get("Status") == STATUS_WARN)
    print(f"\n[+] Report saved: {output_path}")
    print(f"    Groups: {len(rows)} — {fail} FAIL, {warn} WARN, "
          f"{len(rows)-fail-warn} PASS")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=f"Cohesity policy compliance report v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--apikey",              help="Helios API key")
    parser.add_argument("--cluster",             help="Filter to one cluster by name")
    parser.add_argument("--min-retention",       type=int, default=0,
                        help="Minimum local retention in days (e.g. 14)")
    parser.add_argument("--require-replication", action="store_true",
                        help="Flag groups with no replication target")
    parser.add_argument("--require-archival",    action="store_true",
                        help="Flag groups with no archival target")
    parser.add_argument("--rpo-hours",           type=int, default=0,
                        help="Max hours since last successful run (e.g. 24)")
    parser.add_argument("--output",              help="Output .xlsx (auto-generated if omitted)")
    parser.add_argument("--clear-credentials",   action="store_true",
                        help="Remove stored Helios API key and exit")
    parser.add_argument("--ca-bundle", dest="ca_bundle", default=None, metavar="PATH",
                        help="Path to CA bundle for TLS verification (e.g. corporate proxy cert)")
    parser.add_argument("--insecure", action="store_true",
                        help="Disable TLS certificate verification (NOT recommended)")
    args = parser.parse_args()

    global _verify
    if args.insecure:
        _verify = False
        print("=" * 65)
        print("  WARN: --insecure — TLS certificate validation DISABLED.")
        print("        Credentials and tokens may be intercepted (MITM).")
        print("=" * 65)
        try:
            import urllib3 as _u3
            _u3.disable_warnings(_u3.exceptions.InsecureRequestWarning)
        except ImportError:
            requests.packages.urllib3.disable_warnings()
    elif args.ca_bundle:
        _verify = args.ca_bundle

    if args.clear_credentials:
        clear_stored_credentials()
        sys.exit(0)

    api_key  = get_api_key(args.apikey)
    output   = args.output or default_output_path()
    clusters = get_helios_clusters(api_key)

    if args.cluster:
        clusters = [c for c in clusters
                    if c["name"].lower() == args.cluster.lower()]
        if not clusters:
            print(f"ERROR: Cluster '{args.cluster}' not found in Helios.")
            sys.exit(1)

    rows = []
    for c in clusters:
        cname = c["name"]
        cid   = c["clusterId"]
        print(f"\n[*] Processing: {cname}")

        # Load policy map once per cluster
        raw_policies  = hget(api_key, cid, "/v2/data-protect/policies").get("policies", [])
        policy_map    = {p["id"]: p for p in raw_policies}

        groups = hget(api_key, cid,
                      "/v2/data-protect/protection-groups",
                      params={"isActive": "true",
                              "includeTenants": "true"}).get("protectionGroups", [])

        for g in groups:
            gid      = g.get("id", "")
            gname    = g.get("name", "")
            pid      = g.get("policyId", "")
            policy   = policy_map.get(pid, {})
            pname    = policy.get("name", pid)

            # Fetch last run
            runs = hget(api_key, cid,
                        f"/v2/data-protect/protection-groups/{gid}/runs",
                        params={"numRuns": 1,
                                "includeObjectDetails": "false"}).get("runs", [])
            last_end = 0
            last_end_str = ""
            if runs:
                run = runs[0]
                bk  = run.get("localBackupInfo", run.get("backupInfo", {}))
                last_end = bk.get("endTimeUsecs", 0)
                if last_end:
                    from datetime import datetime, timezone
                    last_end_str = datetime.fromtimestamp(
                        last_end / 1_000_000, tz=timezone.utc
                    ).strftime("%Y-%m-%d %H:%M:%S UTC")

            # Retention for display
            backup_pol = policy.get("backupPolicy", {})
            reg        = backup_pol.get("regular", {})
            ret_d      = retention_days(reg.get("retention", {}))

            status, issues = evaluate_group(
                g, policy, last_end,
                args.min_retention, args.require_replication,
                args.require_archival, args.rpo_hours,
            )

            rows.append({
                "Cluster":                 cname,
                "Protection Group":        gname,
                "Policy Name":             pname,
                "Is Active":              "Yes" if g.get("isActive") else "No",
                "Is Paused":              "Yes" if g.get("isPaused") else "No",
                "Local Retention (Days)":  ret_d if ret_d else "",
                "Has Replication":        "Yes" if has_replication(policy) else "No",
                "Has Archival":           "Yes" if has_archival(policy) else "No",
                "Last Successful Run":     last_end_str,
                "Status":                  status,
                "Issues":                  issues,
            })

        print(f"    {len(groups)} group(s) evaluated")

    write_excel(rows, output)


if __name__ == "__main__":
    main()
