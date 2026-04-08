#!/usr/bin/env python3
# =============================================================================
# DISCLAIMER: This code is provided on a best-effort basis and is not in any
# way officially supported or sanctioned by Cohesity. The code is intentionally
# kept simple to retain value as example code. The code in this repository is
# provided as-is and the author accepts no liability for damages resulting
# from its use.
# =============================================================================
"""
protection_group_report.py
------------------------------
Generates a CSV report of protection group run status for Cohesity clusters
running version 7.1 or 7.3.

Supports two authentication modes:
  1. Direct cluster  — username/password against a specific cluster IP/hostname
  2. Helios (SaaS)   — API key against helios.cohesity.com, auto-discovers all
                       connected clusters (or targets a specific one with --cluster)

Default mode: last run only (one row per protection group).
Historical mode: all runs within a date range (one row per run per group).

API reference: https://developer.cohesity.com/
Inspired by:   https://github.com/bseltz-cohesity/scripts

Version history:
  1.0 (2026-04-04) — Initial working release. Last run per group, CSV output,
                     Helios + direct cluster auth.
  2.0 (2026-04-04) — Added --start / --end / --days for historical run data.
                     Each run becomes its own CSV row.
  3.0 (2026-04-04) — Added SLA Violated, Snapshot Expiry, Run Type, Objects
                     Canceled, Replication and Archival status/target/expiry,
                     Storage Consumed (group-level). Smart size formatting
                     (auto TB/GB/MB). Timestamps now reported in the cluster's
                     own configured local timezone (queried from the cluster).
  3.1 (2026-04-04) — Size columns now reported as plain GB numbers (no unit
                     suffix) with "(GB)" in the column header for clean Excel
                     analysis. All timestamps confirmed in cluster local time.
  3.2 (2026-04-04) — Policy ID replaced with Policy Name. Name resolved via
                     GET /v2/data-protect/policies/{id}; results cached so
                     each unique policy is fetched only once per run.
  4.4 (2026-04-06) — Reverted trend mode Storage Consumed to stats/consumers
                     (same source as summary mode). Previous snapshot-summing
                     approach grossly exceeded cluster capacity because
                     bytesWritten does not account for cross-snapshot dedup.
                     Column renamed to "Storage Consumed (As of End Date)"
                     to make clear it is a current-state value, not per-day
                     historical. Startup caveat message added. README updated.
  4.3 (2026-04-05) — Trend mode now includes all summary mode columns
                     alongside the storage trend. Run detail columns (type,
                     status, SLA, objects, sizes, replication, archive) are
                     aggregated across all runs that started on each day:
                     counts/sizes summed, last run used for status/targets,
                     SLA "Yes" if any run violated. Runs On Day shows the
                     count (0 = no backup ran that day).
  4.2 (2026-04-05) — Added Policy Name, Is Active, Is Paused columns to
                     trend mode output. Updated README trend field table.
  4.1 (2026-04-05) — Replaced timeSeriesStats approach (returned 400 —
                     schemaName required, metric unsupported via Helios) with
                     a runs-based storage estimate. For each protection group,
                     queries all runs going back 1 year, then for each calendar
                     day sums bytesWritten for every snapshot active on that day
                     (startTime <= day AND expiryTime > day). Gives an accurate
                     daily storage consumed trend using only the v2 runs API.
  4.0 (2026-04-05) — Added --mode trend framework, --debug flag, trend chart
                     per cluster, Storage Consumed (Current, GB) rename.
  3.4 (2026-04-05) — Report header row color changed to Cohesity green
                     (#00B388).
  3.3 (2026-04-04) — Credentials stored in OS keychain (keyring). --apikey is
                     optional (prompted on first run, saved for future). Direct
                     cluster passwords also stored in keychain. --clear-credentials
                     flag to remove stored credentials. Output changed from CSV to
                     Excel (.xlsx) with styled headers, frozen row, and auto-fit
                     columns. Output filename auto-generated as
                     <script_name>_YYYYMMDD_HHMMSS.xlsx in cwd.

Usage — last run only (default):
  python3 protection_group_report.py
  python3 protection_group_report.py --cluster <ip> --username admin --domain LOCAL

Usage — historical (date range):
  python3 protection_group_report.py --apikey <key> --start 2026-03-01 --end 2026-04-01
  python3 protection_group_report.py --apikey <key> --days 30

Usage — target one cluster via Helios:
  python3 protection_group_report.py --apikey <key> --cluster <cluster-name> --days 7
"""

__version__ = "4.5"

import argparse
import getpass
import os
import sys
import urllib3
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
_KR_SVC_CLUSTER = "cohesity_cluster"


# ---------------------------------------------------------------------------
# Credential store  (OS keychain via keyring)
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring          # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    """
    Return the Helios API key. Priority:
      1. --apikey CLI argument (one-time use, not stored)
      2. Key stored in the OS keychain
      3. Interactive prompt — key is saved to keychain for future runs
    """
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def get_cluster_password(cluster: str, username: str, domain: str) -> str:
    """
    Return the password for direct cluster auth. Prompts if not in keychain,
    then saves it for future runs.
    """
    kr_user = f"{cluster}:{domain}:{username}"
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_CLUSTER, kr_user)
        if stored:
            print(f"[*] Using stored password for {domain}\\{username}@{cluster}")
            return stored
    pwd = getpass.getpass(f"Password for {domain}\\{username}@{cluster}: ").strip()
    if not pwd:
        print("ERROR: Password cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_CLUSTER, kr_user, pwd)
        print(f"[*] Password saved to system keychain for {domain}\\{username}@{cluster}")
    return pwd


def clear_stored_credentials(cluster: str = None, username: str = "admin",
                              domain: str = "LOCAL"):
    """
    Remove stored credentials from the OS keychain.
    Without --cluster: clears the Helios API key.
    With --cluster:    clears the direct-cluster password.
    """
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if cluster:
        kr_user = f"{cluster}:{domain}:{username}"
        if keyring.get_password(_KR_SVC_CLUSTER, kr_user):
            keyring.delete_password(_KR_SVC_CLUSTER, kr_user)
            print(f"[*] Stored password for {domain}\\{username}@{cluster} removed.")
        else:
            print(f"[*] No stored password found for {domain}\\{username}@{cluster}.")
    else:
        if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
            keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
            print("[*] Stored Helios API key removed from system keychain.")
        else:
            print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    """Auto-generate output path: <script_name>_YYYYMMDD_HHMMSS.xlsx in cwd."""
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_v{__version__}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Authentication — Mode 1: Direct cluster (username/password)
# ---------------------------------------------------------------------------

def get_auth_token(cluster: str, username: str, password: str, domain: str) -> str:
    """
    Authenticate to a Cohesity cluster and return a Bearer token.
    Endpoint: POST /v2/users/sessions
    domain: 'LOCAL' for local accounts, or AD domain name (e.g. 'CORP').
    """
    url = f"https://{cluster}/v2/users/sessions"
    payload = {"username": username, "password": password, "domain": domain}
    hdrs = {"Content-Type": "application/json", "Accept": "application/json"}
    try:
        r = requests.post(url, json=payload, headers=hdrs, verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print(f"ERROR: Cannot connect to '{cluster}'. Check hostname/IP and network.")
        sys.exit(1)
    except requests.exceptions.HTTPError as e:
        print(f"ERROR: Authentication failed: {e}\n       Response: {r.text}")
        sys.exit(1)
    token = r.json().get("tokenType", "Bearer") + " " + r.json().get("accessToken", "")
    print(f"[+] Authenticated to {cluster} as {domain}\\{username}")
    return token


def make_headers(token: str) -> dict:
    return {"Authorization": token, "Accept": "application/json",
            "Content-Type": "application/json"}


# ---------------------------------------------------------------------------
# Authentication — Mode 2: Helios (API key)
# ---------------------------------------------------------------------------

HELIOS_HOST = "helios.cohesity.com"


def make_helios_headers(api_key: str, cluster_id: int = None) -> dict:
    """
    Build Helios request headers.
    'accessClusterId' routes the call to a specific on-prem cluster.
    (Confirmed from cohesity-api.ps1 — NOT 'clusterIdentifier'.)
    """
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str, filter_name: str = None) -> list:
    """
    Retrieve all clusters connected to this Helios account.
    Endpoint: GET https://helios.cohesity.com/mcm/clusters/connectionStatus
    """
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=make_helios_headers(api_key), verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError as e:
        print(f"ERROR: Helios auth failed: {e}\n       Check your API key. Response: {r.text}")
        sys.exit(1)

    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])

    if filter_name:
        clusters = [c for c in clusters if filter_name.lower() in c.get("name", "").lower()]
        if not clusters:
            print(f"ERROR: No cluster found matching '{filter_name}'")
            sys.exit(1)

    print(f"[+] Helios: found {len(clusters)} connected cluster(s)")
    return clusters


# ---------------------------------------------------------------------------
# Cluster info
# ---------------------------------------------------------------------------

def get_cluster_timezone(base_url: str, headers: dict) -> str:
    """
    Fetch the cluster's configured IANA timezone string (e.g. 'America/Chicago').
    Endpoint: GET /irisservices/api/v1/public/cluster
    Falls back to UTC if unavailable.
    """
    url = f"{base_url}/irisservices/api/v1/public/cluster"
    try:
        r = requests.get(url, headers=headers, verify=False, timeout=15)
        r.raise_for_status()
        tz_name = r.json().get("timezone") or "UTC"
        ZoneInfo(tz_name)          # validate it's a recognised timezone
        return tz_name
    except Exception:
        return "UTC"


def get_policy_name(base_url: str, headers: dict, policy_id: str,
                    _cache: dict = {}) -> str:
    """
    Resolve a policy ID to its human-readable name.
    Endpoint: GET /v2/data-protect/policies/{id}
    Results are cached in-process so each unique policy is fetched only once.
    """
    if not policy_id or policy_id == "N/A":
        return "N/A"
    cache_key = f"{base_url}:{policy_id}"
    if cache_key in _cache:
        return _cache[cache_key]
    url = f"{base_url}/v2/data-protect/policies/{policy_id}"
    try:
        r = requests.get(url, headers=headers, verify=False, timeout=15)
        r.raise_for_status()
        name = r.json().get("name", policy_id)
    except Exception:
        name = policy_id   # fall back to ID if lookup fails
    _cache[cache_key] = name
    return name


def get_storage_stats(base_url: str, headers: dict) -> dict:
    """
    Fetch current storage consumption per protection group in one bulk call.
    Endpoint: GET /irisservices/api/v1/public/stats/consumers?consumerType=kProtectionRuns
    Returns: {numeric_group_id (int): storageConsumedBytes (int)}

    'storageConsumedBytes' is the total physical space the group currently
    occupies on the cluster (all snapshots combined, post-dedup/compression).
    This is a group-level aggregate, not a per-run figure.
    """
    url = f"{base_url}/irisservices/api/v1/public/stats/consumers"
    params = {"consumerType": "kProtectionRuns", "fetchViewBoxName": "true"}
    try:
        r = requests.get(url, headers=headers, params=params, verify=False, timeout=30)
        r.raise_for_status()
        return {
            item["id"]: item.get("stats", {}).get("storageConsumedBytes", 0)
            for item in (r.json().get("statsList") or [])
            if "id" in item
        }
    except Exception as e:
        print(f"    WARNING: Could not fetch storage stats: {e}")
        return {}


def numeric_group_id(group_id: str) -> int:
    """
    Extract the numeric job ID from a v2 protection group ID.
    Direct cluster IDs are plain integers.
    Helios compound IDs look like '2297498838261838:local:12345' — we want 12345.
    """
    try:
        return int(group_id.split(":")[-1])
    except (ValueError, IndexError, AttributeError):
        try:
            return int(group_id)
        except (ValueError, TypeError):
            return -1


# ---------------------------------------------------------------------------
# Size and time formatting
# ---------------------------------------------------------------------------

def to_gb(byte_count) -> str:
    """
    Convert a byte count to GB, returned as a plain decimal number string.
    Column headers carry the "(GB)" label so cell values stay numeric for
    Excel pivot tables and sorting.
    Uses binary GiB (1024^3 bytes), consistent with Cohesity's raw values.
    """
    if not byte_count:
        return "N/A"
    return f"{byte_count / 1024 ** 3:.2f}"


def format_time(usecs: int, tz_name: str = "UTC") -> str:
    """
    Convert a Cohesity microsecond epoch timestamp to a human-readable string
    in the cluster's local timezone (e.g. '2026-04-01 10:30 CDT').
    Falls back to UTC if the timezone name is invalid.
    """
    if not usecs:
        return "N/A"
    try:
        tz = ZoneInfo(tz_name)
    except ZoneInfoNotFoundError:
        tz = timezone.utc
    dt = datetime.fromtimestamp(usecs / 1_000_000, tz=tz)
    return dt.strftime("%Y-%m-%d %H:%M %Z")


# ---------------------------------------------------------------------------
# Date helpers (CLI range resolution)
# ---------------------------------------------------------------------------

def date_to_usecs(date_str: str) -> int:
    """Convert YYYY-MM-DD to Cohesity microsecond epoch (start of day UTC)."""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        print(f"ERROR: Invalid date '{date_str}'. Use YYYY-MM-DD (e.g. 2026-03-01).")
        sys.exit(1)
    return int(dt.timestamp() * 1_000_000)


def end_of_day_usecs(date_str: str) -> int:
    """Convert YYYY-MM-DD to microseconds at 23:59:59 UTC (inclusive end of day)."""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        print(f"ERROR: Invalid date '{date_str}'. Use YYYY-MM-DD (e.g. 2026-03-01).")
        sys.exit(1)
    return int(dt.timestamp() * 1_000_000)


def now_usecs() -> int:
    return int(datetime.now(tz=timezone.utc).timestamp() * 1_000_000)


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def get_protection_groups(base_url: str, headers: dict) -> list:
    """
    Fetch all active protection groups.
    Endpoint: GET /v2/data-protect/protection-groups
    Pattern confirmed from cohesityInv.ps1 (apiType='ClusterV2').
    """
    url = f"{base_url}/v2/data-protect/protection-groups"
    params = {"isDeleted": "false", "includeTenants": "true"}
    try:
        r = requests.get(url, headers=headers, params=params, verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print(f"    WARNING: Failed to fetch groups ({r.status_code}): {r.text[:200]}")
        return []
    groups = r.json().get("protectionGroups") or []
    print(f"    Found {len(groups)} protection group(s)")
    return groups


def get_runs(base_url: str, headers: dict, group_id: str,
             start_usecs: int = None, end_usecs: int = None) -> list:
    """
    Fetch protection runs for a group.

    Default (no dates): returns the most recent run only (numRuns=1).
    Date range provided: returns all runs between start and end (capped at 1000).

    Endpoint: GET /v2/data-protect/protection-groups/{id}/runs
    Confirmed from cohesityInv.ps1 line 1275 (apiType='ClusterV2'):
      ?startTimeUsecs={}&endTimeUsecs={}&includeObjectDetails=true&includeTenants=true
    """
    url = f"{base_url}/v2/data-protect/protection-groups/{group_id}/runs"
    params = {"includeObjectDetails": "true", "includeTenants": "true"}
    if start_usecs and end_usecs:
        params["startTimeUsecs"] = start_usecs
        params["endTimeUsecs"]   = end_usecs
        params["numRuns"]        = 1000   # safety cap for very long ranges
    else:
        params["numRuns"] = 1
    try:
        r = requests.get(url, headers=headers, params=params, verify=False, timeout=60)
        r.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print(f"      WARNING: Could not fetch runs ({r.status_code}): {r.text[:150]}")
        return []
    return r.json().get("runs") or []




# ---------------------------------------------------------------------------
# Run stats parser
# ---------------------------------------------------------------------------

def extract_run_stats(run: dict, tz_name: str = "UTC") -> dict:
    """
    Parse a v2 protection group run and return all available metrics.

    Fields extracted:
      From localBackupInfo:
        status, runType, isSlaViolated, startTimeUsecs, endTimeUsecs,
        successfulObjectsCount, failedObjectsCount, cancelledObjectsCount,
        localSnapshotStats.bytesRead, bytesWritten, logicalSizeBytes

      From objects[0].localSnapshotInfo.snapshotInfo:
        expiryTimeUsecs  (snapshot retention date)

      From replicationInfo.replicationTargetResults[]:
        status, clusterName, expiryTimeUsecs  (per replication target)

      From archivalInfo.archivalTargetResults[]:
        status, targetName, targetType, expiryTimeUsecs  (per archive target)

    Multiple replication/archive targets are pipe-separated in each column.
    Timestamps are formatted in the cluster's local timezone (tz_name).
    Sizes are auto-formatted to MB/GB/TB based on magnitude.
    """
    empty = {
        "run_type":              "N/A",
        "sla_violated":          "N/A",
        "run_status":            "No runs in range",
        "run_start":             "N/A",
        "run_end":               "N/A",
        "duration_mins":         "N/A",
        "snapshot_expiry":       "N/A",
        "objects_succeeded":     "N/A",
        "objects_failed":        "N/A",
        "objects_canceled":      "N/A",
        "data_read":             "N/A",
        "data_written":          "N/A",
        "logical_size":          "N/A",
        "replication_status":    "N/A",
        "replication_targets":   "N/A",
        "replication_expiry":    "N/A",
        "archive_status":        "N/A",
        "archive_targets":       "N/A",
        "archive_expiry":        "N/A",
    }

    if not run:
        return empty

    backup = run.get("localBackupInfo", {})
    if not backup:
        return empty

    status      = backup.get("status", "Unknown")
    run_type_raw = backup.get("runType", "")
    run_type    = run_type_raw.lstrip("k") if run_type_raw else "N/A"  # kRegular → Regular
    sla_violated = "Yes" if backup.get("isSlaViolated") else "No"

    start_usecs = backup.get("startTimeUsecs", 0)
    end_usecs   = backup.get("endTimeUsecs", 0)
    duration_mins = (
        round((end_usecs - start_usecs) / 1_000_000 / 60, 1)
        if start_usecs and end_usecs else "N/A"
    )

    # Snapshot expiry from the first object (all objects in a run share the same expiry)
    snapshot_expiry = "N/A"
    objects = run.get("objects", [])
    if objects:
        first_snap = (objects[0]
                      .get("localSnapshotInfo", {})
                      .get("snapshotInfo", {}))
        expiry_usecs = first_snap.get("expiryTimeUsecs", 0)
        snapshot_expiry = format_time(expiry_usecs, tz_name) if expiry_usecs else "N/A"

    # Object counts — use direct counts from localBackupInfo (faster, no loop needed)
    succeeded = backup.get("successfulObjectsCount",  "N/A")
    failed    = backup.get("failedObjectsCount",      "N/A")
    canceled  = backup.get("cancelledObjectsCount",   "N/A")

    # Byte stats
    snap_stats = backup.get("localSnapshotStats", {})

    # Replication — summarise all targets into pipe-separated columns
    rep_results = run.get("replicationInfo", {}).get("replicationTargetResults", [])
    if rep_results:
        rep_status  = " | ".join(t.get("status", "")      for t in rep_results)
        rep_targets = " | ".join(t.get("clusterName", "") for t in rep_results)
        rep_expiry  = " | ".join(
            format_time(t.get("expiryTimeUsecs", 0), tz_name) for t in rep_results
        )
    else:
        rep_status = rep_targets = rep_expiry = "None"

    # Archival — summarise all targets into pipe-separated columns
    arch_results = run.get("archivalInfo", {}).get("archivalTargetResults", [])
    if arch_results:
        arch_status  = " | ".join(t.get("status", "")     for t in arch_results)
        arch_targets = " | ".join(t.get("targetName", "") for t in arch_results)
        arch_expiry  = " | ".join(
            format_time(t.get("expiryTimeUsecs", 0), tz_name) for t in arch_results
        )
    else:
        arch_status = arch_targets = arch_expiry = "None"

    return {
        "run_type":            run_type,
        "sla_violated":        sla_violated,
        "run_status":          status,
        "run_start":           format_time(start_usecs, tz_name),
        "run_end":             format_time(end_usecs, tz_name),
        "duration_mins":       duration_mins,
        "snapshot_expiry":     snapshot_expiry,
        "objects_succeeded":   succeeded,
        "objects_failed":      failed,
        "objects_canceled":    canceled,
        "data_read":           to_gb(snap_stats.get("bytesRead")),
        "data_written":        to_gb(snap_stats.get("bytesWritten")),
        "logical_size":        to_gb(snap_stats.get("logicalSizeBytes")),
        "replication_status":  rep_status,
        "replication_targets": rep_targets,
        "replication_expiry":  rep_expiry,
        "archive_status":      arch_status,
        "archive_targets":     arch_targets,
        "archive_expiry":      arch_expiry,
    }


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

# Columns whose headers should carry a "(GB)" unit label.
# Cell values are plain decimals so Excel can sort/pivot them numerically.
# "Storage Consumed" is always a current-state snapshot (not per-run historical)
# so it is labelled "(Current, GB)" to avoid confusion in summary/historical mode.
_GB_COL_RENAMES = {
    "Data Read":        "Data Read (GB)",
    "Data Written":     "Data Written (GB)",
    "Logical Size":     "Logical Size (GB)",
    "Storage Consumed": "Storage Consumed (Current, GB)",
}

# ---------------------------------------------------------------------------
# Trend mode columns  (--mode trend)
# ---------------------------------------------------------------------------

TREND_COLUMNS = [
    # Group identity / context
    ("Date",                     "date"),
    ("Cluster",                  "cluster"),
    ("Protection Group",         "protection_group"),
    ("Environment",              "environment"),
    ("Policy Name",              "policy_name"),
    ("Is Active",                "is_active"),
    ("Is Paused",                "is_paused"),
    # Storage — current state as of report end date (same value on every row for a group)
    ("Storage Consumed (As of End Date, Bytes)", "storage_consumed_bytes"),
    ("Storage Consumed (As of End Date, TB)",    "storage_consumed_tb"),
    # Run detail for the day (aggregated across all runs that started on this date)
    ("Runs On Day",              "runs_on_day"),
    ("Run Type",                 "run_type"),
    ("Sla Violated",             "sla_violated"),
    ("Run Status",               "run_status"),
    ("Run Start",                "run_start"),
    ("Run End",                  "run_end"),
    ("Duration Mins",            "duration_mins"),
    ("Snapshot Expiry",          "snapshot_expiry"),
    ("Objects Succeeded",        "objects_succeeded"),
    ("Objects Failed",           "objects_failed"),
    ("Objects Canceled",         "objects_canceled"),
    ("Data Read (GB)",           "data_read"),
    ("Data Written (GB)",        "data_written"),
    ("Logical Size (GB)",        "logical_size"),
    ("Replication Status",       "replication_status"),
    ("Replication Targets",      "replication_targets"),
    ("Replication Expiry",       "replication_expiry"),
    ("Archive Status",           "archive_status"),
    ("Archive Targets",          "archive_targets"),
    ("Archive Expiry",           "archive_expiry"),
]

_TREND_TB_COL_IDX = next(i for i, (_, k) in enumerate(TREND_COLUMNS, start=1)
                          if k == "storage_consumed_tb")   # = 9


def build_report(cluster_label: str, base_url: str, headers: dict, groups: list,
                 start_usecs: int = None, end_usecs: int = None) -> list:
    """
    For each protection group, fetch runs and build report rows.

    Default (no dates): one row per group (last run).
    Historical (dates given): one row per run per group.

    Per-cluster setup (done once each):
      - Cluster timezone   → all timestamps in cluster's local time
      - Storage stats bulk fetch → storageConsumedBytes per group
    """
    historical = start_usecs is not None and end_usecs is not None
    report_rows = []

    # Fetch timezone and storage stats once for this cluster
    print("    Fetching cluster timezone...")
    tz_name = get_cluster_timezone(base_url, headers)
    print(f"    Cluster timezone: {tz_name}")

    print("    Fetching storage stats...")
    storage_map = get_storage_stats(base_url, headers)  # {numeric_id: bytes}

    for group in groups:
        group_id    = group.get("id", "")
        group_name  = group.get("name", "Unknown")
        environment = group.get("environment", "").lstrip("k")   # kVMware → VMware
        policy_id   = group.get("policyId", "N/A")
        is_active   = group.get("isActive", True)
        is_paused   = group.get("isPaused", False)

        num_id        = numeric_group_id(group_id)
        storage_consumed = to_gb(storage_map.get(num_id, 0) or None)
        policy_name   = get_policy_name(base_url, headers, policy_id)

        mode_label = "runs" if historical else "last run"
        print(f"      Fetching {mode_label} for: {group_name}")
        runs = get_runs(base_url, headers, group_id, start_usecs, end_usecs)

        run_stats_list = [extract_run_stats(r, tz_name) for r in runs] if runs \
                         else [extract_run_stats({}, tz_name)]

        for i, run_stats in enumerate(run_stats_list, start=1):
            row = {
                "Cluster":           cluster_label,
                "Protection Group":  group_name,
                "Environment":       environment,
                "Is Active":         is_active,
                "Is Paused":         is_paused,
                "Policy Name":       policy_name,
                "Storage Consumed":  storage_consumed,
            }
            if historical:
                row["Run #"] = i
            row.update(run_stats)
            # CSV-friendly column names; rename size columns to carry "(GB)" label
            titled = {k.replace("_", " ").title(): v for k, v in row.items()}
            report_rows.append(
                {_GB_COL_RENAMES.get(k, k): v for k, v in titled.items()}
            )

    return report_rows


def _sum_gb(values: list) -> str:
    """Sum a list of GB decimal strings (some may be 'N/A'). Returns formatted GB string."""
    total = 0.0
    has_value = False
    for v in values:
        if v != "N/A":
            try:
                total += float(v)
                has_value = True
            except (ValueError, TypeError):
                pass
    return f"{total:.2f}" if has_value else "N/A"


def aggregate_day_runs(runs: list, tz_name: str = "UTC") -> dict:
    """
    Aggregate run-level stats for all runs that started on a given day.

    0 runs  → all fields "N/A" / "No runs"
    1 run   → pass-through from extract_run_stats
    N runs  → object counts and data sizes summed; last run used for
              status, type, targets, and expiry; SLA = "Yes" if any run
              violated; Run Start = earliest, Run End = latest.
    """
    _empty = {
        "runs_on_day":         0,
        "run_type":            "N/A",
        "sla_violated":        "N/A",
        "run_status":          "No runs",
        "run_start":           "N/A",
        "run_end":             "N/A",
        "duration_mins":       "N/A",
        "snapshot_expiry":     "N/A",
        "objects_succeeded":   "N/A",
        "objects_failed":      "N/A",
        "objects_canceled":    "N/A",
        "data_read":           "N/A",
        "data_written":        "N/A",
        "logical_size":        "N/A",
        "replication_status":  "N/A",
        "replication_targets": "N/A",
        "replication_expiry":  "N/A",
        "archive_status":      "N/A",
        "archive_targets":     "N/A",
        "archive_expiry":      "N/A",
    }

    if not runs:
        return _empty

    stats = [extract_run_stats(r, tz_name) for r in runs]
    last  = stats[-1]

    if len(stats) == 1:
        return {"runs_on_day": 1, **last}

    # Multiple runs on the same day — aggregate
    def _sum_int(key):
        total = sum(s[key] for s in stats if isinstance(s.get(key), int))
        return total if total or any(isinstance(s.get(key), int) for s in stats) else "N/A"

    def _sum_dur(key):
        vals = [s[key] for s in stats if isinstance(s.get(key), (int, float))]
        return round(sum(vals), 1) if vals else "N/A"

    starts = [s["run_start"] for s in stats if s.get("run_start") not in ("N/A", None)]
    ends   = [s["run_end"]   for s in stats if s.get("run_end")   not in ("N/A", None)]

    return {
        "runs_on_day":         len(runs),
        "run_type":            last["run_type"],
        "sla_violated":        "Yes" if any(s["sla_violated"] == "Yes" for s in stats) else "No",
        "run_status":          last["run_status"],
        "run_start":           starts[0]  if starts else "N/A",
        "run_end":             ends[-1]   if ends   else "N/A",
        "duration_mins":       _sum_dur("duration_mins"),
        "snapshot_expiry":     last["snapshot_expiry"],
        "objects_succeeded":   _sum_int("objects_succeeded"),
        "objects_failed":      _sum_int("objects_failed"),
        "objects_canceled":    _sum_int("objects_canceled"),
        "data_read":           _sum_gb([s["data_read"]    for s in stats]),
        "data_written":        _sum_gb([s["data_written"] for s in stats]),
        "logical_size":        _sum_gb([s["logical_size"] for s in stats]),
        "replication_status":  last["replication_status"],
        "replication_targets": last["replication_targets"],
        "replication_expiry":  last["replication_expiry"],
        "archive_status":      last["archive_status"],
        "archive_targets":     last["archive_targets"],
        "archive_expiry":      last["archive_expiry"],
    }


def build_trend_report(cluster_label: str, base_url: str, headers: dict,
                        groups: list, start_usecs: int, end_usecs: int,
                        debug: bool = False) -> list:
    """
    Trend mode: one row per protection group per day.

    Storage Consumed is sourced from GET /stats/consumers (same as summary
    mode) — a point-in-time snapshot of current cluster storage at the time
    the script runs. This value is stamped identically on every date row for
    a given protection group.

    LIMITATION: Cohesity does not expose a per-day historical storage consumed
    value per protection group via any public API accessible through Helios.
    A previous implementation attempted to estimate this by summing bytesWritten
    across active snapshots per day, but the result grossly exceeded actual
    cluster capacity because bytesWritten does not account for cross-snapshot
    deduplication. The current approach is honest: the number is accurate but
    static across all dates for a group. Use the Run Detail columns (Data
    Written, Objects count, etc.) for per-day trend analysis of backup activity.
    """
    print("    Fetching cluster timezone...")
    tz_name = get_cluster_timezone(base_url, headers)
    print(f"    Cluster timezone: {tz_name}")

    print("    Fetching storage stats (current state)...")
    storage_map = get_storage_stats(base_url, headers)  # {numeric_id: bytes}

    from datetime import date as _date
    from collections import defaultdict
    start_date = datetime.fromtimestamp(start_usecs / 1e6, tz=timezone.utc).date()
    end_date   = datetime.fromtimestamp(end_usecs   / 1e6, tz=timezone.utc).date()
    date_list  = [start_date + timedelta(days=i)
                  for i in range((end_date - start_date).days + 1)]

    rows = []
    for group in groups:
        group_id    = group.get("id", "")
        group_name  = group.get("name", "Unknown")
        environment = group.get("environment", "").lstrip("k")
        policy_id   = group.get("policyId", "N/A")
        is_active   = group.get("isActive", True)
        is_paused   = group.get("isPaused", False)
        policy_name = get_policy_name(base_url, headers, policy_id)
        num_id      = numeric_group_id(group_id)

        # Current storage — same value on every date row for this group
        consumed_bytes = storage_map.get(num_id, 0) or 0
        consumed_tb    = round(consumed_bytes / 1_000_000_000_000, 4)

        print(f"      [{group_name}] querying runs...")
        all_runs = get_runs(base_url, headers, group_id, start_usecs, end_usecs)

        # Index runs by the calendar date they started
        runs_by_date: dict = defaultdict(list)
        for run in (all_runs or []):
            backup  = run.get("localBackupInfo") or {}
            start_t = backup.get("startTimeUsecs", 0)
            if start_t:
                run_date = datetime.fromtimestamp(
                    start_t / 1e6, tz=timezone.utc).date()
                runs_by_date[run_date].append(run)

        for d in date_list:
            run_stats = aggregate_day_runs(runs_by_date.get(d, []), tz_name)
            rows.append({
                "date":                   d.isoformat(),
                "cluster":                cluster_label,
                "protection_group":       group_name,
                "environment":            environment,
                "policy_name":            policy_name,
                "is_active":              is_active,
                "is_paused":              is_paused,
                "storage_consumed_bytes": consumed_bytes,
                "storage_consumed_tb":    consumed_tb,
                **run_stats,
            })

        print(f"        {len(date_list)} row(s), storage consumed: {consumed_tb:.4f} TB")

    return rows


# ---------------------------------------------------------------------------
# Trend chart helpers  (one sheet per cluster, mirrors fortknox style)
# ---------------------------------------------------------------------------

def _safe_pg_sheet_name(name: str, existing: list) -> str:
    """Sanitise and deduplicate an Excel sheet name (max 31 chars)."""
    safe = name.translate(str.maketrans("", "", r":\/?*[]"))[:31]
    if safe in existing:
        safe = safe[:28] + f"_{len(existing)}"
    return safe


def _make_pg_chart(title: str, report_ws, series_list: list,
                   LineChart, Reference, SeriesLabel, Legend):
    """
    Build a styled LineChart for a single cluster's protection groups.

    series_list: [(label, start_row, end_row), ...]
    X axis: Date (col 1 in Report sheet)
    Y axis: Storage Consumed (TB) (col _TREND_TB_COL_IDX in Report sheet)
    Tick density auto-reduces for long ranges.
    """
    chart = LineChart()
    chart.style  = 10
    chart.height = 16
    chart.width  = 28

    # Title — 16 pt bold; plain string fallback
    try:
        from openpyxl.chart.title import Title
        from openpyxl.chart.text import RichText as ChartRichText, TxChoice
        from openpyxl.drawing.text import (
            Paragraph, RegularTextRun, RichTextProperties,
            ParagraphProperties, CharacterProperties,
        )
        chart.title = Title(
            tx=TxChoice(rich=ChartRichText(
                bodyPr=RichTextProperties(),
                p=[Paragraph(
                    pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1600, b=True)),
                    r=[RegularTextRun(t=title)]
                )]
            )),
            overlay=False,
        )
    except Exception:
        chart.title = title

    # Y axis
    chart.y_axis.title         = "Storage Consumed (TB)"
    chart.y_axis.numFmt        = "0.000"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.tickLblPos    = "nextTo"
    chart.y_axis.delete        = False

    # X axis (string dates — no numFmt)
    chart.x_axis.title         = "Date"
    chart.x_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos    = "low"
    chart.x_axis.delete        = False

    # Auto-reduce tick density for long ranges
    n_days = max((end - start + 1) for _, start, end in series_list) if series_list else 1
    if n_days > 60:
        chart.x_axis.tickLblSkip = 7
    elif n_days > 14:
        chart.x_axis.tickLblSkip = 2

    # Legend at bottom, no overlap
    legend = Legend()
    legend.position = "b"
    legend.overlay  = False
    chart.legend    = legend

    # Use StrRef so Excel renders string dates as labels, not numbers.
    # openpyxl's set_categories() generates <c:numRef> which causes blank x-axis.
    try:
        from openpyxl.chart.data_source import DataSource, StrRef
        from openpyxl.utils import get_column_letter as _gcl
        _use_strref = True
    except ImportError:
        _use_strref = False

    cat_formula = None
    for label, start_row, end_row in series_list:
        chart.add_data(
            Reference(report_ws, min_col=_TREND_TB_COL_IDX,
                      min_row=start_row, max_row=end_row))
        chart.series[-1].title = SeriesLabel(v=label)

        if cat_formula is None:
            col_a = _gcl(1) if _use_strref else "A"
            cat_formula = (f"'{report_ws.title}'!${col_a}${start_row}"
                           f":${col_a}${end_row}")

    if cat_formula:
        if _use_strref:
            str_cat = DataSource(strRef=StrRef(f=cat_formula))
            for s in chart.series:
                s.cat = str_cat
        else:
            chart.set_categories(Reference(report_ws, min_col=1,
                                           min_row=series_list[0][1],
                                           max_row=series_list[0][2]))

    return chart


def _add_pg_trend_charts(wb, group_ranges: dict):
    """
    Add one chart sheet per cluster to wb.

    group_ranges: {(cluster, group): {"start": int, "end": int, "any_nonzero": bool}}
    """
    try:
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.chart.legend import Legend
    except ImportError:
        print("WARNING: openpyxl chart module unavailable — charts skipped.")
        return

    report_ws = wb["Report"]
    clusters  = sorted({k[0] for k in group_ranges})
    used_names = [s.title for s in wb.worksheets]
    total = 0

    for cluster in clusters:
        series_list = []
        for (c, group), info in sorted(group_ranges.items(), key=lambda x: x[0][1]):
            if c != cluster or not info["any_nonzero"]:
                continue
            series_list.append((group, info["start"], info["end"]))

        if not series_list:
            print(f"    [{cluster}] all-zero — chart skipped")
            continue

        chart = _make_pg_chart(
            f"Local Storage \u2014 {cluster}",
            report_ws, series_list,
            LineChart, Reference, SeriesLabel, Legend)

        sheet_name = _safe_pg_sheet_name(cluster, used_names)
        used_names.append(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        ws.add_chart(chart, "A1")
        total += 1
        print(f"    Chart sheet '{sheet_name}': {len(series_list)} series")

    if total == 0:
        print("    No non-zero Storage Consumed data — all charts skipped.")
    else:
        print(f"    {total} chart sheet(s) added (one per cluster)")


def write_excel(rows: list, output_file: str, mode: str = "summary"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required.  Install with:  pip install openpyxl")
        sys.exit(1)

    if not rows:
        print("No data to write.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Styled header row — Cohesity green
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="70AD47")

    if mode == "trend":
        # --- Trend mode: fixed TREND_COLUMNS, sorted for contiguous chart series ---
        col_names = [col for col, _ in TREND_COLUMNS]
        ws.append(col_names)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        sorted_rows = sorted(rows, key=lambda r: (
            r["protection_group"], r["cluster"], r["date"]))

        group_ranges: dict = {}
        for excel_row, r in enumerate(sorted_rows, start=2):
            key      = (r["cluster"], r["protection_group"])
            consumed = int(r.get("storage_consumed_bytes") or 0)
            if key not in group_ranges:
                group_ranges[key] = {"start": excel_row, "end": excel_row,
                                     "any_nonzero": consumed > 0}
            else:
                group_ranges[key]["end"] = excel_row
                if consumed > 0:
                    group_ranges[key]["any_nonzero"] = True

        for r in sorted_rows:
            ws.append([r[key] for _, key in TREND_COLUMNS])

        headers = col_names

    else:
        # --- Summary / historical mode: dynamic headers from row keys ---
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(list(row.values()))

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(header)),
            max((len(str(ws.cell(r, col_idx).value or ""))
                 for r in range(2, ws.max_row + 1)), default=0)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Trend charts — one sheet per cluster
    if mode == "trend":
        _add_pg_trend_charts(wb, group_ranges)

    wb.save(output_file)
    print(f"\n[+] Report saved to: {output_file}")
    print(f"    Total rows: {len(rows)}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description=f"Cohesity protection group run report  v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  First run — prompts for Helios API key, saves to keychain:
    python3 protection_group_report.py

  Subsequent runs — key retrieved automatically:
    python3 protection_group_report.py --days 30
    python3 protection_group_report.py --start 2026-03-01 --end 2026-04-01

  Direct cluster — password prompted and saved to keychain:
    python3 protection_group_report.py --cluster <ip> --username admin --domain LOCAL

  One-time Helios key override (not saved):
    python3 protection_group_report.py --apikey <key>

  Clear stored Helios API key:
    python3 protection_group_report.py --clear-credentials

  Clear stored cluster password:
    python3 protection_group_report.py --clear-credentials --cluster <ip>
        """
    )
    parser.add_argument("--cluster",
                        help="Cluster hostname/IP (direct auth) or name filter (Helios)")
    parser.add_argument("--username", default="admin",
                        help="Username for direct cluster auth (default: admin)")
    parser.add_argument("--domain", default="LOCAL",
                        help="Auth domain: LOCAL or AD domain name (default: LOCAL)")
    parser.add_argument("--apikey", default=None,
                        help="Helios API key (optional — keychain used if omitted)")
    parser.add_argument("--clear-credentials", action="store_true", dest="clear_creds",
                        help="Remove stored credentials from keychain and exit")
    parser.add_argument("--output", default=None,
                        help="Output Excel filename (default: <script_name>_YYYYMMDD_HHMMSS.xlsx "
                             "in current directory)")
    parser.add_argument("--mode", choices=["summary", "trend"], default="summary",
                        help="summary (default): last run per group or all runs in date range. "
                             "trend: daily storage consumed per group via timeSeriesStats — "
                             "one row per group per day + trend chart per cluster. "
                             "Defaults to last 30 days when no date range is given.")
    parser.add_argument("--debug", action="store_true",
                        help="Print raw timeSeriesStats API response for metric verification")

    dg = parser.add_argument_group("Date range (required for trend mode; optional for summary)")
    dg.add_argument("--start", metavar="YYYY-MM-DD",
                    help="Start date (inclusive)")
    dg.add_argument("--end",   metavar="YYYY-MM-DD",
                    help="End date (inclusive, defaults to today)")
    dg.add_argument("--days",  type=int, metavar="N",
                    help="Last N days (e.g. --days 30)")
    return parser.parse_args()


def resolve_date_range(args):
    """Convert --start/--end/--days to (start_usecs, end_usecs) or (None, None)."""
    if args.days:
        start_dt = datetime.now(tz=timezone.utc) - timedelta(days=args.days)
        return int(start_dt.timestamp() * 1_000_000), now_usecs()
    if args.start:
        return date_to_usecs(args.start), \
               (end_of_day_usecs(args.end) if args.end else now_usecs())
    if args.end:
        print("WARNING: --end given without --start. Ignoring.")
    return None, None


def main():
    args = parse_args()

    # Handle --clear-credentials before anything else
    if args.clear_creds:
        clear_stored_credentials(
            cluster=args.cluster if args.cluster and not _is_helios_mode(args) else None,
            username=args.username,
            domain=args.domain,
        )
        sys.exit(0)

    output_path = args.output or default_output_path()
    start_usecs, end_usecs = resolve_date_range(args)

    # Trend mode requires a date range — default to last 30 days
    if args.mode == "trend" and start_usecs is None:
        start_usecs = int((datetime.now(tz=timezone.utc) - timedelta(days=30))
                          .timestamp() * 1_000_000)
        end_usecs   = now_usecs()
        print("[*] Trend mode: no date range given — defaulting to last 30 days")

    print(f"\n[*] Protection group report  v{__version__}")
    print(f"[*] Mode:   {args.mode}")
    print(f"[*] Output: {output_path}")

    if start_usecs:
        s = datetime.fromtimestamp(start_usecs / 1e6, tz=timezone.utc).strftime("%Y-%m-%d")
        e = datetime.fromtimestamp(end_usecs   / 1e6, tz=timezone.utc).strftime("%Y-%m-%d")
        print(f"[*] Date range: {s} → {e}")
    else:
        print("[*] Date range: last run per protection group (no range given)")

    if args.mode == "trend":
        print()
        print("    LIMITATION — Storage Consumed:")
        print("    Cohesity does not provide a historical per-day storage consumed value")
        print("    per protection group via any public API accessible through Helios.")
        print("    The 'Storage Consumed (As of End Date)' column reflects the CURRENT")
        print(f"   state of the cluster as of when this script ran — it will show the")
        print("    same value on every date row for a given protection group.")
        print("    For per-day backup activity trends, use the Run Detail columns")
        print("    (Data Written, Objects Succeeded, Run Status, etc.).")
        print()

    all_rows = []

    # --- Helios mode ---
    if _is_helios_mode(args):
        api_key = get_api_key(args.apikey)
        print(f"[*] Helios mode — connecting to {HELIOS_HOST}")
        clusters = get_helios_clusters(api_key, filter_name=args.cluster)

        for cluster_info in clusters:
            cluster_name = cluster_info.get("name", "Unknown")
            cluster_id   = cluster_info.get("clusterId")
            print(f"\n[*] Processing cluster: {cluster_name} (id={cluster_id})")

            h        = make_helios_headers(api_key, cluster_id=cluster_id)
            base_url = f"https://{HELIOS_HOST}"

            print("    Fetching protection groups...")
            groups = get_protection_groups(base_url, h)
            if not groups:
                continue

            if args.mode == "trend":
                rows = build_trend_report(cluster_name, base_url, h, groups,
                                          start_usecs, end_usecs, debug=args.debug)
            else:
                rows = build_report(cluster_name, base_url, h, groups,
                                    start_usecs, end_usecs)
            all_rows.extend(rows)

    # --- Direct cluster mode ---
    elif args.cluster:
        print(f"[*] Direct mode — connecting to: {args.cluster}")
        password = get_cluster_password(args.cluster, args.username, args.domain)
        token    = get_auth_token(args.cluster, args.username, password, args.domain)
        headers  = make_headers(token)
        base_url = f"https://{args.cluster}"

        print("[*] Fetching protection groups...")
        groups = get_protection_groups(base_url, headers)

        if args.mode == "trend":
            all_rows = build_trend_report(args.cluster, base_url, headers, groups,
                                          start_usecs, end_usecs, debug=args.debug)
        else:
            all_rows = build_report(args.cluster, base_url, headers, groups,
                                    start_usecs, end_usecs)

    else:
        print("ERROR: Provide --apikey / keychain (Helios) or --cluster + credentials (direct).")
        print("       Run with --help for usage examples.")
        sys.exit(1)

    write_excel(all_rows, output_path, mode=args.mode)


def _is_helios_mode(args) -> bool:
    """True when running against Helios (API key present, or no cluster given)."""
    return bool(args.apikey) or not args.cluster


if __name__ == "__main__":
    main()
