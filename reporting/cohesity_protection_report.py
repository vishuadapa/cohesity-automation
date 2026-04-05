#!/usr/bin/env python3
"""
cohesity_protection_report.py
------------------------------
Generates a CSV report of protection group run status for Cohesity clusters
running version 7.1 or 7.3.

Supports two authentication modes:
  1. Direct cluster  — username/password against a specific cluster IP/hostname
  2. Helios (SaaS)   — API key against helios.cohesity.com, auto-discovers all
                       connected clusters (or targets a specific one with --cluster)

Default mode: last run only (one row per protection group).
Historical mode: all runs within a date range (one row per run per group).

API reference: https://developer.cohesity.com/
Inspired by:   https://github.com/bseltz-cohesity/scripts

Version history:
  1.0 (2026-04-04) — Initial working release. Last run per group, CSV output,
                     Helios + direct cluster auth.
  2.0 (2026-04-04) — Added --start / --end / --days for historical run data.
                     Each run becomes its own CSV row.
  3.0 (2026-04-04) — Added SLA Violated, Snapshot Expiry, Run Type, Objects
                     Canceled, Replication and Archival status/target/expiry,
                     Storage Consumed (group-level). Smart size formatting
                     (auto TB/GB/MB). Timestamps now reported in the cluster's
                     own configured local timezone (queried from the cluster).
  3.1 (2026-04-04) — Size columns now reported as plain GB numbers (no unit
                     suffix) with "(GB)" in the column header for clean Excel
                     analysis. All timestamps confirmed in cluster local time.
  3.2 (2026-04-04) — Policy ID replaced with Policy Name. Name resolved via
                     GET /v2/data-protect/policies/{id}; results cached so
                     each unique policy is fetched only once per run.
  3.4 (2026-04-05) — Report header row color changed to Cohesity green
                     (#00B388).
  3.3 (2026-04-04) — Credentials stored in OS keychain (keyring). --apikey is
                     optional (prompted on first run, saved for future). Direct
                     cluster passwords also stored in keychain. --clear-credentials
                     flag to remove stored credentials. Output changed from CSV to
                     Excel (.xlsx) with styled headers, frozen row, and auto-fit
                     columns. Output filename auto-generated as
                     <script_name>_YYYYMMDD_HHMMSS.xlsx in cwd.

Usage — last run only (default):
  python3 cohesity_protection_report.py
  python3 cohesity_protection_report.py --cluster <ip> --username admin --domain LOCAL

Usage — historical (date range):
  python3 cohesity_protection_report.py --apikey <key> --start 2026-03-01 --end 2026-04-01
  python3 cohesity_protection_report.py --apikey <key> --days 30

Usage — target one cluster via Helios:
  python3 cohesity_protection_report.py --apikey <key> --cluster <cluster-name> --days 7
"""

__version__ = "3.4"

import argparse
import getpass
import os
import sys
import urllib3
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

_KR_SVC_HELIOS  = "cohesity_helios"
_KR_USER_HELIOS = "apikey"
_KR_SVC_CLUSTER = "cohesity_cluster"


# ---------------------------------------------------------------------------
# Credential store  (OS keychain via keyring)
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring          # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    """
    Return the Helios API key. Priority:
      1. --apikey CLI argument (one-time use, not stored)
      2. Key stored in the OS keychain
      3. Interactive prompt — key is saved to keychain for future runs
    """
    if cli_key:
        return cli_key
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
        if stored:
            print("[*] Using Helios API key stored in system keychain.")
            return stored
        print("[*] No stored Helios API key found.")
    else:
        print("    NOTE: 'keyring' not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
    key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
    if not key:
        print("ERROR: API key cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_HELIOS, _KR_USER_HELIOS, key)
        print("[*] Helios API key saved to system keychain.")
    return key


def get_cluster_password(cluster: str, username: str, domain: str) -> str:
    """
    Return the password for direct cluster auth. Prompts if not in keychain,
    then saves it for future runs.
    """
    kr_user = f"{cluster}:{domain}:{username}"
    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KR_SVC_CLUSTER, kr_user)
        if stored:
            print(f"[*] Using stored password for {domain}\\{username}@{cluster}")
            return stored
    pwd = getpass.getpass(f"Password for {domain}\\{username}@{cluster}: ").strip()
    if not pwd:
        print("ERROR: Password cannot be empty.")
        sys.exit(1)
    if _keyring_available():
        import keyring
        keyring.set_password(_KR_SVC_CLUSTER, kr_user, pwd)
        print(f"[*] Password saved to system keychain for {domain}\\{username}@{cluster}")
    return pwd


def clear_stored_credentials(cluster: str = None, username: str = "admin",
                              domain: str = "LOCAL"):
    """
    Remove stored credentials from the OS keychain.
    Without --cluster: clears the Helios API key.
    With --cluster:    clears the direct-cluster password.
    """
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    if cluster:
        kr_user = f"{cluster}:{domain}:{username}"
        if keyring.get_password(_KR_SVC_CLUSTER, kr_user):
            keyring.delete_password(_KR_SVC_CLUSTER, kr_user)
            print(f"[*] Stored password for {domain}\\{username}@{cluster} removed.")
        else:
            print(f"[*] No stored password found for {domain}\\{username}@{cluster}.")
    else:
        if keyring.get_password(_KR_SVC_HELIOS, _KR_USER_HELIOS):
            keyring.delete_password(_KR_SVC_HELIOS, _KR_USER_HELIOS)
            print("[*] Stored Helios API key removed from system keychain.")
        else:
            print("[*] No stored Helios API key found — nothing to clear.")


def default_output_path() -> str:
    """Auto-generate output path: <script_name>_YYYYMMDD_HHMMSS.xlsx in cwd."""
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Authentication — Mode 1: Direct cluster (username/password)
# ---------------------------------------------------------------------------

def get_auth_token(cluster: str, username: str, password: str, domain: str) -> str:
    """
    Authenticate to a Cohesity cluster and return a Bearer token.
    Endpoint: POST /v2/users/sessions
    domain: 'LOCAL' for local accounts, or AD domain name (e.g. 'CORP').
    """
    url = f"https://{cluster}/v2/users/sessions"
    payload = {"username": username, "password": password, "domain": domain}
    hdrs = {"Content-Type": "application/json", "Accept": "application/json"}
    try:
        r = requests.post(url, json=payload, headers=hdrs, verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print(f"ERROR: Cannot connect to '{cluster}'. Check hostname/IP and network.")
        sys.exit(1)
    except requests.exceptions.HTTPError as e:
        print(f"ERROR: Authentication failed: {e}\n       Response: {r.text}")
        sys.exit(1)
    token = r.json().get("tokenType", "Bearer") + " " + r.json().get("accessToken", "")
    print(f"[+] Authenticated to {cluster} as {domain}\\{username}")
    return token


def make_headers(token: str) -> dict:
    return {"Authorization": token, "Accept": "application/json",
            "Content-Type": "application/json"}


# ---------------------------------------------------------------------------
# Authentication — Mode 2: Helios (API key)
# ---------------------------------------------------------------------------

HELIOS_HOST = "helios.cohesity.com"


def make_helios_headers(api_key: str, cluster_id: int = None) -> dict:
    """
    Build Helios request headers.
    'accessClusterId' routes the call to a specific on-prem cluster.
    (Confirmed from cohesity-api.ps1 — NOT 'clusterIdentifier'.)
    """
    h = {"apiKey": api_key, "Accept": "application/json",
         "Content-Type": "application/json"}
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str, filter_name: str = None) -> list:
    """
    Retrieve all clusters connected to this Helios account.
    Endpoint: GET https://helios.cohesity.com/mcm/clusters/connectionStatus
    """
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=make_helios_headers(api_key), verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError as e:
        print(f"ERROR: Helios auth failed: {e}\n       Check your API key. Response: {r.text}")
        sys.exit(1)

    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])

    if filter_name:
        clusters = [c for c in clusters if filter_name.lower() in c.get("name", "").lower()]
        if not clusters:
            print(f"ERROR: No cluster found matching '{filter_name}'")
            sys.exit(1)

    print(f"[+] Helios: found {len(clusters)} connected cluster(s)")
    return clusters


# ---------------------------------------------------------------------------
# Cluster info
# ---------------------------------------------------------------------------

def get_cluster_timezone(base_url: str, headers: dict) -> str:
    """
    Fetch the cluster's configured IANA timezone string (e.g. 'America/Chicago').
    Endpoint: GET /irisservices/api/v1/public/cluster
    Falls back to UTC if unavailable.
    """
    url = f"{base_url}/irisservices/api/v1/public/cluster"
    try:
        r = requests.get(url, headers=headers, verify=False, timeout=15)
        r.raise_for_status()
        tz_name = r.json().get("timezone") or "UTC"
        ZoneInfo(tz_name)          # validate it's a recognised timezone
        return tz_name
    except Exception:
        return "UTC"


def get_policy_name(base_url: str, headers: dict, policy_id: str,
                    _cache: dict = {}) -> str:
    """
    Resolve a policy ID to its human-readable name.
    Endpoint: GET /v2/data-protect/policies/{id}
    Results are cached in-process so each unique policy is fetched only once.
    """
    if not policy_id or policy_id == "N/A":
        return "N/A"
    cache_key = f"{base_url}:{policy_id}"
    if cache_key in _cache:
        return _cache[cache_key]
    url = f"{base_url}/v2/data-protect/policies/{policy_id}"
    try:
        r = requests.get(url, headers=headers, verify=False, timeout=15)
        r.raise_for_status()
        name = r.json().get("name", policy_id)
    except Exception:
        name = policy_id   # fall back to ID if lookup fails
    _cache[cache_key] = name
    return name


def get_storage_stats(base_url: str, headers: dict) -> dict:
    """
    Fetch current storage consumption per protection group in one bulk call.
    Endpoint: GET /irisservices/api/v1/public/stats/consumers?consumerType=kProtectionRuns
    Returns: {numeric_group_id (int): storageConsumedBytes (int)}

    'storageConsumedBytes' is the total physical space the group currently
    occupies on the cluster (all snapshots combined, post-dedup/compression).
    This is a group-level aggregate, not a per-run figure.
    """
    url = f"{base_url}/irisservices/api/v1/public/stats/consumers"
    params = {"consumerType": "kProtectionRuns", "fetchViewBoxName": "true"}
    try:
        r = requests.get(url, headers=headers, params=params, verify=False, timeout=30)
        r.raise_for_status()
        return {
            item["id"]: item.get("stats", {}).get("storageConsumedBytes", 0)
            for item in (r.json().get("statsList") or [])
            if "id" in item
        }
    except Exception as e:
        print(f"    WARNING: Could not fetch storage stats: {e}")
        return {}


def numeric_group_id(group_id: str) -> int:
    """
    Extract the numeric job ID from a v2 protection group ID.
    Direct cluster IDs are plain integers.
    Helios compound IDs look like '2297498838261838:local:12345' — we want 12345.
    """
    try:
        return int(group_id.split(":")[-1])
    except (ValueError, IndexError, AttributeError):
        try:
            return int(group_id)
        except (ValueError, TypeError):
            return -1


# ---------------------------------------------------------------------------
# Size and time formatting
# ---------------------------------------------------------------------------

def to_gb(byte_count) -> str:
    """
    Convert a byte count to GB, returned as a plain decimal number string.
    Column headers carry the "(GB)" label so cell values stay numeric for
    Excel pivot tables and sorting.
    Uses binary GiB (1024^3 bytes), consistent with Cohesity's raw values.
    """
    if not byte_count:
        return "N/A"
    return f"{byte_count / 1024 ** 3:.2f}"


def format_time(usecs: int, tz_name: str = "UTC") -> str:
    """
    Convert a Cohesity microsecond epoch timestamp to a human-readable string
    in the cluster's local timezone (e.g. '2026-04-01 10:30 CDT').
    Falls back to UTC if the timezone name is invalid.
    """
    if not usecs:
        return "N/A"
    try:
        tz = ZoneInfo(tz_name)
    except ZoneInfoNotFoundError:
        tz = timezone.utc
    dt = datetime.fromtimestamp(usecs / 1_000_000, tz=tz)
    return dt.strftime("%Y-%m-%d %H:%M %Z")


# ---------------------------------------------------------------------------
# Date helpers (CLI range resolution)
# ---------------------------------------------------------------------------

def date_to_usecs(date_str: str) -> int:
    """Convert YYYY-MM-DD to Cohesity microsecond epoch (start of day UTC)."""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        print(f"ERROR: Invalid date '{date_str}'. Use YYYY-MM-DD (e.g. 2026-03-01).")
        sys.exit(1)
    return int(dt.timestamp() * 1_000_000)


def end_of_day_usecs(date_str: str) -> int:
    """Convert YYYY-MM-DD to microseconds at 23:59:59 UTC (inclusive end of day)."""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        print(f"ERROR: Invalid date '{date_str}'. Use YYYY-MM-DD (e.g. 2026-03-01).")
        sys.exit(1)
    return int(dt.timestamp() * 1_000_000)


def now_usecs() -> int:
    return int(datetime.now(tz=timezone.utc).timestamp() * 1_000_000)


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def get_protection_groups(base_url: str, headers: dict) -> list:
    """
    Fetch all active protection groups.
    Endpoint: GET /v2/data-protect/protection-groups
    Pattern confirmed from cohesityInv.ps1 (apiType='ClusterV2').
    """
    url = f"{base_url}/v2/data-protect/protection-groups"
    params = {"isDeleted": "false", "includeTenants": "true"}
    try:
        r = requests.get(url, headers=headers, params=params, verify=False, timeout=30)
        r.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print(f"    WARNING: Failed to fetch groups ({r.status_code}): {r.text[:200]}")
        return []
    groups = r.json().get("protectionGroups") or []
    print(f"    Found {len(groups)} protection group(s)")
    return groups


def get_runs(base_url: str, headers: dict, group_id: str,
             start_usecs: int = None, end_usecs: int = None) -> list:
    """
    Fetch protection runs for a group.

    Default (no dates): returns the most recent run only (numRuns=1).
    Date range provided: returns all runs between start and end (capped at 1000).

    Endpoint: GET /v2/data-protect/protection-groups/{id}/runs
    Confirmed from cohesityInv.ps1 line 1275 (apiType='ClusterV2'):
      ?startTimeUsecs={}&endTimeUsecs={}&includeObjectDetails=true&includeTenants=true
    """
    url = f"{base_url}/v2/data-protect/protection-groups/{group_id}/runs"
    params = {"includeObjectDetails": "true", "includeTenants": "true"}
    if start_usecs and end_usecs:
        params["startTimeUsecs"] = start_usecs
        params["endTimeUsecs"]   = end_usecs
        params["numRuns"]        = 1000   # safety cap for very long ranges
    else:
        params["numRuns"] = 1
    try:
        r = requests.get(url, headers=headers, params=params, verify=False, timeout=60)
        r.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print(f"      WARNING: Could not fetch runs ({r.status_code}): {r.text[:150]}")
        return []
    return r.json().get("runs") or []


# ---------------------------------------------------------------------------
# Run stats parser
# ---------------------------------------------------------------------------

def extract_run_stats(run: dict, tz_name: str = "UTC") -> dict:
    """
    Parse a v2 protection group run and return all available metrics.

    Fields extracted:
      From localBackupInfo:
        status, runType, isSlaViolated, startTimeUsecs, endTimeUsecs,
        successfulObjectsCount, failedObjectsCount, cancelledObjectsCount,
        localSnapshotStats.bytesRead, bytesWritten, logicalSizeBytes

      From objects[0].localSnapshotInfo.snapshotInfo:
        expiryTimeUsecs  (snapshot retention date)

      From replicationInfo.replicationTargetResults[]:
        status, clusterName, expiryTimeUsecs  (per replication target)

      From archivalInfo.archivalTargetResults[]:
        status, targetName, targetType, expiryTimeUsecs  (per archive target)

    Multiple replication/archive targets are pipe-separated in each column.
    Timestamps are formatted in the cluster's local timezone (tz_name).
    Sizes are auto-formatted to MB/GB/TB based on magnitude.
    """
    empty = {
        "run_type":              "N/A",
        "sla_violated":          "N/A",
        "run_status":            "No runs in range",
        "run_start":             "N/A",
        "run_end":               "N/A",
        "duration_mins":         "N/A",
        "snapshot_expiry":       "N/A",
        "objects_succeeded":     "N/A",
        "objects_failed":        "N/A",
        "objects_canceled":      "N/A",
        "data_read":             "N/A",
        "data_written":          "N/A",
        "logical_size":          "N/A",
        "replication_status":    "N/A",
        "replication_targets":   "N/A",
        "replication_expiry":    "N/A",
        "archive_status":        "N/A",
        "archive_targets":       "N/A",
        "archive_expiry":        "N/A",
    }

    if not run:
        return empty

    backup = run.get("localBackupInfo", {})
    if not backup:
        return empty

    status      = backup.get("status", "Unknown")
    run_type_raw = backup.get("runType", "")
    run_type    = run_type_raw.lstrip("k") if run_type_raw else "N/A"  # kRegular → Regular
    sla_violated = "Yes" if backup.get("isSlaViolated") else "No"

    start_usecs = backup.get("startTimeUsecs", 0)
    end_usecs   = backup.get("endTimeUsecs", 0)
    duration_mins = (
        round((end_usecs - start_usecs) / 1_000_000 / 60, 1)
        if start_usecs and end_usecs else "N/A"
    )

    # Snapshot expiry from the first object (all objects in a run share the same expiry)
    snapshot_expiry = "N/A"
    objects = run.get("objects", [])
    if objects:
        first_snap = (objects[0]
                      .get("localSnapshotInfo", {})
                      .get("snapshotInfo", {}))
        expiry_usecs = first_snap.get("expiryTimeUsecs", 0)
        snapshot_expiry = format_time(expiry_usecs, tz_name) if expiry_usecs else "N/A"

    # Object counts — use direct counts from localBackupInfo (faster, no loop needed)
    succeeded = backup.get("successfulObjectsCount",  "N/A")
    failed    = backup.get("failedObjectsCount",      "N/A")
    canceled  = backup.get("cancelledObjectsCount",   "N/A")

    # Byte stats
    snap_stats = backup.get("localSnapshotStats", {})

    # Replication — summarise all targets into pipe-separated columns
    rep_results = run.get("replicationInfo", {}).get("replicationTargetResults", [])
    if rep_results:
        rep_status  = " | ".join(t.get("status", "")      for t in rep_results)
        rep_targets = " | ".join(t.get("clusterName", "") for t in rep_results)
        rep_expiry  = " | ".join(
            format_time(t.get("expiryTimeUsecs", 0), tz_name) for t in rep_results
        )
    else:
        rep_status = rep_targets = rep_expiry = "None"

    # Archival — summarise all targets into pipe-separated columns
    arch_results = run.get("archivalInfo", {}).get("archivalTargetResults", [])
    if arch_results:
        arch_status  = " | ".join(t.get("status", "")     for t in arch_results)
        arch_targets = " | ".join(t.get("targetName", "") for t in arch_results)
        arch_expiry  = " | ".join(
            format_time(t.get("expiryTimeUsecs", 0), tz_name) for t in arch_results
        )
    else:
        arch_status = arch_targets = arch_expiry = "None"

    return {
        "run_type":            run_type,
        "sla_violated":        sla_violated,
        "run_status":          status,
        "run_start":           format_time(start_usecs, tz_name),
        "run_end":             format_time(end_usecs, tz_name),
        "duration_mins":       duration_mins,
        "snapshot_expiry":     snapshot_expiry,
        "objects_succeeded":   succeeded,
        "objects_failed":      failed,
        "objects_canceled":    canceled,
        "data_read":           to_gb(snap_stats.get("bytesRead")),
        "data_written":        to_gb(snap_stats.get("bytesWritten")),
        "logical_size":        to_gb(snap_stats.get("logicalSizeBytes")),
        "replication_status":  rep_status,
        "replication_targets": rep_targets,
        "replication_expiry":  rep_expiry,
        "archive_status":      arch_status,
        "archive_targets":     arch_targets,
        "archive_expiry":      arch_expiry,
    }


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

# Columns whose headers should carry a "(GB)" unit label.
# Cell values are plain decimals so Excel can sort/pivot them numerically.
_GB_COL_RENAMES = {
    "Data Read":        "Data Read (GB)",
    "Data Written":     "Data Written (GB)",
    "Logical Size":     "Logical Size (GB)",
    "Storage Consumed": "Storage Consumed (GB)",
}


def build_report(cluster_label: str, base_url: str, headers: dict, groups: list,
                 start_usecs: int = None, end_usecs: int = None) -> list:
    """
    For each protection group, fetch runs and build report rows.

    Default (no dates): one row per group (last run).
    Historical (dates given): one row per run per group.

    Per-cluster setup (done once each):
      - Cluster timezone   → all timestamps in cluster's local time
      - Storage stats bulk fetch → storageConsumedBytes per group
    """
    historical = start_usecs is not None and end_usecs is not None
    report_rows = []

    # Fetch timezone and storage stats once for this cluster
    print("    Fetching cluster timezone...")
    tz_name = get_cluster_timezone(base_url, headers)
    print(f"    Cluster timezone: {tz_name}")

    print("    Fetching storage stats...")
    storage_map = get_storage_stats(base_url, headers)  # {numeric_id: bytes}

    for group in groups:
        group_id    = group.get("id", "")
        group_name  = group.get("name", "Unknown")
        environment = group.get("environment", "").lstrip("k")   # kVMware → VMware
        policy_id   = group.get("policyId", "N/A")
        is_active   = group.get("isActive", True)
        is_paused   = group.get("isPaused", False)

        num_id        = numeric_group_id(group_id)
        storage_consumed = to_gb(storage_map.get(num_id, 0) or None)
        policy_name   = get_policy_name(base_url, headers, policy_id)

        mode_label = "runs" if historical else "last run"
        print(f"      Fetching {mode_label} for: {group_name}")
        runs = get_runs(base_url, headers, group_id, start_usecs, end_usecs)

        run_stats_list = [extract_run_stats(r, tz_name) for r in runs] if runs \
                         else [extract_run_stats({}, tz_name)]

        for i, run_stats in enumerate(run_stats_list, start=1):
            row = {
                "Cluster":           cluster_label,
                "Protection Group":  group_name,
                "Environment":       environment,
                "Is Active":         is_active,
                "Is Paused":         is_paused,
                "Policy Name":       policy_name,
                "Storage Consumed":  storage_consumed,
            }
            if historical:
                row["Run #"] = i
            row.update(run_stats)
            # CSV-friendly column names; rename size columns to carry "(GB)" label
            titled = {k.replace("_", " ").title(): v for k, v in row.items()}
            report_rows.append(
                {_GB_COL_RENAMES.get(k, k): v for k, v in titled.items()}
            )

    return report_rows


def write_excel(rows: list, output_file: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required.  Install with:  pip install openpyxl")
        sys.exit(1)

    if not rows:
        print("No data to write.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = list(rows[0].keys())

    # Styled header row — Cohesity green
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="00B388")
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in rows:
        ws.append(list(row.values()))

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(header)),
            max((len(str(ws.cell(r, col_idx).value or ""))
                 for r in range(2, ws.max_row + 1)), default=0)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_file)
    print(f"\n[+] Report saved to: {output_file}")
    print(f"    Total rows: {len(rows)}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description=f"Cohesity protection group run report  v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  First run — prompts for Helios API key, saves to keychain:
    python3 cohesity_protection_report.py

  Subsequent runs — key retrieved automatically:
    python3 cohesity_protection_report.py --days 30
    python3 cohesity_protection_report.py --start 2026-03-01 --end 2026-04-01

  Direct cluster — password prompted and saved to keychain:
    python3 cohesity_protection_report.py --cluster <ip> --username admin --domain LOCAL

  One-time Helios key override (not saved):
    python3 cohesity_protection_report.py --apikey <key>

  Clear stored Helios API key:
    python3 cohesity_protection_report.py --clear-credentials

  Clear stored cluster password:
    python3 cohesity_protection_report.py --clear-credentials --cluster <ip>
        """
    )
    parser.add_argument("--cluster",
                        help="Cluster hostname/IP (direct auth) or name filter (Helios)")
    parser.add_argument("--username", default="admin",
                        help="Username for direct cluster auth (default: admin)")
    parser.add_argument("--domain", default="LOCAL",
                        help="Auth domain: LOCAL or AD domain name (default: LOCAL)")
    parser.add_argument("--apikey", default=None,
                        help="Helios API key (optional — keychain used if omitted)")
    parser.add_argument("--clear-credentials", action="store_true", dest="clear_creds",
                        help="Remove stored credentials from keychain and exit")
    parser.add_argument("--output", default=None,
                        help="Output Excel filename (default: <script_name>_YYYYMMDD_HHMMSS.xlsx "
                             "in current directory)")

    dg = parser.add_argument_group("Historical date range (optional)")
    dg.add_argument("--start", metavar="YYYY-MM-DD",
                    help="Start date for historical runs (inclusive)")
    dg.add_argument("--end",   metavar="YYYY-MM-DD",
                    help="End date for historical runs (inclusive, defaults to today)")
    dg.add_argument("--days",  type=int, metavar="N",
                    help="Shortcut: last N days (e.g. --days 30)")
    return parser.parse_args()


def resolve_date_range(args):
    """Convert --start/--end/--days to (start_usecs, end_usecs) or (None, None)."""
    if args.days:
        start_dt = datetime.now(tz=timezone.utc) - timedelta(days=args.days)
        return int(start_dt.timestamp() * 1_000_000), now_usecs()
    if args.start:
        return date_to_usecs(args.start), \
               (end_of_day_usecs(args.end) if args.end else now_usecs())
    if args.end:
        print("WARNING: --end given without --start. Ignoring.")
    return None, None


def main():
    args = parse_args()

    # Handle --clear-credentials before anything else
    if args.clear_creds:
        clear_stored_credentials(
            cluster=args.cluster if args.cluster and not _is_helios_mode(args) else None,
            username=args.username,
            domain=args.domain,
        )
        sys.exit(0)

    output_path = args.output or default_output_path()
    start_usecs, end_usecs = resolve_date_range(args)

    print(f"\n[*] Protection group report  v{__version__}")
    print(f"[*] Output: {output_path}")
    if start_usecs:
        s = datetime.fromtimestamp(start_usecs / 1e6, tz=timezone.utc).strftime("%Y-%m-%d")
        e = datetime.fromtimestamp(end_usecs   / 1e6, tz=timezone.utc).strftime("%Y-%m-%d")
        print(f"[*] Historical mode: {s} → {e} (timestamps in cluster local time)")
    else:
        print("[*] Default mode: last run per protection group")

    all_rows = []

    # --- Helios mode (API key present, or no --cluster given, or --cluster is a name not an IP) ---
    if _is_helios_mode(args):
        api_key = get_api_key(args.apikey)
        print(f"[*] Helios mode — connecting to {HELIOS_HOST}")
        clusters = get_helios_clusters(api_key, filter_name=args.cluster)

        for cluster_info in clusters:
            cluster_name = cluster_info.get("name", "Unknown")
            cluster_id   = cluster_info.get("clusterId")
            print(f"\n[*] Processing cluster: {cluster_name} (id={cluster_id})")

            h        = make_helios_headers(api_key, cluster_id=cluster_id)
            base_url = f"https://{HELIOS_HOST}"

            print("    Fetching protection groups...")
            groups = get_protection_groups(base_url, h)
            if groups:
                rows = build_report(cluster_name, base_url, h, groups,
                                    start_usecs, end_usecs)
                all_rows.extend(rows)

    # --- Direct cluster mode ---
    elif args.cluster:
        print(f"[*] Direct mode — connecting to: {args.cluster}")
        password = get_cluster_password(args.cluster, args.username, args.domain)
        token    = get_auth_token(args.cluster, args.username, password, args.domain)
        headers  = make_headers(token)
        base_url = f"https://{args.cluster}"

        print("[*] Fetching protection groups...")
        groups   = get_protection_groups(base_url, headers)
        all_rows = build_report(args.cluster, base_url, headers, groups,
                                start_usecs, end_usecs)

    else:
        print("ERROR: Provide --apikey / keychain (Helios) or --cluster + credentials (direct).")
        print("       Run with --help for usage examples.")
        sys.exit(1)

    write_excel(all_rows, output_path)


def _is_helios_mode(args) -> bool:
    """True when running against Helios (API key present, or no cluster given)."""
    return bool(args.apikey) or not args.cluster


if __name__ == "__main__":
    main()
