#!/usr/bin/env python3
# =============================================================================
# DISCLAIMER: This code is provided on a best-effort basis and is not in any
# way officially supported or sanctioned by Cohesity. The code is intentionally
# kept simple to retain value as example code. The code in this repository is
# provided as-is and the author accepts no liability for damages resulting
# from its use.
# =============================================================================
"""
fortknox_vault_report.py
------------------------
Reports FortKnox (RPaaS) vault data transfer by protection group.

Iterates every Helios-connected cluster, fetches the "Data Transferred to
External Targets" report (GET /irisservices/api/v1/public/reports/dataTransferToVaults)
via Helios routing, and emits one row per protection group per vault.

API endpoints used:
  Cluster list:     GET  https://helios.cohesity.com/mcm/clusters/connectionStatus
  FortKnox vaults:  GET  https://helios.cohesity.com/irisservices/api/v1/public/vaults
                         ?includeFortKnoxVault=true
  Transfer report:  GET  https://helios.cohesity.com/irisservices/api/v1/public/reports/dataTransferToVaults
                         ?startTimeMsecs=<ms>&endTimeMsecs=<ms>&vaultIds=<id>&vaultIds=<id>

Requires Helios API key — FortKnox is a Helios-managed feature.

Version history:
  4.21 (2026-04-21) — Renamed Storage Consumed columns to Remote Storage
                     Consumed to clarify they reflect vault (remote) storage.
                     Added TB (÷ 1e12, 4 dp) companion columns for all four
                     Activities fields, interleaved after each Bytes column:
                     Activities: Data Read (TB), Data Written (TB), Logical
                     Transferred (TB), Physical Transferred (TB). About tab
                     field reference updated accordingly.
  4.20 (2026-04-20) — Removed Logical Transferred (Bytes) and Physical
                     Transferred (Bytes) columns (formerly G and H). These
                     were sourced from GET /public/protectionRuns which is
                     now unused. Deleted get_protection_runs_transfer() and
                     all related callers, row fields, and About tab entries.
  4.19 (2026-04-20) — Fixed protection runs data limited to last ~10 days.
                     Root cause: the v2 protection-groups runs endpoint caps
                     results to a small page size (typically 10 runs per page)
                     regardless of numRuns.  get_activities_transfer() now
                     loops using the paginationCookie returned in each response
                     until no further cookie is present.  get_protection_runs_
                     transfer() (v1) now uses time-based pagination: requests
                     pages of 1000 runs descending by time, advancing
                     endTimeUsecs to one microsecond before the earliest run
                     on the previous page, until fewer than a full page is
                     returned or the start boundary is reached.  Both
                     functions print the total run count retrieved when
                     --debug is set.
  4.18 (2026-04-20) — All green bar fills changed to 70AD47 (About tab
                     sections/headers and Report tab source-banner and
                     column-header rows).
  4.17 (2026-04-20) — Reworked Activities data collection to match the
                     Protection Activities report model: Backup and Vault
                     are separate run entries in the API, not phases of the
                     same run. Backup activity (Data Read, Data Written) is
                     now collected independently by backup start date;
                     Vault activity (Logical, Physical Transferred) is
                     collected independently by vault start date. Results
                     are merged by (group, date) — on days where both kick
                     off at roughly the same time all four fields populate
                     the same row. --debug now reports backup run count and
                     vault run count separately per group.
  4.16 (2026-04-20) — Fixed Activities: Data Read / Data Written date
                     alignment. Previously anchored to the local backup
                     start time; now anchored to the FortKnox archival
                     activity start time (fk_results[0].startTimeUsecs),
                     so Data Read and Data Written always land on the same
                     row/date as the corresponding Logical and Physical
                     Transferred values. Backup spanning midnight no longer
                     causes a one-day offset.
  4.15 (2026-04-20) — About tab: all color fills unified to Cohesity green
                     (70AD47) — removed the secondary light-green (70AD47);
                     section labels now use the same green bar style as the
                     title. Report tab: added source-banner row 1 showing
                     which API/report each column group originates from
                     (Computed, Helios MCM, dataTransferToVaults,
                     protectionRuns, Protection Activities); column headers
                     moved to row 2, data to row 3; both rows frozen; header
                     fill also updated to 70AD47.
  4.14 (2026-04-20) — Fixed Activities: Logical/Physical Transferred always
                     showing 0. Two root causes: (1) wrong field names —
                     v2 API uses logicalBytesTransferred / physicalBytesTransferred,
                     not logicalTransferredBytes / physicalTransferredBytes;
                     (2) vault name match could silently fail — added targetType
                     fallback matching using all known FortKnox type strings
                     (krpaas, kfortknox, kfort_knox, krpaasarchival and plain
                     equivalents). _xfer_bytes() now checks stats sub-dict then
                     top-level of each archival result. --debug now prints raw
                     archivalTargetResult keys and all (targetName, targetType)
                     pairs seen in the response for easy diagnosis.
  4.13 (2026-04-20) — Added four "Activities:" columns sourced from the
                     protection activities report filtered to cloud vault
                     (FortKnox) archival runs: Data Read, Data Written,
                     Logical Transferred, Physical Transferred. Fetched via
                     GET /v2/data-protect/protection-groups/{id}/runs and
                     matched to FortKnox vault names. Summary mode aggregates
                     all days per group; trend mode aligns per day. Two new
                     APIs added to About sheet reference. get_fortknox_vaults()
                     now returns (ids, names) replacing get_fortknox_vault_ids().
  4.12 (2026-04-20) — Added "About" tab (first sheet) containing the command
                     run, report parameters, field reference (column name,
                     description, source field path, API endpoint), and APIs
                     used. Fixed default_output_path() docstring — filename
                     already included _v<version>_ since v4.3.
  4.11 (2026-04-17) — Security: TLS verification enabled by default; added
                     --ca-bundle and --insecure CLI flags. Excel formula
                     injection hardening via _safe_cell().
  4.10 (2026-04-07) — Fixed blank Logical/Physical Transferred: vault ID in
                     protectionRuns copyRun is nested at
                     target.archivalTarget.vaultId, not target.vaultId.
                     The vault ID match always failed, leaving all rows at 0.
  4.9 (2026-04-07) — Fixed Logical Transferred and Physical Transferred to be
                     time-windowed instead of cumulative lifetime totals.
                     Root cause: numLogicalBytesTransferred /
                     numPhysicalBytesTransferred in dataTransferPerProtectionJob
                     are confirmed cumulative and are NOT scoped by the API
                     time window. Fix: added GET /public/protectionRuns call
                     (one per cluster) which returns per-run stats that are
                     inherently time-windowed. Sums logicalBytesTransferred /
                     physicalBytesTransferred across all kSuccess archival
                     runs targeting FortKnox vault IDs within the query range.
                     storageConsumed unchanged (already matched UI). In trend
                     mode, runs are bucketed by calendar day so each row gets
                     only that day's transfers.
  4.8 (2026-04-07) — Added Storage Consumed (TiB) column (÷ 1,099,511,627,776)
                     alongside the existing TB column so values match the Helios
                     UI directly. Fixed --days / default date range to snap to
                     00:00:00 of the start day (calendar-day boundary) instead
                     of a rolling window from the current time. Fixed
                     end_of_day_msecs to include full last second (23:59:59.999)
                     so the final second of a period is not excluded.
  4.7 (2026-04-06) — Fixed x-axis dates not showing on trend charts. Root
                     cause: openpyxl set_categories() emits <c:numRef> in
                     chart XML; Excel discards string values passed via numRef.
                     Fix: assign DataSource(strRef=StrRef(...)) directly to
                     each series' .cat — forces <c:strRef> so Excel renders
                     date strings as text labels on the x-axis.
  4.6 (2026-04-05) — Fixed axis labels not rendering: removed numFmt from
                     category (string) x-axis, added delete=False on both
                     axes. X-axis tick density auto-reduces: every day for
                     ≤14 days, every 2 days for 15-60 days, weekly for >60.
  4.5 (2026-04-05) — One chart sheet per cluster (named after the cluster)
                     instead of a single combined sheet. Each chart shows all
                     protection groups for that cluster as separate series.
  4.4 (2026-04-05) — Trend chart: 16 pt bold title, legend moved to bottom
                     (no overlap), x-axis labelled "Date" with yyyy-mm-dd
                     format, y-axis labelled "Storage Consumed (TB)". Header
                     row color changed to Cohesity green (#70AD47).
  4.3 (2026-04-04) — Secure credential storage via OS keychain (keyring).
                     --apikey is now optional; key is prompted once, saved to
                     the system keychain, and retrieved automatically on future
                     runs. --clear-credentials removes the stored key.
                     Output filename auto-generated as
                     <script_name>_YYYYMMDD_HHMMSS.xlsx in cwd.
  4.2 (2026-04-04) — Added Storage Consumed (TB) column (÷ 1,000,000,000,000,
                     4 decimal places). Trend chart now plots TB values instead
                     of raw bytes for human-readable Y axis.
  4.1 (2026-04-04) — Chart references Report sheet directly — no data duplication
                     on the chart sheet. Report rows sorted by (group, cluster,
                     vault, date) so each series is a contiguous range. All-zero
                     Storage Consumed series excluded from chart. Fixed chart
                     size (18 × 28 cm) to fit a single screen.
  4.0 (2026-04-04) — Combined all protection groups onto a single "Trend Charts"
                     sheet. Pivot table (Date × Group) formatted as an Excel
                     Table with ▼ dropdown filters on every column header for
                     group/date navigation (native slicers require pivot tables
                     not supported by openpyxl). Chart sums storage across
                     vaults per group per day. One series per protection group.
  3.9 (2026-04-04) — Output changed from CSV to Excel (.xlsx) via openpyxl.
                     In --mode trend, one chart sheet is added per protection
                     group showing Storage Consumed over time. Multiple vaults
                     for the same group appear as separate series. Header row
                     styled and columns auto-fitted. Requires: pip install openpyxl.
  3.8 (2026-04-04) — Added startup note explaining zero transfer values and
                     that Storage Consumed is the primary metric. Same caveat
                     added to README.
  3.7 (2026-04-04) — Added --mode trend. In trend mode the date range is split
                     into individual calendar days (cluster timezone); each day
                     gets its own API call and rows so users can plot a daily
                     storage utilization trend per protection group. Vault IDs
                     are fetched once per cluster and reused across all days.
                     Default remains --mode summary (existing behaviour).
  3.6 (2026-04-04) — Timezone auto-detected from cluster via GET /cluster.
                     Removed --timezone flag — no longer needed. Date
                     boundaries computed in cluster local time automatically.
                     Startup banner shows detected timezone and resolved
                     timestamps for verification.
  3.5 (2026-04-04) — Added --timezone flag. Date boundaries (--start/--end/
                     --days) are now computed in the specified timezone so the
                     script matches the Helios UI which uses the cluster's local
                     time (e.g. America/Chicago for US Central). Startup banner
                     shows resolved timestamps with timezone label.
  3.4 (2026-04-04) — Added --start-msecs / --end-msecs flags to pass exact
                     millisecond timestamps from Chrome DevTools, eliminating
                     time boundary mismatches vs the UI. --debug now prints
                     the exact request URL for direct comparison with Chrome.
  3.3 (2026-04-04) — Output raw bytes — no conversion. Column headers updated
                     to (Bytes). Allows direct comparison to API payload values
                     confirmed via Chrome DevTools.
  3.2 (2026-04-04) — Unit changed from GB to TB (÷ 1,000,000,000,000). UI
                     displays binary TiB; 1 TiB = 1.09951 TB so script values
                     are ~9.95% higher than UI figures — correct decimal
                     equivalent. Column headers updated to (TB). 4 decimal
                     places to preserve precision at TB scale.
  3.1 (2026-04-04) — GB conversion corrected to decimal SI (÷ 1,000,000,000).
                     Added Period Start / Period End columns so stacked CSV
                     runs can be used to build a storage utilization trend.
  3.0 (2026-04-04) — Full rewrite. Single source of truth: dataTransferToVaults
                     report API only. Removed Helios activity API and Reporting
                     API component calls — data was inaccurate. One row per
                     protection group per vault per cluster.
  2.4 (2026-04-04) — Per-group retained storage via dataTransferToVaults; nested
                     response structure fix.
  2.0 (2026-04-04) — Helios activity API + Reporting API component 1600.
  1.0 (2026-04-04) — Initial release.

Usage:
  python3 fortknox_vault_report.py                         # prompts for key on first run
  python3 fortknox_vault_report.py --days 30               # uses stored key
  python3 fortknox_vault_report.py --mode trend --days 30  # trend mode
  python3 fortknox_vault_report.py --ca-bundle /path/to/ca.pem  # corporate proxy cert
  python3 fortknox_vault_report.py --clear-credentials     # remove stored key
"""

__version__ = "4.21"

import argparse
import getpass
import os
import sys
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import requests

HELIOS_HOST     = "helios.cohesity.com"
_KEYRING_SVC    = "cohesity_helios"
_KEYRING_USER   = "apikey"

_verify = True   # overridden in main() via --insecure / --ca-bundle

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(val):
    if isinstance(val, str) and val and val[0] in _FORMULA_PREFIXES:
        return "'" + val
    return val


# ---------------------------------------------------------------------------
# Credential store  (OS keychain via keyring)
# ---------------------------------------------------------------------------

def _keyring_available() -> bool:
    try:
        import keyring          # noqa: F401
        return True
    except ImportError:
        return False


def get_api_key(cli_key: str = None) -> str:
    """
    Return the Helios API key using this priority order:
      1. --apikey CLI argument (one-time use, not stored)
      2. Key stored in the OS keychain from a previous run
      3. Interactive prompt — key is then saved to the keychain

    The key is never written to disk in plaintext and never visible
    in the process list.
    """
    if cli_key:
        return cli_key

    if _keyring_available():
        import keyring
        stored = keyring.get_password(_KEYRING_SVC, _KEYRING_USER)
        if stored:
            print("[*] Using API key stored in system keychain.")
            return stored
        print("[*] No stored API key found.")
        key = getpass.getpass("    Enter Helios API key (will be saved to keychain): ").strip()
        if not key:
            print("ERROR: API key cannot be empty.")
            sys.exit(1)
        keyring.set_password(_KEYRING_SVC, _KEYRING_USER, key)
        print("[*] API key saved to system keychain.")
        return key
    else:
        print("    NOTE: 'keyring' package not installed — key will not be saved.")
        print("          Install with:  pip install keyring")
        key = getpass.getpass("    Enter Helios API key: ").strip()
        if not key:
            print("ERROR: API key cannot be empty.")
            sys.exit(1)
        return key


def clear_stored_credentials():
    """Remove the stored API key from the OS keychain."""
    if not _keyring_available():
        print("ERROR: 'keyring' package not installed. Nothing to clear.")
        sys.exit(1)
    import keyring
    existing = keyring.get_password(_KEYRING_SVC, _KEYRING_USER)
    if existing:
        keyring.delete_password(_KEYRING_SVC, _KEYRING_USER)
        print("[*] Stored API key removed from system keychain.")
    else:
        print("[*] No stored API key found — nothing to clear.")


def default_output_path() -> str:
    """Auto-generate output filename: <script_name>_v<version>_YYYYMMDD_HHMMSS.xlsx in cwd."""
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    return os.path.join(os.getcwd(), f"{script_name}_v{__version__}_{timestamp}.xlsx")


# ---------------------------------------------------------------------------
# Auth / cluster helpers
# ---------------------------------------------------------------------------

def helios_headers(api_key: str, cluster_id: int = None) -> dict:
    h = {
        "apiKey":       api_key,
        "Accept":       "application/json",
        "Content-Type": "application/json",
    }
    if cluster_id is not None:
        h["accessClusterId"] = str(cluster_id)
    return h


def get_helios_clusters(api_key: str) -> list:
    """Return [{name, clusterId}, ...] for all Helios-connected clusters."""
    url = f"https://{HELIOS_HOST}/mcm/clusters/connectionStatus"
    try:
        r = requests.get(url, headers=helios_headers(api_key), verify=_verify, timeout=30)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Helios. Check your network/VPN.")
        sys.exit(1)
    except requests.exceptions.HTTPError:
        print(f"ERROR: Helios auth failed ({r.status_code}). Check your API key.")
        sys.exit(1)

    clusters = r.json()
    if not isinstance(clusters, list):
        clusters = clusters.get("clusterList", [])

    result = [{"name": c.get("name", ""), "clusterId": c.get("clusterId")}
              for c in clusters]
    print(f"[+] Connected to Helios — {len(result)} cluster(s)")
    return result


def get_cluster_timezone(api_key: str, cluster_id: int, cluster_name: str) -> str:
    """Query the cluster's configured timezone via GET /irisservices/api/v1/public/cluster."""
    url = f"https://{HELIOS_HOST}/irisservices/api/v1/public/cluster"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         verify=_verify, timeout=15)
        r.raise_for_status()
        tz = r.json().get("timezone", "")
        if tz:
            return tz
    except Exception:
        pass
    print(f"    WARNING: Could not query timezone for {cluster_name}, falling back to UTC")
    return "UTC"


def get_fortknox_vaults(api_key: str, cluster_id: int) -> tuple:
    """Return (vault_ids: list, vault_names: set) for FortKnox vaults on a cluster."""
    url = f"https://{HELIOS_HOST}/irisservices/api/v1/public/vaults"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         params={"includeFortKnoxVault": "true"},
                         verify=_verify, timeout=30)
        r.raise_for_status()
        vaults = r.json() or []
        ids   = [v["id"]   for v in vaults if "id" in v]
        names = {v["name"] for v in vaults if v.get("name")}
        return ids, names
    except Exception as e:
        print(f"    WARNING: Could not fetch vault IDs: {e}")
        return [], set()


def get_protection_group_map(api_key: str, cluster_id: int, cluster_name: str) -> dict:
    """Return {group_name: group_id} for all active protection groups on a cluster."""
    url = f"https://{HELIOS_HOST}/v2/data-protect/protection-groups"
    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         params={"isActive": "true", "includeTenants": "true"},
                         verify=_verify, timeout=30)
        r.raise_for_status()
        groups = r.json().get("protectionGroups") or []
        return {g["name"]: g["id"] for g in groups if g.get("id") and g.get("name")}
    except Exception as e:
        print(f"    WARNING: Could not fetch protection groups for {cluster_name}: {e}")
        return {}


_FK_TARGET_TYPES = frozenset({
    "krpaas", "kfortknox", "kfort_knox", "krpaasarchival",  # k-prefixed (v1/mixed)
    "rpaas",  "fortknox",  "fort_knox",  "rpaasarchival",   # plain (v2)
    "cloud",                                                  # generic cloud vault
})


def get_activities_transfer(api_key: str, cluster_id: int, cluster_name: str,
                             vault_names: set, target_groups: set,
                             start_msecs: int, end_msecs: int,
                             tz, debug: bool = False) -> dict:
    """
    Fetch protection activity stats for FortKnox cloud-vault archival runs.

    Calls GET /v2/data-protect/protection-groups to list groups, then for each
    relevant group calls GET /v2/data-protect/protection-groups/{id}/runs
    filtered to the time window.  Archival results are matched to FortKnox
    vaults first by targetName (against the known vault name set), then by
    targetType as a fallback (see _FK_TARGET_TYPES).  target_groups limits
    which groups are queried (pass an empty set to query all groups).

    Backup and Vault appear as separate run entries in the Protection
    Activities API — a backup-only run has localBackupInfo but no
    archivalInfo; a vault-only run has archivalInfo but no localBackupInfo.
    This function collects each activity type independently by its own
    start date and then merges them:

      • Backup activity  → Data Read / Data Written keyed by backup start date
      • Vault activity   → Logical / Physical Transferred keyed by vault start date

    On days where both activities happen (typically most days, since the
    daily backup and the vault archival of the previous snapshot both kick
    off at roughly the same wall-clock time), all four fields populate the
    same row.  On days where only one type runs, the other pair is 0.

    target_groups limits which groups are queried (pass an empty set to
    query all groups).

    Returns:
      {(group_name, date_str): {
         "act_data_read":    int,   # Backup activity: localSnapshotStats.bytesRead
         "act_data_written": int,   # Backup activity: localSnapshotStats.bytesWritten
         "act_logical":      int,   # Vault activity: archivalTargetResults logicalBytesTransferred
         "act_physical":     int,   # Vault activity: archivalTargetResults physicalBytesTransferred
      }}
    """
    group_map = get_protection_group_map(api_key, cluster_id, cluster_name)
    if not group_map:
        return {}

    relevant = {name: gid for name, gid in group_map.items()
                if not target_groups or name in target_groups}
    if not relevant:
        return {}

    def _is_fk(ar: dict) -> bool:
        if ar.get("targetName", "") in vault_names:
            return True
        ttype = (ar.get("targetType") or ar.get("archivalTargetType") or "").lower()
        return ttype in _FK_TARGET_TYPES

    def _xfer_bytes(ar: dict, field: str) -> int:
        stats = ar.get("stats") or {}
        return int(stats.get(field) or ar.get(field) or 0)

    def _date(usecs: int) -> str:
        return datetime.fromtimestamp(usecs / 1_000_000, tz=tz).strftime("%Y-%m-%d")

    def _ensure(d, key):
        if key not in d:
            d[key] = {"act_data_read": 0, "act_data_written": 0,
                      "act_logical":   0, "act_physical":     0}

    result: dict = {}

    for group_name, group_id in relevant.items():
        url = f"https://{HELIOS_HOST}/v2/data-protect/protection-groups/{group_id}/runs"
        params = {
            "startTimeUsecs":       start_msecs * 1000,
            "endTimeUsecs":         end_msecs   * 1000,
            "numRuns":              10000,
            "includeObjectDetails": "false",
        }
        try:
            all_runs = []
            page_params = dict(params)
            while True:
                r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                                 params=page_params, verify=_verify, timeout=60)
                if debug and not all_runs:
                    print(f"\n[DEBUG] activities {cluster_name}/{group_name}:\n        {r.url}")
                r.raise_for_status()
                data = r.json()
                page_runs = data.get("runs") or []
                all_runs.extend(page_runs)
                cookie = data.get("paginationCookie")
                if not cookie or not page_runs:
                    break
                page_params = {"paginationCookie": cookie, "numRuns": params["numRuns"]}
            runs = all_runs
        except Exception as e:
            print(f"    WARNING: Activities fetch failed for {cluster_name}/{group_name}: {e}")
            continue

        if debug and runs:
            sample_ar = ((runs[0].get("archivalInfo") or {})
                         .get("archivalTargetResults") or [{}])[0]
            print(f"[DEBUG] {cluster_name}/{group_name} — {len(runs)} run(s) total (paginated); "
                  f"sample archivalTargetResult keys: {list(sample_ar.keys())}")
            if sample_ar.get("stats"):
                print(f"[DEBUG] sample stats keys: {list(sample_ar['stats'].keys())}")
            vault_pairs = {(a.get("targetName"), a.get("targetType"))
                           for run in runs
                           for a in ((run.get("archivalInfo") or {})
                                     .get("archivalTargetResults") or [])}
            print(f"[DEBUG] all (targetName, targetType) pairs: {vault_pairs}")
            print(f"[DEBUG] known vault_names set: {vault_names}")

        bk_runs = vk_runs = 0
        for run in runs:
            # --- Backup activity (Activity Type = Backup) ---
            local_info  = run.get("localBackupInfo") or {}
            local_stats = local_info.get("localSnapshotStats") or {}
            bk_usecs    = local_info.get("startTimeUsecs") or 0
            data_read    = int(local_stats.get("bytesRead",    0) or 0)
            data_written = int(local_stats.get("bytesWritten", 0) or 0)
            if bk_usecs and (data_read or data_written):
                key = (group_name, _date(bk_usecs))
                _ensure(result, key)
                result[key]["act_data_read"]    += data_read
                result[key]["act_data_written"] += data_written
                bk_runs += 1

            # --- Vault activity (Activity Type = Vault) ---
            archival_results = (run.get("archivalInfo") or {}).get("archivalTargetResults") or []
            fk_results = [a for a in archival_results if _is_fk(a)]
            for ar in fk_results:
                vk_usecs = ar.get("startTimeUsecs") or bk_usecs
                if not vk_usecs:
                    continue
                key = (group_name, _date(vk_usecs))
                _ensure(result, key)
                result[key]["act_logical"]  += _xfer_bytes(ar, "logicalBytesTransferred")
                result[key]["act_physical"] += _xfer_bytes(ar, "physicalBytesTransferred")
                vk_runs += 1

        if debug:
            hits = sum(1 for k in result if k[0] == group_name)
            print(f"[DEBUG] {cluster_name}/{group_name} — "
                  f"{bk_runs} backup run(s), {vk_runs} vault run(s), "
                  f"{hits} distinct date(s) recorded")

    return result


# ---------------------------------------------------------------------------
# Report API
# ---------------------------------------------------------------------------

def get_data_transfer_report(api_key: str, cluster_id: int, cluster_name: str,
                              vault_ids: list, start_msecs: int, end_msecs: int,
                              debug: bool = False) -> list:
    """
    Fetch "Data Transferred to External Targets" report for a cluster.

    Returns a flat list of row dicts, one per protection group per vault:
      cluster, vault_name, vault_type, protection_group, storage_consumed_bytes
    """
    if not vault_ids:
        return []

    url = f"https://{HELIOS_HOST}/irisservices/api/v1/public/reports/dataTransferToVaults"
    params = [("startTimeMsecs", start_msecs), ("endTimeMsecs", end_msecs)]
    for vid in vault_ids:
        params.append(("vaultIds", vid))

    try:
        r = requests.get(url, headers=helios_headers(api_key, cluster_id),
                         params=params, verify=_verify, timeout=60)
        if debug:
            import json
            print(f"\n[DEBUG] {cluster_name} — exact request URL:\n        {r.url}")
        r.raise_for_status()
        resp = r.json()
    except Exception as e:
        print(f"    WARNING: Report fetch failed for {cluster_name}: {e}")
        return []

    if debug:
        import json
        print(f"[DEBUG] {cluster_name} — response keys: {list(resp.keys()) if isinstance(resp, dict) else type(resp)}")
        for k, v in (resp.items() if isinstance(resp, dict) else []):
            if isinstance(v, list) and v:
                print(f"[DEBUG] {k}[0] = {json.dumps(v[0], indent=2)[:1000]}")
                break

    rows = []
    for vault_summary in (resp.get("dataTransferSummary") or []):
        vault_name = vault_summary.get("vaultName", "")
        vault_type = vault_summary.get("vaultType", "N/A")
        for job in (vault_summary.get("dataTransferPerProtectionJob") or []):
            consumed_bytes = job.get("storageConsumed", 0) or 0
            rows.append({
                "cluster":                cluster_name,
                "vault_name":             vault_name,
                "vault_type":             vault_type,
                "protection_group":       job.get("protectionJobName", ""),
                "storage_consumed_bytes": consumed_bytes,
                "storage_consumed_tb":    round(consumed_bytes / 1_000_000_000_000, 4),
                "storage_consumed_tib":   round(consumed_bytes / 1_099_511_627_776, 4),
            })
    return rows


# ---------------------------------------------------------------------------
# Size / date helpers
# ---------------------------------------------------------------------------

def fmt_bytes(byte_count) -> str:
    if not byte_count:
        return "0"
    return str(int(byte_count))


def resolve_tz(tz_name: str):
    """Return a tzinfo object for tz_name, falling back to UTC on error."""
    try:
        return ZoneInfo(tz_name)
    except (ZoneInfoNotFoundError, KeyError):
        print(f"WARNING: Unknown timezone '{tz_name}', falling back to UTC.")
        return timezone.utc


def now_msecs() -> int:
    return int(datetime.now(tz=timezone.utc).timestamp() * 1000)


def date_to_msecs(date_str: str, tz) -> int:
    """Start of day (00:00:00) in tz."""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=tz)
    except ValueError:
        print(f"ERROR: Invalid date '{date_str}'. Use YYYY-MM-DD.")
        sys.exit(1)
    return int(dt.timestamp() * 1000)


def end_of_day_msecs(date_str: str, tz) -> int:
    """End of day (23:59:59) in tz."""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, microsecond=999000, tzinfo=tz)
    except ValueError:
        print(f"ERROR: Invalid date '{date_str}'. Use YYYY-MM-DD.")
        sys.exit(1)
    return int(dt.timestamp() * 1000)


def resolve_date_range(args, tz) -> tuple:
    """Return (start_msecs, end_msecs). Default: last 7 days."""
    if args.start_msecs or args.end_msecs:
        start = int(args.start_msecs) if args.start_msecs else now_msecs() - 7 * 86400 * 1000
        end   = int(args.end_msecs)   if args.end_msecs   else now_msecs()
        return start, end
    if args.days:
        start = (datetime.now(tz=tz) - timedelta(days=args.days)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        return int(start.timestamp() * 1000), now_msecs()
    if args.start:
        return date_to_msecs(args.start, tz), \
               (end_of_day_msecs(args.end, tz) if args.end else now_msecs())
    if args.end:
        print("WARNING: --end given without --start. Ignoring.")
    start = (datetime.now(tz=tz) - timedelta(days=7)).replace(
        hour=0, minute=0, second=0, microsecond=0)
    return int(start.timestamp() * 1000), now_msecs()


# ---------------------------------------------------------------------------
# Trend helper
# ---------------------------------------------------------------------------

def iter_days(start_msecs: int, end_msecs: int, tz):
    """
    Yield (day_start_msecs, day_end_msecs, date_str) for each calendar day
    in the range [start_msecs, end_msecs], computed in tz.
    """
    # Snap to the start of the calendar day containing start_msecs
    start_dt = datetime.fromtimestamp(start_msecs / 1000, tz=tz).replace(
        hour=0, minute=0, second=0, microsecond=0)
    end_dt   = datetime.fromtimestamp(end_msecs   / 1000, tz=tz)

    current = start_dt
    while current <= end_dt:
        day_start = int(current.timestamp() * 1000)
        day_end   = int(current.replace(hour=23, minute=59, second=59).timestamp() * 1000)
        yield day_start, day_end, current.strftime("%Y-%m-%d")
        current += timedelta(days=1)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

COLUMNS = [
    ("Period Start",                          "period_start"),
    ("Period End",                            "period_end"),
    ("Cluster",                               "cluster"),
    ("Vault Name",                            "vault_name"),
    ("Vault Type",                            "vault_type"),
    ("Protection Group",                      "protection_group"),
    ("Remote Storage Consumed (Bytes)",              "storage_consumed_bytes"),
    ("Remote Storage Consumed (TB)",                 "storage_consumed_tb"),
    ("Remote Storage Consumed (TiB)",                "storage_consumed_tib"),
    # --- Protection Activities (cloud vault filter) ---
    ("Activities: Data Read (Bytes)",                "act_data_read"),
    ("Activities: Data Read (TB)",                   "act_data_read_tb"),
    ("Activities: Data Written (Bytes)",             "act_data_written"),
    ("Activities: Data Written (TB)",                "act_data_written_tb"),
    ("Activities: Logical Transferred (Bytes)",      "act_logical"),
    ("Activities: Logical Transferred (TB)",         "act_logical_tb"),
    ("Activities: Physical Transferred (Bytes)",     "act_physical"),
    ("Activities: Physical Transferred (TB)",        "act_physical_tb"),
]

# Source label for each column key — used to build the source-banner row in the Report sheet
_COLUMN_SOURCE = {
    "period_start":           "Computed",
    "period_end":             "Computed",
    "cluster":                "Helios MCM",
    "vault_name":             "dataTransferToVaults",
    "vault_type":             "dataTransferToVaults",
    "protection_group":       "dataTransferToVaults",
    "storage_consumed_bytes": "dataTransferToVaults",
    "storage_consumed_tb":    "Computed",
    "storage_consumed_tib":   "Computed",
    "act_data_read":          "Protection Activities",
    "act_data_read_tb":       "Computed",
    "act_data_written":       "Protection Activities",
    "act_data_written_tb":    "Computed",
    "act_logical":            "Protection Activities",
    "act_logical_tb":         "Computed",
    "act_physical":           "Protection Activities",
    "act_physical_tb":        "Computed",
}

# Column index (1-based) of the TB column in the Report sheet — used by chart
_TB_COL_IDX = next(i for i, (_, k) in enumerate(COLUMNS, start=1)
                   if k == "storage_consumed_tb")


def _safe_sheet_name(name: str, existing: list) -> str:
    """Sanitise and deduplicate an Excel sheet name (max 31 chars)."""
    safe = name.translate(str.maketrans("", "", r":\/?*[]"))
    safe = safe[:31]
    if safe in existing:
        safe = safe[:28] + f"_{len(existing)}"
    return safe


def _make_chart(title: str, report_ws, series_list: list,
                LineChart, Reference, SeriesLabel, Legend):
    """
    Build a single LineChart with consistent styling.

    series_list: [(label, start_row, end_row), ...]
    X-axis tick density is reduced automatically for long date ranges:
      ≤14 days  → every day
      15-60 days → every 2 days
      >60 days   → every 7 days (weekly)
    """
    chart = LineChart()
    chart.style  = 10
    chart.height = 16   # cm — room for legend below without overlap
    chart.width  = 28   # cm

    # --- Chart title: 16 pt bold; plain string fallback ---
    try:
        from openpyxl.chart.title import Title
        from openpyxl.chart.text import RichText as ChartRichText, TxChoice
        from openpyxl.drawing.text import (
            Paragraph, RegularTextRun, RichTextProperties,
            ParagraphProperties, CharacterProperties,
        )
        chart.title = Title(
            tx=TxChoice(rich=ChartRichText(
                bodyPr=RichTextProperties(),
                p=[Paragraph(
                    pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1600, b=True)),
                    r=[RegularTextRun(t=title)]
                )]
            )),
            overlay=False,
        )
    except Exception:
        chart.title = title

    # --- Y axis (value) ---
    chart.y_axis.title        = "Remote Storage Consumed (TB)"
    chart.y_axis.numFmt       = "0.000"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.tickLblPos   = "nextTo"
    chart.y_axis.delete       = False

    # --- X axis (category — string dates, no numFmt) ---
    chart.x_axis.title        = "Date"
    chart.x_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos   = "low"
    chart.x_axis.delete       = False

    # Reduce tick density for long ranges so dates don't overlap
    n_days = max((end - start + 1) for _, start, end in series_list) if series_list else 1
    if n_days > 60:
        chart.x_axis.tickLblSkip = 7    # weekly label
    elif n_days > 14:
        chart.x_axis.tickLblSkip = 2    # every-other-day label
    # else: ≤14 days — show every date, leave tickLblSkip at default

    # --- Legend at bottom, no overlap ---
    legend = Legend()
    legend.position = "b"
    legend.overlay  = False
    chart.legend = legend

    # Build all series first, then assign string categories to each one.
    # openpyxl's set_categories() generates <c:numRef> in the XML — Excel
    # treats string date values as invalid numbers and shows nothing.
    # Using StrRef forces <c:strRef> so Excel renders them as text labels.
    try:
        from openpyxl.chart.data_source import DataSource, StrRef
        from openpyxl.utils import get_column_letter as _gcl
        _use_strref = True
    except ImportError:
        _use_strref = False

    cat_formula = None
    for label, start_row, end_row in series_list:
        data_ref = Reference(report_ws, min_col=_TB_COL_IDX,
                             min_row=start_row, max_row=end_row)
        chart.add_data(data_ref)
        chart.series[-1].title = SeriesLabel(v=label)

        # Capture category formula from the first (longest) series
        if cat_formula is None:
            col_a = _gcl(1) if _use_strref else "A"
            cat_formula = (f"'{report_ws.title}'!${col_a}${start_row}"
                           f":${col_a}${end_row}")

    # Assign string categories to every series
    if cat_formula:
        if _use_strref:
            str_cat = DataSource(strRef=StrRef(f=cat_formula))
            for s in chart.series:
                s.cat = str_cat
        else:
            # Fallback: let openpyxl use numRef (dates may not display)
            chart.set_categories(Reference(report_ws, min_col=1,
                                           min_row=series_list[0][1],
                                           max_row=series_list[0][2]))

    return chart


def _add_trend_chart(wb, group_ranges: dict):
    """
    Add one chart sheet per cluster, each showing all protection groups
    for that cluster as separate series.

    References the Report sheet directly — no data is duplicated.
    All-zero series are excluded. Sheet names are truncated to 31 chars.

    group_ranges: {(cluster, group, vault): {"start": int, "end": int,
                                              "any_nonzero": bool}}
    """
    try:
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.chart.legend import Legend
    except ImportError:
        print("WARNING: openpyxl chart module unavailable — chart skipped.")
        return

    report_ws = wb["Report"]

    # Group series by cluster
    clusters_seen = sorted({k[0] for k in group_ranges})
    sheet_names_used = [s.title for s in wb.worksheets]
    total_charts = 0

    for cluster in clusters_seen:
        # Collect non-zero series for this cluster, sorted by group name
        series_list = []
        for (c, group, vault), info in sorted(group_ranges.items(),
                                              key=lambda x: (x[0][1], x[0][2])):
            if c != cluster or not info["any_nonzero"]:
                continue
            vault_count = sum(1 for k in group_ranges
                              if k[0] == cluster and k[1] == group)
            label = group if vault_count == 1 else f"{group} [{vault}]"
            series_list.append((label, info["start"], info["end"]))

        if not series_list:
            print(f"    [{cluster}] all-zero — chart skipped")
            continue

        title = f"FortKnox \u2014 {cluster}"
        chart = _make_chart(title, report_ws, series_list,
                            LineChart, Reference, SeriesLabel, Legend)

        sheet_name = _safe_sheet_name(cluster, sheet_names_used)
        sheet_names_used.append(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        ws.add_chart(chart, "A1")
        total_charts += 1
        print(f"    Chart sheet '{sheet_name}': {len(series_list)} series")

    if total_charts == 0:
        print("    No non-zero Storage Consumed data — all charts skipped.")
    else:
        print(f"    {total_charts} chart sheet(s) added (one per cluster)")


_FIELD_REF = [
    # (Column Name, Description, Source field path, API endpoint)
    ("Period Start",
     "Start of the reporting window",
     "Computed from --start / --days / --start-msecs",
     "—"),
    ("Period End",
     "End of the reporting window",
     "Computed from --end / --days / --end-msecs",
     "—"),
    ("Cluster",
     "Name of the Cohesity cluster",
     "connectionStatus[].name",
     "GET /mcm/clusters/connectionStatus"),
    ("Vault Name",
     "Name of the FortKnox vault",
     "dataTransferSummary[].vaultName",
     "GET /public/reports/dataTransferToVaults"),
    ("Vault Type",
     "Vault type identifier (e.g. kFortKnox, kRPaaS)",
     "dataTransferSummary[].vaultType",
     "GET /public/reports/dataTransferToVaults"),
    ("Protection Group",
     "Name of the protection job/group",
     "dataTransferSummary[].dataTransferPerProtectionJob[].protectionJobName",
     "GET /public/reports/dataTransferToVaults"),
    ("Remote Storage Consumed (Bytes)",
     "Total retained bytes in the remote (FortKnox) vault for this protection "
     "group across all snapshots (cumulative — not scoped to the query window). "
     "Matches Helios UI.",
     "dataTransferSummary[].dataTransferPerProtectionJob[].storageConsumed",
     "GET /public/reports/dataTransferToVaults"),
    ("Remote Storage Consumed (TB)",
     "Remote Storage Consumed ÷ 1,000,000,000,000  (decimal terabytes, 4 dp)",
     "Computed",
     "—"),
    ("Remote Storage Consumed (TiB)",
     "Remote Storage Consumed ÷ 1,099,511,627,776  (binary tebibytes, 4 dp). "
     "Matches the unit shown in the Helios UI.",
     "Computed",
     "—"),
    # --- Activities columns ---
    ("Activities: Data Read (Bytes)",
     "Bytes read from the backup source during the Backup activity, keyed by "
     "the backup start date. Collected independently of vault activity.",
     "runs[].localBackupInfo.localSnapshotStats.bytesRead",
     "GET /v2/data-protect/protection-groups/{id}/runs"),
    ("Activities: Data Read (TB)",
     "Activities: Data Read ÷ 1,000,000,000,000  (decimal terabytes, 4 dp)",
     "Computed",
     "—"),
    ("Activities: Data Written (Bytes)",
     "Bytes written to local storage during the Backup activity, keyed by the "
     "backup start date.",
     "runs[].localBackupInfo.localSnapshotStats.bytesWritten",
     "GET /v2/data-protect/protection-groups/{id}/runs"),
    ("Activities: Data Written (TB)",
     "Activities: Data Written ÷ 1,000,000,000,000  (decimal terabytes, 4 dp)",
     "Computed",
     "—"),
    ("Activities: Logical Transferred (Bytes)",
     "Logical bytes transferred to the FortKnox cloud vault target across all "
     "qualifying archival copy runs. Summed from per-run archival target stats.",
     "runs[].archivalInfo.archivalTargetResults[].stats.logicalBytesTransferred",
     "GET /v2/data-protect/protection-groups/{id}/runs"),
    ("Activities: Logical Transferred (TB)",
     "Activities: Logical Transferred ÷ 1,000,000,000,000  (decimal terabytes, 4 dp)",
     "Computed",
     "—"),
    ("Activities: Physical Transferred (Bytes)",
     "Physical (post-dedup/compression) bytes transferred to the FortKnox "
     "cloud vault target. Summed from per-run archival target stats.",
     "runs[].archivalInfo.archivalTargetResults[].stats.physicalBytesTransferred",
     "GET /v2/data-protect/protection-groups/{id}/runs"),
    ("Activities: Physical Transferred (TB)",
     "Activities: Physical Transferred ÷ 1,000,000,000,000  (decimal terabytes, 4 dp)",
     "Computed",
     "—"),
]

_API_REF = [
    # (Type, Endpoint, Routing, Purpose)
    ("Helios MCM",
     "GET /mcm/clusters/connectionStatus",
     "Direct — helios.cohesity.com",
     "Lists all Helios-connected clusters with their IDs and names"),
    ("Cluster v1 public",
     "GET /irisservices/api/v1/public/cluster",
     "Via Helios proxy  (accessClusterId header)",
     "Reads the cluster's configured timezone for date boundary calculation"),
    ("Cluster v1 public",
     "GET /irisservices/api/v1/public/vaults?includeFortKnoxVault=true",
     "Via Helios proxy  (accessClusterId header)",
     "Retrieves FortKnox vault IDs for the cluster"),
    ("Cluster v1 public",
     "GET /irisservices/api/v1/public/reports/dataTransferToVaults",
     "Via Helios proxy  (accessClusterId header)",
     "Storage consumed and vault membership per protection group (cumulative)"),
    ("Cluster v2",
     "GET /v2/data-protect/protection-groups",
     "Via Helios proxy  (accessClusterId header)",
     "Lists all active protection groups to obtain group IDs for the activities fetch"),
    ("Cluster v2",
     "GET /v2/data-protect/protection-groups/{id}/runs",
     "Via Helios proxy  (accessClusterId header)",
     "Per-run backup and archival stats for the Activities columns; filtered to runs "
     "with FortKnox cloud vault archival results (matched by vault name)"),
]


def _sheet_about(wb, meta: dict):
    """Insert an 'About' sheet at position 0 with run info, field reference, and API reference."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title="About", index=0)

    GREEN = "70AD47"
    WHITE = "FFFFFF"

    _fill = PatternFill(fill_type="solid", fgColor=GREEN)

    def _title_row(text, ncols=4):
        nonlocal _row
        ws.cell(_row, 1, text).font = Font(bold=True, size=13, color=WHITE)
        ws.cell(_row, 1).fill      = _fill
        ws.cell(_row, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[_row].height = 20
        if ncols > 1:
            ws.merge_cells(f"A{_row}:{get_column_letter(ncols)}{_row}")
        _row += 1

    def _section(text):
        nonlocal _row
        _row += 1
        ws.cell(_row, 1, text).font = Font(bold=True, size=11, color=WHITE)
        ws.cell(_row, 1).fill      = _fill
        ws.merge_cells(f"A{_row}:D{_row}")
        ws.row_dimensions[_row].height = 18
        _row += 1

    def _header_row(labels):
        nonlocal _row
        for col, h in enumerate(labels, start=1):
            c = ws.cell(_row, col, h)
            c.font      = Font(bold=True, color=WHITE)
            c.fill      = _fill
            c.alignment = Alignment(horizontal="center")
        _row += 1

    def _kv(key, val):
        nonlocal _row
        ws.cell(_row, 1, key).font = Font(bold=True)
        ws.cell(_row, 2, val)
        _row += 1

    _row = 1

    _title_row(f"FortKnox Vault Report  v{__version__}  —  Run Information", ncols=4)

    _section("Command")
    ws.cell(_row, 1, meta["command"])
    ws.merge_cells(f"A{_row}:D{_row}")
    _row += 1

    _section("Parameters")
    _kv("Script Version",  __version__)
    _kv("Generated",       meta["generated"])
    _kv("Mode",            meta["mode"])
    _kv("Period Start",    meta["period_start"])
    _kv("Period End",      meta["period_end"])
    _kv("Cluster Filter",  meta["cluster_filter"])
    _kv("Vault Filter",    meta["vault_filter"])
    _kv("Cluster Timezone", meta["tz_name"])

    _section("Report Fields")
    _header_row(["Column Name", "Description", "Source Field Path", "API Endpoint"])
    for col_name, desc, source, api in _FIELD_REF:
        ws.cell(_row, 1, col_name)
        ws.cell(_row, 2, desc)
        ws.cell(_row, 3, source)
        ws.cell(_row, 4, api)
        _row += 1

    _section("APIs Used")
    _header_row(["Type", "Endpoint", "Routing", "Purpose"])
    for api_type, endpoint, routing, purpose in _API_REF:
        ws.cell(_row, 1, api_type)
        ws.cell(_row, 2, endpoint)
        ws.cell(_row, 3, routing)
        ws.cell(_row, 4, purpose)
        _row += 1

    # Wrap text and auto-fit all four columns
    for col_idx in range(1, 5):
        col_max = 0
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(r, col_idx)
            if cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                col_max = max(col_max, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(col_max + 2, 60)


def write_excel(rows: list, output_file: str, mode: str, meta: dict = None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required.  Install with:  pip install openpyxl")
        sys.exit(1)

    if not rows:
        print("No FortKnox data found for the requested scope.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    if meta:
        _sheet_about(wb, meta)

    cohesity_fill = PatternFill(fill_type="solid", fgColor="70AD47")
    header_font   = Font(bold=True, color="FFFFFF")

    # --- Row 1: source-banner row (merged cells per consecutive same-source group) ---
    source_labels = [_COLUMN_SOURCE.get(key, "Computed") for _, key in COLUMNS]
    # Build (start_col, end_col, label) spans for consecutive same-source runs
    spans = []
    start_col = 1
    for i, label in enumerate(source_labels):
        col = i + 1
        if col == len(source_labels) or source_labels[i + 1] != label:
            spans.append((start_col, col, label))
            start_col = col + 1

    for span_start, span_end, label in spans:
        cell = ws.cell(1, span_start, label)
        cell.font      = header_font
        cell.fill      = cohesity_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if span_end > span_start:
            ws.merge_cells(
                start_row=1, start_column=span_start,
                end_row=1,   end_column=span_end
            )
    ws.row_dimensions[1].height = 18

    # --- Row 2: column headers ---
    for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(2, col_idx, col_name)
        cell.font      = header_font
        cell.fill      = cohesity_fill
        cell.alignment = Alignment(horizontal="center")

    # Sort by (group, cluster, vault, date) so each series occupies contiguous rows
    # — required for the chart to reference Report sheet ranges directly.
    sorted_rows = sorted(rows, key=lambda r: (
        r["protection_group"], r["cluster"], r["vault_name"], r["period_start"]
    ))

    # Track Excel row range per (cluster, group, vault): data starts at row 3
    group_ranges: dict = {}
    for excel_row, r in enumerate(sorted_rows, start=3):
        key = (r["cluster"], r["protection_group"], r["vault_name"])
        consumed = int(r.get("storage_consumed_bytes") or 0)
        if key not in group_ranges:
            group_ranges[key] = {"start": excel_row, "end": excel_row,
                                 "any_nonzero": consumed > 0}
        else:
            group_ranges[key]["end"] = excel_row
            if consumed > 0:
                group_ranges[key]["any_nonzero"] = True

    # Data rows — byte columns written as integers so Excel can sort/chart them
    for r in sorted_rows:
        ws.append([
            int(r[key]) if key.endswith("_bytes") else _safe_cell(r[key]) if isinstance(r[key], str) else r[key]
            for _, key in COLUMNS
        ])

    # Auto-fit column widths
    for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(col_name),
            max((len(str(ws.cell(r, col_idx).value or "")) for r in range(3, ws.max_row + 1)),
                default=0)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    # Freeze both banner and header rows
    ws.freeze_panes = "A3"

    # Trend chart (trend mode only) — references Report sheet, no data duplication
    if mode == "trend":
        _add_trend_chart(wb, group_ranges)

    wb.save(output_file)
    print(f"\n[+] Report saved to: {output_file}  ({len(rows)} rows)")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description=f"Cohesity FortKnox data transfer report  v{__version__}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  First run (prompts for API key, saves to keychain):
    python3 fortknox_vault_report.py

  Subsequent runs (key retrieved from keychain automatically):
    python3 fortknox_vault_report.py --days 30
    python3 fortknox_vault_report.py --start 2026-03-01 --end 2026-04-01 --mode trend
    python3 fortknox_vault_report.py --cluster <name>

  One-time override (key not saved):
    python3 fortknox_vault_report.py --apikey <key>

  Clear stored credentials:
    python3 fortknox_vault_report.py --clear-credentials
        """
    )
    parser.add_argument("--apikey",  default=None,
                        help="Helios API key (optional — if omitted the key stored in the "
                             "system keychain is used, or you will be prompted)")
    parser.add_argument("--clear-credentials", action="store_true", dest="clear_creds",
                        help="Remove the stored API key from the system keychain and exit")
    parser.add_argument("--cluster", help="Filter to a specific cluster (partial name match)")
    parser.add_argument("--vault",   help="Filter to a specific vault (partial name match)")
    parser.add_argument("--output",  default=None,
                        help="Output Excel filename (default: <script_name>_YYYYMMDD_HHMMSS.xlsx "
                             "in the current directory)")
    parser.add_argument("--mode",    choices=["summary", "trend"], default="summary",
                        help="summary (default): one row per protection group covering the full "
                             "date range. trend: one row per protection group per day, enabling "
                             "storage utilization trend lines over time.")
    parser.add_argument("--debug",   action="store_true",
                        help="Print raw API response samples")
    parser.add_argument("--ca-bundle", dest="ca_bundle", default=None, metavar="PATH",
                        help="Path to CA bundle for TLS verification (e.g. corporate proxy cert)")
    parser.add_argument("--insecure", action="store_true",
                        help="Disable TLS certificate verification (NOT recommended)")

    dg = parser.add_argument_group("Date range (default: last 7 days)")
    dg.add_argument("--days",        type=int, metavar="N",        help="Last N days")
    dg.add_argument("--start",       metavar="YYYY-MM-DD",         help="Start date (inclusive)")
    dg.add_argument("--end",         metavar="YYYY-MM-DD",         help="End date (inclusive, defaults to today)")
    dg.add_argument("--start-msecs", metavar="MS", dest="start_msecs",
                    help="Exact start timestamp in milliseconds (paste from Chrome DevTools URL)")
    dg.add_argument("--end-msecs",   metavar="MS", dest="end_msecs",
                    help="Exact end timestamp in milliseconds (paste from Chrome DevTools URL)")
    return parser.parse_args()


def main():
    args = parse_args()

    global _verify
    if args.insecure:
        _verify = False
        print("=" * 65)
        print("  WARN: --insecure — TLS certificate validation DISABLED.")
        print("        Credentials and tokens may be intercepted (MITM).")
        print("=" * 65)
        try:
            import urllib3 as _u3
            _u3.disable_warnings(_u3.exceptions.InsecureRequestWarning)
        except ImportError:
            requests.packages.urllib3.disable_warnings()
    elif args.ca_bundle:
        _verify = args.ca_bundle

    # Handle --clear-credentials before doing anything else
    if args.clear_creds:
        clear_stored_credentials()
        sys.exit(0)

    api_key     = get_api_key(args.apikey)
    output_path = args.output or default_output_path()

    print(f"\n[*] FortKnox data transfer report  v{__version__}")
    print(f"[*] Output: {output_path}")
    if args.cluster:
        print(f"[*] Cluster filter: '{args.cluster}'")
    if args.vault:
        print(f"[*] Vault filter:   '{args.vault}'")

    clusters = get_helios_clusters(api_key)

    # Pick the reference cluster for timezone detection:
    # use the first cluster matching --cluster filter, or the first overall.
    ref_clusters = [c for c in clusters
                    if not args.cluster or args.cluster.lower() in c["name"].lower()]
    if not ref_clusters:
        print("ERROR: No clusters match the given --cluster filter.")
        sys.exit(1)

    ref = ref_clusters[0]
    print(f"[*] Querying timezone from '{ref['name']}'...")
    tz_name = get_cluster_timezone(api_key, ref["clusterId"], ref["name"])
    tz      = resolve_tz(tz_name)
    print(f"[*] Cluster timezone: {tz_name}")

    start_msecs, end_msecs = resolve_date_range(args, tz)
    s = datetime.fromtimestamp(start_msecs / 1000, tz=tz).strftime("%Y-%m-%d %H:%M %Z")
    e = datetime.fromtimestamp(end_msecs   / 1000, tz=tz).strftime("%Y-%m-%d %H:%M %Z")
    print(f"[*] Date range: {s} → {e}")
    print(f"[*] Mode: {args.mode}")
    print()

    # Filter cluster list once
    target_clusters = [c for c in clusters
                       if not args.cluster or args.cluster.lower() in c["name"].lower()]

    # Fetch vault IDs and names once per cluster (reused across all days in trend mode)
    cluster_vault_ids = {}
    for c in target_clusters:
        cname = c["name"]
        print(f"  [{cname}] fetching vault IDs...")
        vault_ids, vault_names = get_fortknox_vaults(api_key, c["clusterId"])
        if not vault_ids:
            print(f"  [{cname}] no FortKnox vaults — skipping")
        else:
            cluster_vault_ids[cname] = (c["clusterId"], vault_ids, vault_names)

    if not cluster_vault_ids:
        print("No clusters with FortKnox vaults found.")
        sys.exit(0)

    all_rows = []

    def _zero_act(r: dict):
        r["act_data_read"]       = 0
        r["act_data_read_tb"]    = 0.0
        r["act_data_written"]    = 0
        r["act_data_written_tb"] = 0.0
        r["act_logical"]         = 0
        r["act_logical_tb"]      = 0.0
        r["act_physical"]        = 0
        r["act_physical_tb"]     = 0.0

    if args.mode == "summary":
        for cname, (cid, vault_ids, vault_names) in cluster_vault_ids.items():
            print(f"  [{cname}] fetching summary report (storage consumed)...")
            rows = get_data_transfer_report(api_key, cid, cname,
                                            vault_ids, start_msecs, end_msecs,
                                            debug=args.debug)
            if args.vault:
                rows = [r for r in rows if args.vault.lower() in r["vault_name"].lower()]

            target_groups = {r["protection_group"] for r in rows}
            print(f"  [{cname}] fetching protection activities (cloud vault filter)...")
            activities = get_activities_transfer(
                api_key, cid, cname, vault_names, target_groups,
                start_msecs, end_msecs, tz, debug=args.debug)

            # Aggregate activities to per-group totals across all days
            act_by_group: dict = {}
            for (grp, _date), vals in activities.items():
                if grp not in act_by_group:
                    act_by_group[grp] = {"act_data_read": 0, "act_data_written": 0,
                                         "act_logical":   0, "act_physical":     0}
                for k, v in vals.items():
                    act_by_group[grp][k] += v

            for r in rows:
                r["period_start"] = s
                r["period_end"]   = e
                _zero_act(r)
                act = act_by_group.get(r["protection_group"], {})
                for k in ("act_data_read", "act_data_written", "act_logical", "act_physical"):
                    r[k] = act.get(k, 0)
                r["act_data_read_tb"]    = round(r["act_data_read"]    / 1e12, 4)
                r["act_data_written_tb"] = round(r["act_data_written"] / 1e12, 4)
                r["act_logical_tb"]      = round(r["act_logical"]      / 1e12, 4)
                r["act_physical_tb"]     = round(r["act_physical"]     / 1e12, 4)
            print(f"  [{cname}] {len(rows)} row(s)")
            all_rows.extend(rows)

    else:  # trend
        days = list(iter_days(start_msecs, end_msecs, tz))
        print(f"[*] Trend mode: {len(days)} day(s) × {len(cluster_vault_ids)} cluster(s)")
        for cname, (cid, vault_ids, vault_names) in cluster_vault_ids.items():
            print(f"  [{cname}] fetching protection activities (cloud vault filter)...")
            activities = get_activities_transfer(
                api_key, cid, cname, vault_names, set(),
                start_msecs, end_msecs, tz, debug=args.debug)

            print(f"  [{cname}] fetching daily storage consumed...")
            for day_start, day_end, date_str in days:
                rows = get_data_transfer_report(api_key, cid, cname,
                                                vault_ids, day_start, day_end,
                                                debug=args.debug)
                if args.vault:
                    rows = [r for r in rows if args.vault.lower() in r["vault_name"].lower()]
                for r in rows:
                    r["period_start"] = date_str
                    r["period_end"]   = date_str
                    _zero_act(r)
                    act = activities.get((r["protection_group"], date_str), {})
                    for k in ("act_data_read", "act_data_written", "act_logical", "act_physical"):
                        r[k] = act.get(k, 0)
                    r["act_data_read_tb"]    = round(r["act_data_read"]    / 1e12, 4)
                    r["act_data_written_tb"] = round(r["act_data_written"] / 1e12, 4)
                    r["act_logical_tb"]      = round(r["act_logical"]      / 1e12, 4)
                    r["act_physical_tb"]     = round(r["act_physical"]     / 1e12, 4)
                print(f"    {date_str}: {len(rows)} row(s)")
                all_rows.extend(rows)

    meta = {
        "command":        " ".join(sys.argv),
        "mode":           args.mode,
        "period_start":   s,
        "period_end":     e,
        "cluster_filter": args.cluster or "(all)",
        "vault_filter":   args.vault   or "(all)",
        "tz_name":        tz_name,
        "generated":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    write_excel(all_rows, output_path, args.mode, meta)


if __name__ == "__main__":
    main()
